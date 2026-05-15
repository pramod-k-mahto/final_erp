import json
import difflib

from fastapi import APIRouter, Depends, HTTPException, Request, File, UploadFile
from fastapi.responses import StreamingResponse
import io
from sqlalchemy.orm import Session, aliased, selectinload
from sqlalchemy import func
from datetime import datetime, date

from .. import models, schemas
from ..auth import get_current_user
from ..database import get_db
from ..voucher_service import get_next_voucher_number
from ..stock_service import StockValuationService
from ..dependencies import get_company_secure, validate_transaction_date


router = APIRouter(prefix="/companies/{company_id}", tags=["purchases"])


def _resolve_import_line_rate(
    purchase_type: str,
    line_rate: float,
    foreign_currency_rate: float | None,
    exchange_rate: float | None,
) -> tuple[float, float | None, float | None]:
    """Return (resolved_rate, stored_fc_rate, stored_exchange_rate).

    For IMPORT bills:
    - Both FC fields present and > 0: derive rate = fc_rate * exchange_rate.
    - Exactly one FC field: raise 400 (ambiguous).
    - Neither: use the submitted line_rate unchanged.
    For LOCAL bills: always pass through line_rate; zero-out the FC fields.
    """
    if purchase_type != "IMPORT":
        return line_rate, None, None

    has_fc = foreign_currency_rate is not None and foreign_currency_rate > 0
    has_ex = exchange_rate is not None and exchange_rate > 0

    if has_fc and has_ex:
        derived = round(foreign_currency_rate * exchange_rate, 2)
        return derived, foreign_currency_rate, exchange_rate

    if has_fc != has_ex:
        raise HTTPException(
            status_code=400,
            detail=(
                "For IMPORT purchase, both 'foreign_currency_rate' and 'exchange_rate' "
                "must be provided together, or neither."
            ),
        )

    # Neither provided – keep submitted rate, store NULLs
    return line_rate, None, None


@router.get("/hs-codes/{item_id}")
def get_item_hs_codes(
    company_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Returns a list of unique HS codes previously used for this item across item master, purchase bills, and sales invoices.
    """
    hs_codes = set()
    
    item = db.query(models.Item).filter(models.Item.id == item_id, models.Item.company_id == company_id).first()
    if item and item.hsn_sac_code:
        hs_codes.add(item.hsn_sac_code.strip())

    purchase_results = (
        db.query(models.PurchaseBillLine.hs_code)
        .join(models.PurchaseBill)
        .filter(
            models.PurchaseBill.company_id == company_id,
            models.PurchaseBillLine.item_id == item_id,
            models.PurchaseBillLine.hs_code.isnot(None),
            models.PurchaseBillLine.hs_code != "",
        )
        .distinct()
        .all()
    )
    for r in purchase_results:
        if r[0]:
            hs_codes.add(r[0].strip())

    sales_results = (
        db.query(models.SalesInvoiceLine.hs_code)
        .join(models.SalesInvoice)
        .filter(
            models.SalesInvoice.company_id == company_id,
            models.SalesInvoiceLine.item_id == item_id,
            models.SalesInvoiceLine.hs_code.isnot(None),
            models.SalesInvoiceLine.hs_code != "",
        )
        .distinct()
        .all()
    )
    for r in sales_results:
        if r[0]:
            hs_codes.add(r[0].strip())

    return list(hs_codes)



def _simulate_qty_after_removing_source(
    *,
    starting_qty: float,
    movements: list[tuple[datetime, int, float, str, int]],
    remove_source_type: str,
    remove_source_id: int,
) -> tuple[bool, float]:
    qty = float(starting_qty)
    for _, _, qty_delta, source_type, source_id in movements:
        if source_type == remove_source_type and int(source_id) == int(remove_source_id):
            continue
        qty += float(qty_delta)
        if qty < -1e-9:
            return False, qty
    return True, qty


def _ensure_purchase_bill_can_be_deleted(
    *,
    db: Session,
    company_id: int,
    bill: models.PurchaseBill,
    impacted_pairs: set[tuple[int, int]],
) -> None:
    StockLedger = models.StockLedger
    Item = models.Item

    bill_posted_at = datetime.combine(bill.date, datetime.min.time())

    item_ids = {item_id for item_id, _ in impacted_pairs}
    items = (
        db.query(Item.id, Item.opening_stock, Item.allow_negative_stock)
        .filter(Item.company_id == company_id, Item.id.in_(item_ids))
        .all()
    )
    item_map = {int(r.id): r for r in items}

    for item_id, warehouse_id in impacted_pairs:
        item_row = item_map.get(int(item_id))
        if item_row is None:
            continue
        if bool(getattr(item_row, "allow_negative_stock", False)):
            continue

        opening = float(getattr(item_row, "opening_stock", 0) or 0)

        qty_before = (
            db.query(func.coalesce(func.sum(StockLedger.qty_delta), 0))
            .filter(
                StockLedger.company_id == company_id,
                StockLedger.item_id == item_id,
                StockLedger.warehouse_id == warehouse_id,
                StockLedger.reversed_at.is_(None),
                StockLedger.posted_at < bill_posted_at,
            )
            .scalar()
        )
        starting_qty = opening + float(qty_before or 0)

        movement_rows = (
            db.query(
                StockLedger.posted_at,
                StockLedger.id,
                StockLedger.qty_delta,
                StockLedger.source_type,
                StockLedger.source_id,
            )
            .filter(
                StockLedger.company_id == company_id,
                StockLedger.item_id == item_id,
                StockLedger.warehouse_id == warehouse_id,
                StockLedger.reversed_at.is_(None),
                StockLedger.posted_at >= bill_posted_at,
            )
            .order_by(StockLedger.posted_at.asc(), StockLedger.id.asc())
            .all()
        )
        movements = [
            (
                r.posted_at,
                int(r.id),
                float(r.qty_delta),
                str(r.source_type),
                int(r.source_id),
            )
            for r in movement_rows
        ]

        ok, _ = _simulate_qty_after_removing_source(
            starting_qty=starting_qty,
            movements=movements,
            remove_source_type="PURCHASE_BILL",
            remove_source_id=bill.id,
        )
        if not ok:
            raise HTTPException(
                status_code=409,
                detail="Cannot delete bill: inventory from this bill has already been consumed; use reverse/return instead.",
            )


def _delete_purchase_bill_internal(
    *,
    db: Session,
    company_id: int,
    bill_id: int,
    actor_user_id: int | None,
    skip_consumption_check: bool = False,
) -> dict:
    bill = (
        db.query(models.PurchaseBill)
        .filter(
            models.PurchaseBill.id == bill_id,
            models.PurchaseBill.company_id == company_id,
        )
        .with_for_update()
        .first()
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    bill_lines = (
        db.query(models.PurchaseBillLine)
        .filter(models.PurchaseBillLine.bill_id == bill.id)
        .all()
    )

    # Batch-fetch items for lines that have a warehouse (avoids N+1)
    warehoused_item_ids = list({int(line.item_id) for line in bill_lines if line.warehouse_id is not None})
    item_map: dict[int, models.Item] = {}
    if warehoused_item_ids:
        rows = (
            db.query(models.Item)
            .filter(models.Item.id.in_(warehoused_item_ids), models.Item.company_id == company_id)
            .all()
        )
        item_map = {int(i.id): i for i in rows}

    impacted_pairs: set[tuple[int, int]] = set()
    for line in bill_lines:
        if line.warehouse_id is None:
            continue
        item = item_map.get(int(line.item_id))
        if item is None:
            continue
        if bool(getattr(item, "allow_negative_stock", False)):
            continue
        impacted_pairs.add((int(line.item_id), int(line.warehouse_id)))

    if impacted_pairs and not skip_consumption_check:
        _ensure_purchase_bill_can_be_deleted(
            db=db,
            company_id=company_id,
            bill=bill,
            impacted_pairs=impacted_pairs,
        )

    linked_order = (
        db.query(models.PurchaseOrder)
        .filter(
            models.PurchaseOrder.company_id == company_id,
            models.PurchaseOrder.converted_to_bill_id == bill.id,
        )
        .first()
    )

    if linked_order is not None:
        linked_order.converted_to_bill_id = None
        db.flush()

    linked_returns = (
        db.query(models.PurchaseReturn)
        .filter(
            models.PurchaseReturn.company_id == company_id,
            models.PurchaseReturn.source_bill_id == bill.id,
        )
        .all()
    )
    if linked_returns:
        for r in linked_returns:
            r.source_bill_id = None
        db.flush()

    _reverse_stock_ledger(
        db=db,
        company_id=company_id,
        source_type="PURCHASE_BILL",
        source_id=bill.id,
        created_by=actor_user_id,
    )

    db.query(models.StockMovement).filter(
        models.StockMovement.company_id == company_id,
        models.StockMovement.source_type == "PURCHASE_BILL",
        models.StockMovement.source_id == bill.id,
    ).delete()

    # Also remove inventory batches for this bill (FIFO/valuation)
    tenant_id = db.query(models.Company.tenant_id).filter(models.Company.id == company_id).scalar()
    if tenant_id:
        db.query(models.StockBatch).filter(
            models.StockBatch.tenant_id == tenant_id,
            models.StockBatch.ref_type == "PURCHASE",
            models.StockBatch.ref_id == bill.id,
        ).delete()

    if bill.voucher_id is not None:
        voucher = (
            db.query(models.Voucher)
            .filter(
                models.Voucher.id == bill.voucher_id,
                models.Voucher.company_id == company_id,
            )
            .first()
        )
        bill.voucher_id = None
        db.flush()
        if voucher is not None:
            db.delete(voucher)
            db.flush()

    db.delete(bill)
    db.flush()

    return {
        "success": True,
        "deletedId": bill_id,
        "impacted": [
            {"item_id": item_id, "warehouse_id": warehouse_id}
            for item_id, warehouse_id in sorted(impacted_pairs)
        ],
    }


def _reverse_stock_ledger(
    *,
    db: Session,
    company_id: int,
    source_type: str,
    source_id: int,
    created_by: int | None,
) -> None:
    StockLedger = models.StockLedger
    now = datetime.utcnow()
    rows = (
        db.query(StockLedger)
        .filter(
            StockLedger.company_id == company_id,
            StockLedger.source_type == source_type,
            StockLedger.source_id == source_id,
            StockLedger.reversed_at.is_(None),
        )
        .all()
    )
    if not rows:
        return

    for r in rows:
        r.reversed_at = now
        db.add(
            StockLedger(
                company_id=r.company_id,
                warehouse_id=r.warehouse_id,
                item_id=r.item_id,
                qty_delta=-float(r.qty_delta),
                unit_cost=r.unit_cost,
                source_type=r.source_type,
                source_id=r.source_id,
                source_line_id=r.source_line_id,
                posted_at=now,
                reversal_of_ledger_id=r.id,
                created_by=created_by,
            )
        )


def _get_company(db: Session, company_id: int, user: models.User) -> models.Company:
    return get_company_secure(db, company_id, user)


def _get_default_stock_ledger_id(db: Session, *, company_id: int) -> int | None:
    Ledger = models.Ledger
    LedgerGroup = models.LedgerGroup

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "CLOSING_STOCK",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "OPENING_STOCK",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    stock_group = (
        db.query(LedgerGroup)
        .filter(
            LedgerGroup.company_id == company_id,
            LedgerGroup.name == "Stock-in-Hand",
        )
        .first()
    )
    if stock_group is None:
        # If the default chart wasn't seeded for this company, create the
        # required group so inventory purchases can still post to an asset
        # ledger instead of falling back to an expense ledger.
        stock_group = LedgerGroup(
            company_id=company_id,
            name="Stock-in-Hand",
            group_type=models.LedgerGroupType.ASSET,
            parent_group_id=None,
        )
        db.add(stock_group)
        db.flush()

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.group_id == stock_group.id,
            Ledger.name.in_(["Closing Stock", "Inventory", "Stock"]),
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = Ledger(
        company_id=company_id,
        group_id=stock_group.id,
        name="Closing Stock",
        code="CLOSING_STOCK",
        opening_balance=0,
        opening_balance_type=models.OpeningBalanceType.DEBIT,
        is_active=True,
    )
    db.add(ledger)
    db.flush()
    return ledger.id


def _get_default_purchase_asset_ledger_id(db: Session, *, company_id: int) -> int | None:
    Ledger = models.Ledger
    LedgerGroup = models.LedgerGroup

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "PURCHASES_ASSET",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    stock_group = (
        db.query(LedgerGroup)
        .filter(
            LedgerGroup.company_id == company_id,
            LedgerGroup.name == "Stock-in-Hand",
        )
        .first()
    )
    if stock_group is None:
        stock_group = LedgerGroup(
            company_id=company_id,
            name="Stock-in-Hand",
            group_type=models.LedgerGroupType.ASSET,
            parent_group_id=None,
        )
        db.add(stock_group)
        db.flush()

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.group_id == stock_group.id,
            Ledger.name.in_(["Purchase", "Purchases", "Purchase (Inventory)"]),
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = Ledger(
        company_id=company_id,
        group_id=stock_group.id,
        name="Purchase",
        code="PURCHASES_ASSET",
        opening_balance=0,
        opening_balance_type=models.OpeningBalanceType.DEBIT,
        is_active=True,
    )
    db.add(ledger)
    db.flush()
    return ledger.id


# -------- Suppliers --------


@router.get("/suppliers", response_model=list[schemas.SupplierRead])
def list_suppliers(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    suppliers = (
        db.query(models.Supplier)
        .filter(models.Supplier.company_id == company_id)
        .order_by(models.Supplier.name)
        .all()
    )
    return suppliers


@router.post("/suppliers", response_model=schemas.SupplierRead)
def create_supplier(
    company_id: int,
    supplier_in: schemas.SupplierCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    supplier_data = supplier_in.model_dump()

    # If caller explicitly provides a ledger_id, honor it; otherwise create a
    # dedicated ledger for this supplier under the company's "Sundry Creditors"
    # group so each supplier has its own ledger.
    if supplier_data.get("ledger_id") is None:
        creditor_group = (
            db.query(models.LedgerGroup)
            .filter(
                models.LedgerGroup.company_id == company_id,
                models.LedgerGroup.name == "Sundry Creditors",
            )
            .first()
        )
        if not creditor_group:
            raise HTTPException(
                status_code=400,
                detail="Ledger group 'Sundry Creditors' not found for this company",
            )

        if creditor_group.group_type in (models.LedgerGroupType.ASSET, models.LedgerGroupType.EXPENSE):
            ob_type = models.OpeningBalanceType.DEBIT
        else:
            ob_type = models.OpeningBalanceType.CREDIT

        supplier_ledger = models.Ledger(
            company_id=company_id,
            group_id=creditor_group.id,
            name=supplier_data.get("name", "Supplier"),
            code=None,
            opening_balance=0,
            opening_balance_type=ob_type,
            is_active=True,
        )
        db.add(supplier_ledger)
        db.flush()
        supplier_data["ledger_id"] = supplier_ledger.id
    supplier = models.Supplier(
        company_id=company_id,
        created_by_id=current_user.id,
        updated_by_id=current_user.id,
        tenant_id=current_user.tenant_id,
        **supplier_data,
    )
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    return supplier


@router.get("/suppliers/{supplier_id}", response_model=schemas.SupplierRead)
def get_supplier(
    company_id: int,
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return supplier


@router.put("/suppliers/{supplier_id}", response_model=schemas.SupplierRead)
def update_supplier(
    company_id: int,
    supplier_id: int,
    supplier_in: schemas.SupplierUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    for field, value in supplier_in.model_dump(exclude_unset=True).items():
        setattr(supplier, field, value)
    supplier.updated_by_id = current_user.id
    db.commit()
    db.refresh(supplier)
    return supplier


@router.delete("/suppliers/{supplier_id}")
def delete_supplier(
    company_id: int,
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    db.delete(supplier)
    db.commit()
    return {"detail": "Deleted"}
def _get_counterparty_ledger_for_payment_mode(
    db: Session,
    company_id: int,
    *,
    payment_mode_id: int | None,
    fallback_ledger_id: int,
) -> int:
    """Resolve the counterparty ledger for purchase flows.

    Priority:
    1) If payment_mode_id is provided: use PaymentMode.ledger_id (and validate company + active).
    2) Else: fall back to provided fallback_ledger_id (e.g. supplier.ledger_id for credit purchases).
    """

    if payment_mode_id is not None:
        pm = (
            db.query(models.PaymentMode)
            .filter(
                models.PaymentMode.id == payment_mode_id,
                models.PaymentMode.company_id == company_id,
                models.PaymentMode.is_active == True,
            )
            .first()
        )
        if not pm:
            raise HTTPException(status_code=400, detail="Invalid payment_mode_id")

        # Treat a payment mode named 'Credit' (or similar casing) as a true
        # credit purchase: do not use its ledger; instead, fall back to the
        # supplier's ledger so Sundry Creditors are affected as expected.
        if pm.name.strip().lower() == "credit":
            return fallback_ledger_id

        return pm.ledger_id

    return fallback_ledger_id


# -------- Purchase Bills --------


def _build_purchase_voucher(
    db: Session,
    company_id: int,
    bill: models.PurchaseBill,
    payment_mode_id: int | None,
    purchase_ledger_id: int | None = None,
    input_tax_ledger_id: int | None = None,
    payment_ledger_id: int | None = None,
    existing_voucher: models.Voucher | None = None,
) -> models.Voucher:
    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == bill.supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=400, detail="Supplier not found")

    # Load company so we can automatically fall back to its default
    # purchase and input tax ledgers when item-level bindings are not
    # configured.
    company = (
        db.query(models.Company)
        .filter(models.Company.id == company_id)
        .first()
    )
    if not company:
        raise HTTPException(status_code=400, detail="Company not found")

    default_stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)
    default_purchase_asset_ledger_id = _get_default_purchase_asset_ledger_id(
        db, company_id=company_id
    )

    # Aggregate lines by expense, tax, and dimensions
    expense_totals: dict[tuple[int, int | None, int | None], float] = {}
    tax_totals: dict[tuple[int, int | None, int | None], float] = {}
    
    grand_total = 0.0

    # Work from the persisted bill lines to avoid issues where the in-memory
    # relationship bill.lines is empty or stale at the time this helper is
    # called. This ensures the voucher always reflects the actual stored bill
    # details.
    bill_lines = (
        db.query(models.PurchaseBillLine)
        .filter(
            models.PurchaseBillLine.bill_id == bill.id,
        )
        .all()
    )

    # Batch-fetch all items and warehouses in 2 queries instead of N+1
    _bill_item_ids = list({int(l.item_id) for l in bill_lines})
    _bill_items = (
        db.query(models.Item)
        .filter(models.Item.id.in_(_bill_item_ids), models.Item.company_id == company_id)
        .all()
    )
    _bill_item_map: dict[int, models.Item] = {int(i.id): i for i in _bill_items}

    _bill_wh_ids = list({int(l.warehouse_id) for l in bill_lines if l.warehouse_id})
    _bill_wh_map: dict[int, models.Warehouse] = {}
    if _bill_wh_ids:
        _bill_wh_rows = db.query(models.Warehouse).filter(models.Warehouse.id.in_(_bill_wh_ids)).all()
        _bill_wh_map = {int(w.id): w for w in _bill_wh_rows}

    for line in bill_lines:
        item = _bill_item_map.get(int(line.item_id))
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

        # Resolve dimensions for this line (Hub & Spoke fallback)
        _wh = _bill_wh_map.get(int(line.warehouse_id)) if line.warehouse_id else None
        warehouse_dept_id = _wh.department_id if _wh else None
        warehouse_proj_id = _wh.project_id if _wh else None
        line_dept_id = line.department_id or warehouse_dept_id or bill.department_id
        line_proj_id = line.project_id or warehouse_proj_id or bill.project_id
        
        subtotal = float(line.quantity) * float(line.rate) - float(line.discount)
        tax = subtotal * float(line.tax_rate) / 100.0
        grand_total += subtotal + tax

        # Expense aggregation
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False))
        
        if is_fixed_asset and item.expense_ledger_id:
            eff_ledger_id = item.expense_ledger_id
        elif purchase_ledger_id is not None:
            eff_ledger_id = purchase_ledger_id
        else:
            is_stock_item = not bool(getattr(item, "allow_negative_stock", False)) and not is_fixed_asset
            if is_stock_item:
                eff_ledger_id = (
                    company.default_purchase_ledger_id
                    or default_purchase_asset_ledger_id
                    or default_stock_ledger_id
                )
            else:
                eff_ledger_id = (
                    item.expense_ledger_id
                    or company.default_purchase_ledger_id
                    or default_purchase_asset_ledger_id
                    or default_stock_ledger_id
                )
        
        if eff_ledger_id is None:
            raise HTTPException(
                status_code=400,
                detail="Item missing expense/inventory ledger and no default purchase ledger configured",
            )
        
        exp_key = (eff_ledger_id, line_dept_id, line_proj_id)
        expense_totals[exp_key] = expense_totals.get(exp_key, 0.0) + subtotal

        # Tax aggregation
        if tax:
            eff_input_tax_ledger_id = None
            if getattr(line, "duty_tax_id", None):
                dt = db.query(models.DutyTax).filter(models.DutyTax.id == line.duty_tax_id).first()
                if dt and dt.ledger_id:
                    eff_input_tax_ledger_id = dt.ledger_id

            if eff_input_tax_ledger_id is None:
                eff_input_tax_ledger_id = input_tax_ledger_id or item.input_tax_ledger_id or company.default_input_tax_ledger_id
                
            if eff_input_tax_ledger_id is None:
                eff_input_tax_ledger_id = item.expense_ledger_id or company.default_purchase_ledger_id

            if eff_input_tax_ledger_id is None:
                raise HTTPException(
                    status_code=400,
                    detail="Item missing input tax ledger and no default input tax ledger configured",
                )
            tax_key = (eff_input_tax_ledger_id, line_dept_id, line_proj_id)
            tax_totals[tax_key] = tax_totals.get(tax_key, 0.0) + tax

    # ── TDS deduction ──────────────────────────────────────────────────
    # When TDS is applied, the supplier is paid net of TDS.
    # Accounting:
    #   DR  Expense / Stock                 (subtotal)
    #   DR  Input VAT                       (tax)
    #   CR  Supplier (Sundry Creditor)      (grand_total - tds_amount)
    #   CR  TDS Payable ledger              (tds_amount)
    tds_amount_value = 0.0
    tds_payable_ledger_id: int | None = None
    if getattr(bill, "apply_tds", False) and getattr(bill, "tds_amount", None):
        tds_amount_value = float(bill.tds_amount or 0)
        tds_payable_ledger_id = getattr(bill, "tds_ledger_id", None)

    id_rows: list = []
    import_supplier_cr = 0.0
    if getattr(bill, "apply_import_duties", False) and (
        (getattr(bill, "purchase_type", "") or "").upper() == "IMPORT"
    ):
        id_rows = (
            db.query(models.PurchaseBillImportDutyLine)
            .filter(models.PurchaseBillImportDutyLine.bill_id == bill.id)
            .order_by(models.PurchaseBillImportDutyLine.sort_order)
            .all()
        )
        import_supplier_cr = sum(float(r.credit_amount or 0) for r in id_rows)

    supplier_net = grand_total - tds_amount_value + import_supplier_cr

    # Persist purchase bill vouchers as PURCHASE_BILL type.
    if existing_voucher is not None:
        voucher = existing_voucher
        voucher.voucher_date = bill.date
        voucher.bill_date = bill.bill_date
        voucher.narration = f"Purchase bill {bill.reference or bill.id}"
        voucher.payment_mode_id = payment_mode_id
        
        # Clear existing lines and allocations
        db.query(models.VoucherLine).filter(
            models.VoucherLine.voucher_id == voucher.id
        ).delete()
        db.query(models.VoucherAllocation).filter(
            models.VoucherAllocation.voucher_id == voucher.id
        ).delete()
        db.flush()
    else:
        voucher_number, fiscal_year, next_seq = get_next_voucher_number(
            db, company_id, models.VoucherType.PURCHASE_BILL, bill.date
        )
        voucher = models.Voucher(
            company_id=company_id,
            voucher_date=bill.date,
            voucher_type=models.VoucherType.PURCHASE_BILL,
            fiscal_year=fiscal_year,
            voucher_sequence=next_seq,
            voucher_number=voucher_number,
            narration=f"Purchase bill {bill.reference or bill.id}",
            payment_mode_id=payment_mode_id,
            bill_date=bill.bill_date,
        )
        db.add(voucher)
        db.flush()

    payment_mode: models.PaymentMode | None = None
    if payment_mode_id is not None:
        payment_mode = (
            db.query(models.PaymentMode)
            .filter(
                models.PaymentMode.id == payment_mode_id,
                models.PaymentMode.company_id == company_id,
                models.PaymentMode.is_active == True,
            )
            .first()
        )
        if not payment_mode:
            raise HTTPException(status_code=400, detail="Invalid payment_mode_id")

    is_credit_mode = bool(
        payment_mode is not None and payment_mode.name.strip().lower() == "credit"
    )

    # Always book the liability to the supplier ledger.
    counterparty_ledger_id = supplier.ledger_id

    db.add(
        models.VoucherLine(
            voucher_id=voucher.id,
            ledger_id=counterparty_ledger_id,
            debit=0,
            credit=supplier_net,
            department_id=bill.department_id,
            project_id=bill.project_id,
        )
    )

    # CR TDS Payable ledger (if TDS is applied)
    if tds_amount_value and tds_payable_ledger_id:
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=tds_payable_ledger_id,
                debit=0,
                credit=tds_amount_value,
                department_id=bill.department_id,
                project_id=bill.project_id,
            )
        )

    # If the bill is marked as paid via a cash/bank payment mode, also record
    # the settlement leg so the cash/bank ledger is affected and the supplier
    # ledger shows both the bill and the payment.
    if payment_mode is not None and not is_credit_mode:
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=counterparty_ledger_id,
                debit=supplier_net,
                credit=0,
                department_id=bill.department_id,
                project_id=bill.project_id,
            )
        )
        eff_pm_ledger_id = payment_ledger_id or payment_mode.ledger_id
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=eff_pm_ledger_id,
                debit=0,
                credit=supplier_net,
                department_id=bill.department_id,
                project_id=bill.project_id,
            )
        )

    # DR expense
    for (ledger_id, dept_id, proj_id), amount in expense_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=amount,
                    credit=0,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    # DR tax
    for (ledger_id, dept_id, proj_id), amount in tax_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=amount,
                    credit=0,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    # IMPORT add-on: Dr expense ledgers for import duties; Cr supplier included in supplier_net above.
    for row in id_rows:
        d = float(row.debit_amount or 0)
        if d <= 0:
            continue
        leg_chk = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == row.ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not leg_chk:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid import duty ledger_id {row.ledger_id}",
            )
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=row.ledger_id,
                debit=d,
                credit=0,
                department_id=bill.department_id,
                project_id=bill.project_id,
            )
        )

    return voucher


@router.get("/bills", response_model=list[schemas.PurchaseBillRead])
def list_bills(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    bills_with_ledger = (
        db.query(models.PurchaseBill, models.Ledger.name.label("purchase_ledger_name"))
        .outerjoin(models.Ledger, models.Ledger.id == models.PurchaseBill.purchase_ledger_id)
        .filter(models.PurchaseBill.company_id == company_id)
        .order_by(models.PurchaseBill.date.desc(), models.PurchaseBill.id.desc())
        .all()
    )
    
    results = []
    for bill, ledger_name in bills_with_ledger:
        bill_data = schemas.PurchaseBillRead.model_validate(bill)
        bill_data.purchase_ledger_name = ledger_name
        results.append(bill_data)
    return results


@router.post("/bills", response_model=schemas.PurchaseBillRead)
def create_bill(
    company_id: int,
    bill_in: schemas.PurchaseBillCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):

    # Load company so we can use and, if necessary, backfill its default
    # purchase/input tax ledgers for legacy companies.
    company = _get_company(db, company_id, current_user)
    validate_transaction_date(company, bill_in.date)

    # Validate supplier belongs to this company before inserting the bill
    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == bill_in.supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=400, detail="Supplier not found")

    # For legacy companies that were created before default_purchase_ledger_id
    # and default_input_tax_ledger_id were wired, attempt to backfill those
    # fields from the standard seeded ledgers so header-level binding works
    # automatically going forward.
    if company.default_purchase_ledger_id is None or company.default_input_tax_ledger_id is None:
        from ..models import Ledger  # local import to avoid cycles

        # Backfill default_purchase_ledger_id from stock/current-asset ledgers if missing.
        if company.default_purchase_ledger_id is None:
            stock_ledger = (
                db.query(Ledger)
                .filter(
                    Ledger.company_id == company_id,
                    Ledger.code == "CLOSING_STOCK",
                )
                .first()
            )
            if stock_ledger is None:
                stock_ledger = (
                    db.query(Ledger)
                    .filter(
                        Ledger.company_id == company_id,
                        Ledger.code == "OPENING_STOCK",
                    )
                    .first()
                )
            if stock_ledger is None:
                stock_ledger = (
                    db.query(Ledger)
                    .filter(
                        Ledger.company_id == company_id,
                        Ledger.code == "PURCHASES",
                    )
                    .first()
                )
            if stock_ledger is not None:
                company.default_purchase_ledger_id = stock_ledger.id

        # Backfill default_input_tax_ledger_id from INPUT_TAX/INPUT_VAT if missing.
        if company.default_input_tax_ledger_id is None:
            input_tax_ledger = (
                db.query(Ledger)
                .filter(
                    Ledger.company_id == company_id,
                    Ledger.code.in_(["INPUT_TAX", "INPUT_VAT"]),
                )
                .first()
            )
            if input_tax_ledger is not None:
                company.default_input_tax_ledger_id = input_tax_ledger.id

    # Determine effective header-level ledgers: prefer explicit values from the
    # request, otherwise fall back to company-level defaults (if configured).
    default_purchase_asset_ledger_id = _get_default_purchase_asset_ledger_id(
        db, company_id=company_id
    )
    default_stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)
    effective_purchase_ledger_id = (
        bill_in.purchase_ledger_id
        if bill_in.purchase_ledger_id is not None
        else (
            company.default_purchase_ledger_id
            or default_purchase_asset_ledger_id
            or default_stock_ledger_id
        )
    )
    effective_input_tax_ledger_id = (
        bill_in.input_tax_ledger_id
        if bill_in.input_tax_ledger_id is not None
        else company.default_input_tax_ledger_id
    )

    # Validate that any explicitly or implicitly chosen header-level ledgers
    # actually exist for this company before creating the bill and voucher.
    if effective_purchase_ledger_id is not None:
        purchase_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_purchase_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not purchase_ledger:
            raise HTTPException(status_code=400, detail="Invalid purchase_ledger_id")

    if effective_input_tax_ledger_id is not None:
        input_tax_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_input_tax_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not input_tax_ledger:
            raise HTTPException(status_code=400, detail="Invalid input_tax_ledger_id")

    effective_purchase_type = (bill_in.purchase_type or "LOCAL").upper()
    bill = models.PurchaseBill(
        company_id=company_id,
        supplier_id=bill_in.supplier_id,
        date=bill_in.date,
        due_date=bill_in.due_date,
        reference=bill_in.reference,
        purchase_ledger_id=effective_purchase_ledger_id,
        input_tax_ledger_id=effective_input_tax_ledger_id,
        department_id=bill_in.department_id,
        project_id=bill_in.project_id,
        narration=bill_in.narration,
        payment_mode_id=bill_in.payment_mode_id,
        payment_ledger_id=bill_in.payment_ledger_id,
        bill_date=bill_in.bill_date,
        purchase_type=effective_purchase_type,
        pragyapan_patra_no=bill_in.pragyapan_patra_no if effective_purchase_type == "IMPORT" else None,
        lc_no=bill_in.lc_no if effective_purchase_type == "IMPORT" else None,
        import_invoice_no=bill_in.import_invoice_no if effective_purchase_type == "IMPORT" else None,
        apply_tds=bool(getattr(bill_in, "apply_tds", False)),
        tds_amount=getattr(bill_in, "tds_amount", None),
        tds_ledger_id=getattr(bill_in, "tds_ledger_id", None),
        apply_import_duties=bool(
            effective_purchase_type == "IMPORT"
            and bool(getattr(bill_in, "apply_import_duties", False))
        ),
    )
    db.add(bill)
    db.flush()

    default_warehouse = None

    for line in bill_in.lines:
        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

        effective_warehouse_id = line.warehouse_id
        if effective_warehouse_id is None:
            if default_warehouse is None:
                default_warehouse = (
                    db.query(models.Warehouse)
                    .filter(
                        models.Warehouse.company_id == company_id,
                        models.Warehouse.name == "Main",
                        models.Warehouse.is_active == True,
                    )
                    .first()
                )
                if not default_warehouse:
                    raise HTTPException(status_code=400, detail="Default warehouse 'Main' not found")
            effective_warehouse_id = default_warehouse.id

        warehouse = (
            db.query(models.Warehouse)
            .filter(
                models.Warehouse.id == effective_warehouse_id,
                models.Warehouse.company_id == company_id,
                models.Warehouse.is_active == True,
            )
            .first()
        )
        if not warehouse:
            raise HTTPException(status_code=400, detail="Invalid warehouse_id")

        resolved_rate, stored_fc_rate, stored_ex_rate = _resolve_import_line_rate(
            effective_purchase_type,
            line.rate,
            line.foreign_currency_rate,
            line.exchange_rate,
        )
        bill_line = models.PurchaseBillLine(
            bill_id=bill.id,
            item_id=line.item_id,
            quantity=line.quantity,
            rate=resolved_rate,
            discount=line.discount,
            tax_rate=line.tax_rate,
            hs_code=line.hs_code,
            warehouse_id=effective_warehouse_id,
            department_id=line.department_id or bill_in.department_id,
            project_id=line.project_id or bill_in.project_id,
            remarks=line.remarks,
            foreign_currency_rate=stored_fc_rate,
            exchange_rate=stored_ex_rate,
        )
        db.add(bill_line)
        db.flush()

        # Only track inventory valuation for strict stock items. 
        # Service items and Fixed Assets should not affect inventory valuation.
        # But Fixed Assets DO need StockLedger entries to track quantity.
        is_service = (item.category and item.category.strip().lower() == "service")
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False))
        
        if not is_service:
            if not is_fixed_asset:
                StockValuationService(db).fifo_add_batch(
                    tenant_id=int(current_user.tenant_id or company.tenant_id),
                    product_id=int(line.item_id),
                    qty_in=float(line.quantity),
                    rate=float(line.rate or 0),
                    ref_type="PURCHASE",
                    ref_id=int(bill.id),
                    created_at=datetime.combine(bill.date, datetime.min.time()),
                )
            
            db.add(
                models.StockLedger(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    qty_delta=float(line.quantity),
                    unit_cost=float(line.rate) if line.rate is not None else None,
                    source_type="PURCHASE_BILL",
                    source_id=bill.id,
                    source_line_id=bill_line.id,
                    posted_at=datetime.combine(bill.date, datetime.min.time()),
                    created_by=current_user.id,
                )
            )

            db.add(
                models.StockMovement(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    movement_date=bill.date,
                    source_type="PURCHASE_BILL",
                    source_id=bill.id,
                    qty_in=line.quantity,
                    qty_out=0,
                )
            )

    if bill.apply_import_duties and effective_purchase_type == "IMPORT":
        for i, row in enumerate(getattr(bill_in, "import_duty_lines", None) or []):
            lg = (
                db.query(models.Ledger)
                .filter(
                    models.Ledger.id == row.ledger_id,
                    models.Ledger.company_id == company_id,
                )
                .first()
            )
            if not lg:
                raise HTTPException(status_code=400, detail="Invalid import duty ledger_id")
            db.add(
                models.PurchaseBillImportDutyLine(
                    bill_id=bill.id,
                    sort_order=i,
                    product_label=row.product_label,
                    ledger_id=row.ledger_id,
                    tax_base=float(row.tax_base or 0),
                    debit_amount=float(row.debit_amount or 0),
                    credit_amount=float(row.credit_amount or 0),
                )
            )
        db.flush()

    voucher = _build_purchase_voucher(
        db,
        company_id,
        bill,
        bill_in.payment_mode_id,
        purchase_ledger_id=bill.purchase_ledger_id,
        input_tax_ledger_id=bill.input_tax_ledger_id,
        payment_ledger_id=bill.payment_ledger_id,
    )

    # Link the created voucher back to the bill so voucher_id is populated
    bill.voucher_id = voucher.id

    db.commit()
    db.refresh(bill)
    return bill


@router.get("/bills/export-template")
def export_purchase_bill_template(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import Workbook
    import io
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Invoices"
    
    headers = [
        "Supplier Name", "Date (YYYY-MM-DD)", "Bill Date (YYYY-MM-DD)", "Due Date (YYYY-MM-DD)",
        "Bill Reference", "Payment Mode",
        "Purchase Ledger", "Input Tax Ledger", "Purchase Type (LOCAL/IMPORT)",
        "Pragyapan Patra No.", "LC No.", "Bill No.",
        "Item Name", "Quantity", "FC rate", "exch. rate", "Rate", "Discount", "Tax Rate (%)", "HS Code",
        "Duty Tax Name", "Remarks",
        "Warehouse Name", "Department Name", "Project Name", "Segment Name",
        "Narration"
    ]
    ws.append(headers)

    # Sample LOCAL row (IMPORT-specific columns left blank)
    ws.append([
        "ABC Suppliers", "2023-01-01", "", "", "INV-001", "Credit",
        "", "", "LOCAL",
        "", "", "",
        "Computer Mouse", "10", "", "", "500", "0", "13", "8471.30",
        "", "",
        "Main Warehouse", "Finance", "", "",
        "Bulk purchase of peripherals"
    ])
    # Sample IMPORT row
    ws.append([
        "XYZ Importers", "2023-02-01", "", "2023-02-15", "IMP-001", "Credit",
        "", "", "IMPORT",
        "PP-2023-001", "LC-2023-005", "COMM-INV-100",
        "Laptop", "5", "1200", "133.50", "", "0", "13", "8471.30",
        "", "",
        "Main Warehouse", "", "", "",
        "Import of laptops"
    ])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_bill_template.xlsx"}
    )




@router.get("/bills/{bill_id}", response_model=schemas.PurchaseBillRead)
def get_bill(
    company_id: int,
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    result = (
        db.query(models.PurchaseBill, models.Ledger.name.label("purchase_ledger_name"))
        .options(
            selectinload(models.PurchaseBill.lines),
            selectinload(models.PurchaseBill.import_duty_lines),
        )
        .outerjoin(models.Ledger, models.Ledger.id == models.PurchaseBill.purchase_ledger_id)
        .filter(
            models.PurchaseBill.id == bill_id,
            models.PurchaseBill.company_id == company_id,
        )
        .first()
    )
    if not result:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    bill, ledger_name = result
    bill_data = schemas.PurchaseBillRead.model_validate(bill)
    bill_data.purchase_ledger_name = ledger_name
    return bill_data



def recharge_stock_from_purchase_bill(db: Session, company_id: int, bill: models.PurchaseBill, current_user_id: int | None = None):
    # Get existing stock ledger entries for this bill
    existing_ledger_entries = (
        db.query(models.StockLedger)
        .filter(
            models.StockLedger.company_id == company_id,
            models.StockLedger.source_type == "PURCHASE_BILL",
            models.StockLedger.source_id == bill.id,
            models.StockLedger.reversed_at.is_(None),
        )
        .all()
    )
    
    # Create a map of existing entries by (item_id, warehouse_id)
    existing_ledger_map: dict[tuple[int, int], models.StockLedger] = {}
    for entry in existing_ledger_entries:
        key = (entry.item_id, entry.warehouse_id)
        existing_ledger_map[key] = entry

    # Get existing stock movements for this bill
    existing_movements = (
        db.query(models.StockMovement)
        .filter(
            models.StockMovement.company_id == company_id,
            models.StockMovement.source_type == "PURCHASE_BILL",
            models.StockMovement.source_id == bill.id,
        )
        .all()
    )
    
    # Create a map of existing movements by (item_id, warehouse_id)
    existing_movement_map: dict[tuple[int, int], models.StockMovement] = {}
    for movement in existing_movements:
        key = (movement.item_id, movement.warehouse_id)
        existing_movement_map[key] = movement

    # Also remove old stock batches so we can recreate them if needed
    db.query(models.StockBatch).filter(
        models.StockBatch.tenant_id == int(bill.company.tenant_id),
        models.StockBatch.ref_type == "PURCHASE",
        models.StockBatch.ref_id == bill.id,
    ).delete()

    # Fetch item categories and properties
    item_ids = {line.item_id for line in bill.lines}
    item_settings = (
        db.query(models.Item.id, models.Item.category, models.Item.is_fixed_asset)
        .filter(models.Item.company_id == company_id, models.Item.id.in_(item_ids))
        .all()
    )
    is_service_map = {row.id: (row.category and row.category.strip().lower() == "service") for row in item_settings}
    is_fixed_asset_map = {row.id: bool(row.is_fixed_asset) for row in item_settings}

    for bill_line in bill.lines:
        effective_warehouse_id = bill_line.warehouse_id
        if effective_warehouse_id is None:
            continue

        # Only track inventory for stock items.
        if not is_service_map.get(bill_line.item_id):
            if not is_fixed_asset_map.get(bill_line.item_id):
                StockValuationService(db).fifo_add_batch(
                    tenant_id=int(bill.company.tenant_id),
                    product_id=int(bill_line.item_id),
                    qty_in=float(bill_line.quantity),
                    rate=float(bill_line.rate or 0),
                    ref_type="PURCHASE",
                    ref_id=int(bill.id),
                    created_at=datetime.combine(bill.date, datetime.min.time()),
                )

            # Check if we have an existing ledger entry for this item/warehouse combination
            ledger_key = (bill_line.item_id, effective_warehouse_id)
            existing_ledger = existing_ledger_map.get(ledger_key)
            
            if existing_ledger:
                # Update the existing ledger entry
                existing_ledger.qty_delta = float(bill_line.quantity)
                existing_ledger.unit_cost = float(bill_line.rate) if bill_line.rate is not None else None
                existing_ledger.source_line_id = bill_line.id
                existing_ledger.posted_at = datetime.combine(bill.date, datetime.min.time())
                # Remove from map so we know it's been processed
                del existing_ledger_map[ledger_key]
            else:
                # Create new ledger entry
                db.add(
                    models.StockLedger(
                        company_id=company_id,
                        warehouse_id=effective_warehouse_id,
                        item_id=bill_line.item_id,
                        qty_delta=float(bill_line.quantity),
                        unit_cost=float(bill_line.rate) if bill_line.rate is not None else None,
                        source_type="PURCHASE_BILL",
                        source_id=bill.id,
                        source_line_id=bill_line.id,
                        posted_at=datetime.combine(bill.date, datetime.min.time()),
                        created_by=current_user_id,
                    )
                )

            # Check if we have an existing movement entry for this item/warehouse combination
            movement_key = (bill_line.item_id, effective_warehouse_id)
            existing_movement = existing_movement_map.get(movement_key)
            
            if existing_movement:
                # Update the existing movement entry
                existing_movement.movement_date = bill.date
                existing_movement.qty_in = bill_line.quantity
                existing_movement.qty_out = 0
                # Remove from map so we know it's been processed
                del existing_movement_map[movement_key]
            else:
                # Create new movement entry
                db.add(
                    models.StockMovement(
                        company_id=company_id,
                        warehouse_id=effective_warehouse_id,
                        item_id=bill_line.item_id,
                        movement_date=bill.date,
                        source_type="PURCHASE_BILL",
                        source_id=bill.id,
                        qty_in=bill_line.quantity,
                        qty_out=0,
                    )
                )

    # Delete any remaining ledger and movement entries that weren't matched (items removed from bill)
    for remaining_entry in existing_ledger_map.values():
        db.delete(remaining_entry)
    for remaining_movement in existing_movement_map.values():
        db.delete(remaining_movement)

@router.put("/bills/{bill_id}", response_model=schemas.PurchaseBillRead)
def update_bill(
    company_id: int,
    bill_id: int,
    bill_in: schemas.PurchaseBillUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    if bill_in.date:
        validate_transaction_date(company, bill_in.date)
    bill = (
        db.query(models.PurchaseBill)
        .filter(
            models.PurchaseBill.id == bill_id,
            models.PurchaseBill.company_id == company_id,
        )
        .first()
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    for field, value in bill_in.model_dump(
        exclude_unset=True,
        exclude={"lines", "import_duty_lines"},
    ).items():
        # Do not overwrite non-nullable fields with None; treat None as "no change".
        if field in {"supplier_id", "date"} and value is None:
            continue
        # Also treat payment_mode_id=None as "no change" so existing mode is preserved.
        if field == "payment_mode_id" and value is None:
            continue
        # For header-level ledgers and dimensions, explicitly preserve existing values when None is sent.
        if field in {"purchase_ledger_id", "input_tax_ledger_id", "department_id", "project_id"} and value is None:
            continue
        setattr(bill, field, value)

    # If the purchase type was updated to LOCAL, clear IMPORT-only header fields.
    if (bill.purchase_type or "LOCAL").upper() != "IMPORT":
        bill.pragyapan_patra_no = None
        bill.lc_no = None
        bill.import_invoice_no = None
        bill.apply_import_duties = False
        db.query(models.PurchaseBillImportDutyLine).filter(
            models.PurchaseBillImportDutyLine.bill_id == bill.id
        ).delete()
        db.flush()

    if bill_in.lines is not None:
        # Delete old bill lines (we'll recreate them)
        db.query(models.PurchaseBillLine).filter(
            models.PurchaseBillLine.bill_id == bill.id
        ).delete()

        default_warehouse = None

        for line in bill_in.lines:
            # We must fetch the item to check allow_negative_stock and ensure it exists
            item = (
                db.query(models.Item)
                .filter(
                    models.Item.id == line.item_id,
                    models.Item.company_id == company_id,
                )
                .first()
            )
            if not item:
                raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

            effective_warehouse_id = line.warehouse_id
            if effective_warehouse_id is None:
                if default_warehouse is None:
                    default_warehouse = (
                        db.query(models.Warehouse)
                        .filter(
                            models.Warehouse.company_id == company_id,
                            models.Warehouse.name == "Main",
                            models.Warehouse.is_active == True,
                        )
                        .first()
                    )
                    if not default_warehouse:
                        raise HTTPException(status_code=400, detail="Default warehouse 'Main' not found")
                effective_warehouse_id = default_warehouse.id

            warehouse = (
                db.query(models.Warehouse)
                .filter(
                    models.Warehouse.id == effective_warehouse_id,
                    models.Warehouse.company_id == company_id,
                    models.Warehouse.is_active == True,
                )
                .first()
            )
            if not warehouse:
                raise HTTPException(status_code=400, detail="Invalid warehouse_id")

            update_purchase_type = (bill.purchase_type or "LOCAL").upper()
            resolved_rate, stored_fc_rate, stored_ex_rate = _resolve_import_line_rate(
                update_purchase_type,
                line.rate,
                line.foreign_currency_rate,
                line.exchange_rate,
            )
            bill_line = models.PurchaseBillLine(
                bill_id=bill.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=resolved_rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
                warehouse_id=effective_warehouse_id,
                department_id=line.department_id or bill_in.department_id,
                project_id=line.project_id or bill_in.project_id,
                remarks=line.remarks,
                foreign_currency_rate=stored_fc_rate,
                exchange_rate=stored_ex_rate,
            )
            db.add(bill_line)

        db.flush()
        # Refresh stock entries for the newly created lines
        recharge_stock_from_purchase_bill(db, company_id, bill, current_user.id)

    if "import_duty_lines" in bill_in.model_fields_set or "apply_import_duties" in bill_in.model_fields_set:
        db.query(models.PurchaseBillImportDutyLine).filter(
            models.PurchaseBillImportDutyLine.bill_id == bill.id
        ).delete()
        db.flush()
        upd_pt = (bill.purchase_type or "LOCAL").upper()
        if upd_pt == "IMPORT" and getattr(bill, "apply_import_duties", False):
            rows_in = (
                bill_in.import_duty_lines
                if "import_duty_lines" in bill_in.model_fields_set
                and bill_in.import_duty_lines is not None
                else []
            )
            for i, row in enumerate(rows_in):
                lg = (
                    db.query(models.Ledger)
                    .filter(
                        models.Ledger.id == row.ledger_id,
                        models.Ledger.company_id == company_id,
                    )
                    .first()
                )
                if not lg:
                    raise HTTPException(status_code=400, detail="Invalid import duty ledger_id")
                db.add(
                    models.PurchaseBillImportDutyLine(
                        bill_id=bill.id,
                        sort_order=i,
                        product_label=row.product_label,
                        ledger_id=row.ledger_id,
                        tax_base=float(row.tax_base or 0),
                        debit_amount=float(row.debit_amount or 0),
                        credit_amount=float(row.credit_amount or 0),
                    )
                )
            db.flush()

    # Rebuild the linked voucher using the same logic as create_bill.
    # Use bill.payment_mode_id which reflects the updated state (or existing state if not changed).
    payment_mode_id = bill.payment_mode_id

    # Ensure header-level ledgers are always resolved using the same rules as create_bill:
    # prefer explicit values on the bill, otherwise fall back to company defaults.
    default_purchase_asset_ledger_id = _get_default_purchase_asset_ledger_id(
        db, company_id=company_id
    )
    default_stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)
    effective_purchase_ledger_id = (
        bill.purchase_ledger_id
        if bill.purchase_ledger_id is not None
        else (
            company.default_purchase_ledger_id
            or default_purchase_asset_ledger_id
            or default_stock_ledger_id
        )
    )
    effective_input_tax_ledger_id = (
        bill.input_tax_ledger_id
        if bill.input_tax_ledger_id is not None
        else company.default_input_tax_ledger_id
    )

    # Validate that the resolved header-level ledgers still exist for this company
    # before rebuilding the voucher. This prevents bills from pointing to
    # non-existent ledgers after configuration changes.
    if effective_purchase_ledger_id is not None:
        purchase_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_purchase_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not purchase_ledger:
            raise HTTPException(status_code=400, detail="Invalid purchase_ledger_id")

    if effective_input_tax_ledger_id is not None:
        input_tax_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_input_tax_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not input_tax_ledger:
            raise HTTPException(status_code=400, detail="Invalid input_tax_ledger_id")

    bill.purchase_ledger_id = effective_purchase_ledger_id
    bill.input_tax_ledger_id = effective_input_tax_ledger_id

    # Rebuild the linked voucher.
    existing_voucher: models.Voucher | None = None
    if bill.voucher_id is not None:
        existing_voucher = (
            db.query(models.Voucher)
            .filter(
                models.Voucher.id == bill.voucher_id,
                models.Voucher.company_id == company_id,
            )
            .first()
        )

    # Rebuild voucher only if there are lines on the bill; otherwise there is
    # nothing to account for.
    if bill.lines:
        voucher = _build_purchase_voucher(
            db,
            company_id,
            bill,
            payment_mode_id,
            purchase_ledger_id=effective_purchase_ledger_id,
            input_tax_ledger_id=effective_input_tax_ledger_id,
            payment_ledger_id=bill.payment_ledger_id,
            existing_voucher=existing_voucher,
        )
        bill.voucher_id = voucher.id
    elif existing_voucher is not None:
        # If no lines remain, disconnect the voucher (logic for deletion could be added here if desired)
        bill.voucher_id = None
        db.delete(existing_voucher)
        db.flush()

    db.commit()
    db.refresh(bill)
    return bill


@router.delete("/bills/{bill_id}")
def delete_bill(
    company_id: int,
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    if db.in_transaction():
        result = _delete_purchase_bill_internal(
            db=db,
            company_id=company_id,
            bill_id=bill_id,
            actor_user_id=current_user.id,
            skip_consumption_check=False,
        )
    else:
        with db.begin():
            result = _delete_purchase_bill_internal(
                db=db,
                company_id=company_id,
                bill_id=bill_id,
                actor_user_id=current_user.id,
                skip_consumption_check=False,
            )
    return result


@router.post("/bills/{bill_id}/reverse", response_model=schemas.PurchaseReturnRead)
async def reverse_bill(
    company_id: int,
    bill_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    raw = await request.body()
    data: dict = {}
    if raw:
        try:
            data = json.loads(raw.decode("utf-8"))
            if data is None:
                data = {}
        except Exception:
            data = {}

    reverse_in = schemas.PurchaseBillReverseRequest.model_construct(**(data or {}))
    if getattr(reverse_in, "date", None) is not None and isinstance(reverse_in.date, str):
        try:
            reverse_in.date = datetime.fromisoformat(reverse_in.date).date()
        except Exception:
            reverse_in.date = None
    return _reverse_purchase_bill_internal(
        db=db,
        company_id=company_id,
        bill_id=bill_id,
        reverse_in=reverse_in,
        current_user=current_user,
    )


def _reverse_purchase_bill_internal(
    *,
    db: Session,
    company_id: int,
    bill_id: int,
    reverse_in: schemas.PurchaseBillReverseRequest,
    current_user: models.User,
) -> models.PurchaseReturn:
    _get_company(db, company_id, current_user)

    bill = (
        db.query(models.PurchaseBill)
        .filter(
            models.PurchaseBill.id == bill_id,
            models.PurchaseBill.company_id == company_id,
        )
        .first()
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    effective_ref = (
        reverse_in.reference
        if reverse_in.reference is not None and str(reverse_in.reference).strip()
        else f"REV-{bill_id}"
    )

    existing_return = (
        db.query(models.PurchaseReturn)
        .filter(
            models.PurchaseReturn.company_id == company_id,
            models.PurchaseReturn.source_bill_id == bill_id,
            models.PurchaseReturn.reference == effective_ref,
        )
        .first()
    )
    if existing_return is not None:
        return existing_return

    bill_lines = (
        db.query(models.PurchaseBillLine)
        .filter(models.PurchaseBillLine.bill_id == bill_id)
        .all()
    )
    if not bill_lines:
        raise HTTPException(status_code=400, detail="Bill has no lines")

    impacted_pairs: set[tuple[int, int]] = set()
    for line in bill_lines:
        if line.warehouse_id is None:
            continue
        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        if item is None:
            continue
        # Service items skip stock checks
        if (item.category and item.category.strip().lower() == "service"):
            continue
        impacted_pairs.add((int(line.item_id), int(line.warehouse_id)))

    if impacted_pairs:
        _ensure_purchase_bill_can_be_deleted(
            db=db,
            company_id=company_id,
            bill=bill,
            impacted_pairs=impacted_pairs,
        )

    return_date = reverse_in.date or bill.date

    effective_purchase_return_ledger_id = (
        reverse_in.purchase_return_ledger_id
        if reverse_in.purchase_return_ledger_id is not None
        else bill.purchase_ledger_id
    )
    effective_input_tax_return_ledger_id = (
        reverse_in.input_tax_return_ledger_id
        if reverse_in.input_tax_return_ledger_id is not None
        else bill.input_tax_ledger_id
    )

    purchase_return = models.PurchaseReturn(
        company_id=company_id,
        supplier_id=bill.supplier_id,
        date=return_date,
        reference=effective_ref,
        source_bill_id=bill_id,
        purchase_return_ledger_id=effective_purchase_return_ledger_id,
        input_tax_return_ledger_id=effective_input_tax_return_ledger_id,
        payment_ledger_id=reverse_in.payment_ledger_id,
    )
    db.add(purchase_return)
    db.flush()

    for bill_line in bill_lines:
        db.add(
            models.PurchaseReturnLine(
                return_id=purchase_return.id,
                item_id=bill_line.item_id,
                quantity=float(bill_line.quantity),
                rate=float(bill_line.rate),
                discount=float(bill_line.discount or 0),
                tax_rate=float(bill_line.tax_rate),
                hs_code=bill_line.hs_code,
            )
        )

        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == bill_line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False)) if item else False
        if item is None or (item.category and item.category.strip().lower() == "service"):
            continue
        if bill_line.warehouse_id is None:
            raise HTTPException(
                status_code=400,
                detail="Cannot reverse bill: missing warehouse on stock item line",
            )

        posted_at = datetime.combine(return_date, datetime.min.time())
        db.add(
            models.StockLedger(
                company_id=company_id,
                warehouse_id=bill_line.warehouse_id,
                item_id=bill_line.item_id,
                qty_delta=-float(bill_line.quantity),
                unit_cost=float(bill_line.rate) if bill_line.rate is not None else None,
                source_type="PURCHASE_RETURN",
                source_id=purchase_return.id,
                source_line_id=None,
                posted_at=posted_at,
                created_by=current_user.id,
            )
        )
        db.add(
            models.StockMovement(
                company_id=company_id,
                warehouse_id=bill_line.warehouse_id,
                item_id=bill_line.item_id,
                movement_date=return_date,
                source_type="PURCHASE_RETURN",
                source_id=purchase_return.id,
                qty_in=0,
                qty_out=float(bill_line.quantity),
            )
        )


    _build_purchase_return_voucher(
        db,
        company_id,
        purchase_return,
        reverse_in.payment_mode_id,
        purchase_return_ledger_id=effective_purchase_return_ledger_id,
        input_tax_return_ledger_id=effective_input_tax_return_ledger_id,
        payment_ledger_id=purchase_return.payment_ledger_id,
    )

    db.commit()
    db.refresh(purchase_return)
    return purchase_return


# -------- Purchase Returns --------


def _build_purchase_return_voucher(
    db: Session,
    company_id: int,
    purchase_return: models.PurchaseReturn,
    payment_mode_id: int | None,
    purchase_return_ledger_id: int | None = None,
    input_tax_return_ledger_id: int | None = None,
    payment_ledger_id: int | None = None,
) -> models.Voucher:
    if payment_mode_id in (0, "0"):
        payment_mode_id = None

    supplier = (
        db.query(models.Supplier)
        .filter(
            models.Supplier.id == purchase_return.supplier_id,
            models.Supplier.company_id == company_id,
        )
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=400, detail="Supplier not found")

    # Load company so we can automatically fall back to its default purchase
    # and input tax ledgers when item-level bindings are not configured.
    company = (
        db.query(models.Company)
        .filter(models.Company.id == company_id)
        .first()
    )
    if not company:
        raise HTTPException(status_code=400, detail="Company not found")

    default_stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)
    default_purchase_asset_ledger_id = _get_default_purchase_asset_ledger_id(
        db, company_id=company_id
    )

    # Aggregate lines by expense, tax, and dimensions
    expense_totals: dict[tuple[int, int | None, int | None], float] = {}
    tax_totals: dict[tuple[int, int | None, int | None], float] = {}
    grand_total = 0.0

    for line in purchase_return.lines:
        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

        # Resolve dimensions for this line (Hub & Spoke fallback)
        warehouse_dept_id = line.warehouse.department_id if line.warehouse else None
        warehouse_proj_id = line.warehouse.project_id if line.warehouse else None
        line_dept_id = line.department_id or warehouse_dept_id or purchase_return.department_id
        line_proj_id = line.project_id or warehouse_proj_id or purchase_return.project_id

        subtotal = float(line.quantity) * float(line.rate) - float(line.discount)
        tax = subtotal * float(line.tax_rate) / 100.0
        grand_total += subtotal + tax

        # Expense aggregation
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False))
        
        if is_fixed_asset and item.expense_ledger_id:
            eff_expense_ledger_id = item.expense_ledger_id
        elif purchase_return_ledger_id is not None:
            eff_expense_ledger_id = purchase_return_ledger_id
        else:
            is_stock_item = not (item.category and item.category.strip().lower() == "service") and not is_fixed_asset
            if is_stock_item:
                eff_expense_ledger_id = (
                    default_stock_ledger_id
                    or default_purchase_asset_ledger_id
                    or company.default_purchase_ledger_id
                )
            else:
                eff_expense_ledger_id = (
                    item.expense_ledger_id
                    or default_stock_ledger_id
                    or default_purchase_asset_ledger_id
                    or company.default_purchase_ledger_id
                )
        
        if eff_expense_ledger_id is None:
            raise HTTPException(
                status_code=400,
                detail="Item missing expense/inventory ledger and no default purchase ledger configured",
            )
        exp_key = (eff_expense_ledger_id, line_dept_id, line_proj_id)
        expense_totals[exp_key] = expense_totals.get(exp_key, 0.0) + subtotal

        # Tax aggregation
        if tax:
            eff_tax_ledger_id = None
            if getattr(line, "duty_tax_id", None):
                dt = db.query(models.DutyTax).filter(models.DutyTax.id == line.duty_tax_id).first()
                if dt and dt.ledger_id:
                    eff_tax_ledger_id = dt.ledger_id

            if eff_tax_ledger_id is None:
                eff_tax_ledger_id = input_tax_return_ledger_id or item.input_tax_ledger_id or company.default_input_tax_ledger_id
                
            if eff_tax_ledger_id is None:
                 eff_tax_ledger_id = item.expense_ledger_id or company.default_purchase_ledger_id

            if eff_tax_ledger_id is None:
                raise HTTPException(
                    status_code=400,
                    detail="Item missing input tax ledger and no default input tax ledger configured",
                )
            tax_key = (eff_tax_ledger_id, line_dept_id, line_proj_id)
            tax_totals[tax_key] = tax_totals.get(tax_key, 0.0) + tax

    voucher_number, fiscal_year, next_seq = get_next_voucher_number(
        db, company_id, models.VoucherType.PURCHASE_RETURN, purchase_return.date
    )
    voucher = models.Voucher(
        company_id=company_id,
        voucher_date=purchase_return.date,
        voucher_type=models.VoucherType.PURCHASE_RETURN,
        fiscal_year=fiscal_year,
        voucher_sequence=next_seq,
        voucher_number=voucher_number,
        narration=f"Purchase return {purchase_return.reference or purchase_return.id}",
        payment_mode_id=payment_mode_id,
    )
    db.add(voucher)
    db.flush()

    pm_ledger_id = payment_ledger_id
    if pm_ledger_id is None:
        pm_ledger_id = _get_counterparty_ledger_for_payment_mode(
            db,
            company_id,
            payment_mode_id=payment_mode_id,
            fallback_ledger_id=supplier.ledger_id,
        )

    db.add(
        models.VoucherLine(
            voucher_id=voucher.id,
            ledger_id=pm_ledger_id,
            debit=grand_total,
            credit=0,
            department_id=purchase_return.department_id,
            project_id=purchase_return.project_id,
        )
    )

    # CR expense
    for (ledger_id, dept_id, proj_id), amount in expense_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=0,
                    credit=amount,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    # CR tax
    for (ledger_id, dept_id, proj_id), amount in tax_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=0,
                    credit=amount,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    return voucher


@router.post("/returns", response_model=schemas.PurchaseReturnRead)
def create_purchase_return(
    company_id: int,
    return_in: schemas.PurchaseReturnCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    validate_transaction_date(company, return_in.date)

    # Resolve header-level ledgers for the return: prefer explicit values from
    # the payload. For purchase returns, we default to the company's dedicated
    # "Purchase Return" ledger (code PURCHASE_RETURN) so all returns aggregate
    # into a single contra-purchase account in P&L. Input tax return continues
    # to fall back to the company's default input tax ledger.
    effective_purchase_return_ledger_id = return_in.purchase_return_ledger_id
    if effective_purchase_return_ledger_id is None:
        default_purchase_asset_ledger_id = _get_default_purchase_asset_ledger_id(
            db, company_id=company_id
        )
        default_stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)
        effective_purchase_return_ledger_id = (
            default_stock_ledger_id
            or default_purchase_asset_ledger_id
            or company.default_purchase_ledger_id
        )

    effective_input_tax_return_ledger_id = (
        return_in.input_tax_return_ledger_id
        if return_in.input_tax_return_ledger_id is not None
        else company.default_input_tax_ledger_id
    )

    # Validate that any resolved header-level ledgers actually exist for this company
    # before we create vouchers that reference them.
    if effective_purchase_return_ledger_id is not None:
        purchase_return_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_purchase_return_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not purchase_return_ledger:
            raise HTTPException(status_code=400, detail="Invalid purchase_return_ledger_id")

    if effective_input_tax_return_ledger_id is not None:
        input_tax_return_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.id == effective_input_tax_return_ledger_id,
                models.Ledger.company_id == company_id,
            )
            .first()
        )
        if not input_tax_return_ledger:
            raise HTTPException(status_code=400, detail="Invalid input_tax_return_ledger_id")

    purchase_return = models.PurchaseReturn(
        company_id=company_id,
        supplier_id=return_in.supplier_id,
        date=return_in.date,
        reference=return_in.reference,
        source_bill_id=return_in.source_bill_id,
        department_id=return_in.department_id,
        project_id=return_in.project_id,
        purchase_return_ledger_id=effective_purchase_return_ledger_id,
        input_tax_return_ledger_id=effective_input_tax_return_ledger_id,
        payment_ledger_id=return_in.payment_ledger_id,
    )
    db.add(purchase_return)
    db.flush()

    source_bill = None
    if return_in.source_bill_id is not None:
        source_bill = (
            db.query(models.PurchaseBill)
            .filter(
                models.PurchaseBill.id == return_in.source_bill_id,
                models.PurchaseBill.company_id == company_id,
            )
            .first()
        )

    default_warehouse = None

    for line in return_in.lines:
        effective_warehouse_id = line.warehouse_id
        if effective_warehouse_id is None and source_bill is not None:
            for bill_line in source_bill.lines:
                if bill_line.item_id == line.item_id:
                    effective_warehouse_id = bill_line.warehouse_id
                    break

        if effective_warehouse_id is None:
            if default_warehouse is None:
                default_warehouse = (
                    db.query(models.Warehouse)
                    .filter(
                        models.Warehouse.company_id == company_id,
                        models.Warehouse.name == "Main",
                        models.Warehouse.is_active == True,
                    )
                    .first()
                )
                if not default_warehouse:
                    raise HTTPException(status_code=400, detail="Default warehouse 'Main' not found")
            effective_warehouse_id = default_warehouse.id

        db.add(
            models.PurchaseReturnLine(
                return_id=purchase_return.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=line.rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
                department_id=line.department_id or return_in.department_id,
                project_id=line.project_id or return_in.project_id,
                warehouse_id=effective_warehouse_id,
            )
        )

        warehouse = (
            db.query(models.Warehouse)
            .filter(
                models.Warehouse.id == effective_warehouse_id,
                models.Warehouse.company_id == company_id,
                models.Warehouse.is_active == True,
            )
            .first()
        )
        if not warehouse:
            raise HTTPException(status_code=400, detail="Invalid warehouse_id for purchase return")

        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

        # Only track inventory for stock items. Service items should not affect stock.
        if not (item.category and item.category.strip().lower() == "service"):
            db.add(
                models.StockLedger(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    qty_delta=-float(line.quantity),
                    unit_cost=float(line.rate) if line.rate is not None else None,
                    source_type="PURCHASE_RETURN",
                    source_id=purchase_return.id,
                    source_line_id=None,
                    posted_at=datetime.combine(purchase_return.date, datetime.min.time()),
                    created_by=current_user.id,
                )
            )

            db.add(
                models.StockMovement(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    movement_date=purchase_return.date,
                    source_type="PURCHASE_RETURN",
                    source_id=purchase_return.id,
                    qty_in=0,
                    qty_out=line.quantity,
                )
            )
    voucher = _build_purchase_return_voucher(
        db,
        company_id,
        purchase_return,
        return_in.payment_mode_id,
        purchase_return_ledger_id=effective_purchase_return_ledger_id,
        input_tax_return_ledger_id=effective_input_tax_return_ledger_id,
        payment_ledger_id=purchase_return.payment_ledger_id,
    )
    purchase_return.voucher_id = voucher.id

    db.commit()
    db.refresh(purchase_return)
    return purchase_return


@router.get("/returns", response_model=list[schemas.PurchaseReturnRead])
def list_purchase_returns(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    PurchaseLedger = aliased(models.Ledger)
    TaxLedger = aliased(models.Ledger)
    returns_with_ledger = (
        db.query(
            models.PurchaseReturn,
            PurchaseLedger.name.label("purchase_return_ledger_name"),
            TaxLedger.name.label("input_tax_return_ledger_name")
        )
        .outerjoin(PurchaseLedger, PurchaseLedger.id == models.PurchaseReturn.purchase_return_ledger_id)
        .outerjoin(TaxLedger, TaxLedger.id == models.PurchaseReturn.input_tax_return_ledger_id)
        .filter(models.PurchaseReturn.company_id == company_id)
        .order_by(models.PurchaseReturn.date.desc(), models.PurchaseReturn.id.desc())
        .all()
    )
    
    results = []
    for ret, ledger_name, tax_ledger_name in returns_with_ledger:
        ret_data = schemas.PurchaseReturnRead.model_validate(ret)
        ret_data.purchase_return_ledger_name = ledger_name
        ret_data.input_tax_return_ledger_name = tax_ledger_name
        results.append(ret_data)
    return results


@router.get("/returns/{return_id}", response_model=schemas.PurchaseReturnRead)
def get_purchase_return(
    company_id: int,
    return_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    PurchaseLedger = aliased(models.Ledger)
    TaxLedger = aliased(models.Ledger)
    result = (
        db.query(
            models.PurchaseReturn,
            PurchaseLedger.name.label("purchase_return_ledger_name"),
            TaxLedger.name.label("input_tax_return_ledger_name")
        )
        .outerjoin(PurchaseLedger, PurchaseLedger.id == models.PurchaseReturn.purchase_return_ledger_id)
        .outerjoin(TaxLedger, TaxLedger.id == models.PurchaseReturn.input_tax_return_ledger_id)
        .filter(
            models.PurchaseReturn.id == return_id,
            models.PurchaseReturn.company_id == company_id,
        )
        .first()
    )
    if not result:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    
    ret, ledger_name, tax_ledger_name = result
    ret_data = schemas.PurchaseReturnRead.model_validate(ret)
    ret_data.purchase_return_ledger_name = ledger_name
    ret_data.input_tax_return_ledger_name = tax_ledger_name
    return ret_data


@router.post("/bills/{bill_id}/create-return", response_model=schemas.PurchaseReturnRead)
def create_purchase_return_from_bill(
    company_id: int,
    bill_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    bill = (
        db.query(models.PurchaseBill)
        .filter(
            models.PurchaseBill.id == bill_id,
            models.PurchaseBill.company_id == company_id,
        )
        .first()
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    purchase_return = models.PurchaseReturn(
        company_id=company_id,
        supplier_id=bill.supplier_id,
        date=bill.date,
        reference=None,
        source_bill_id=bill.id,
        purchase_return_ledger_id=bill.purchase_ledger_id,
        input_tax_return_ledger_id=bill.input_tax_ledger_id,
    )
    db.add(purchase_return)
    db.flush()

    for line in bill.lines:
        db.add(
            models.PurchaseReturnLine(
                return_id=purchase_return.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=line.rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
            )
        )

    _build_purchase_return_voucher(db, company_id, purchase_return, None)

    db.commit()
    db.refresh(purchase_return)
    return purchase_return


@router.post("/bills/parse-excel")
async def parse_purchase_bills_excel(
    company_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import load_workbook
    from datetime import date, datetime
    
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Empty or invalid file")
    
    headers = [str(c).strip().lower() for c in rows[0] if c is not None]
    data_rows = rows[1:]
    
    # Pre-fetch lookups
    supplier_objs = db.query(models.Supplier).filter(models.Supplier.company_id == company_id).all()
    suppliers = {s.name.lower(): s.id for s in supplier_objs}
    suppliers_by_key = {s.name.lower(): {"id": s.id, "name": s.name} for s in supplier_objs}
    item_objs = db.query(models.Item).filter(models.Item.company_id == company_id).all()
    items = {i.name.lower(): {"id": i.id, "name": i.name} for i in item_objs}
    warehouses = {w.name.lower(): {"id": w.id, "name": w.name} for w in db.query(models.Warehouse).filter(models.Warehouse.company_id == company_id).all()}
    payment_modes = {pm.name.lower(): pm.id for pm in db.query(models.PaymentMode).filter(models.PaymentMode.company_id == company_id).all()}
    departments = {d.name.lower(): {"id": d.id, "name": d.name} for d in db.query(models.Department).filter(models.Department.company_id == company_id).all()}
    projects = {p.name.lower(): {"id": p.id, "name": p.name} for p in db.query(models.Project).filter(models.Project.company_id == company_id).all()}
    segments = {s.name.lower(): {"id": s.id, "name": s.name} for s in db.query(models.Segment).filter(models.Segment.company_id == company_id).all()}

    # Group lines by (Supplier, Date, Reference, Narration, PaymentMode)
    parsed_bills = []
    bills_map = {}
    
    for idx, row_data in enumerate(data_rows, start=2):
        row = dict(zip(headers, row_data))
        
        supplier_name = str(row.get("supplier name") or "").strip()
        if not supplier_name:
            continue
            
        bill_date_val = row.get("date (yyyy-mm-dd)")
        bill_date = date.today().isoformat()
        if bill_date_val:
            if isinstance(bill_date_val, (date, datetime)):
                bill_date = (bill_date_val.date() if isinstance(bill_date_val, datetime) else bill_date_val).isoformat()
            else:
                try:
                    bill_date = datetime.strptime(str(bill_date_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass

        # Optional separate bill_date
        bill_date2_val = row.get("bill date (yyyy-mm-dd)")
        bill_date2 = None
        if bill_date2_val:
            if isinstance(bill_date2_val, (date, datetime)):
                bill_date2 = (bill_date2_val.date() if isinstance(bill_date2_val, datetime) else bill_date2_val).isoformat()
            else:
                try:
                    bill_date2 = datetime.strptime(str(bill_date2_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass

        # Optional due_date
        due_date_val = row.get("due date (yyyy-mm-dd)")
        due_date_str = None
        if due_date_val:
            if isinstance(due_date_val, (date, datetime)):
                due_date_str = (due_date_val.date() if isinstance(due_date_val, datetime) else due_date_val).isoformat()
            else:
                try:
                    due_date_str = datetime.strptime(str(due_date_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass

        reference = str(row.get("bill reference") or "").strip()
        narration = str(row.get("narration") or "").strip()
        pm_name = str(row.get("payment mode") or "").strip()
        purch_ledger_name = str(row.get("purchase ledger") or "").strip()
        tax_ledger_name = str(row.get("input tax ledger") or "").strip()
        purchase_type = str(row.get("purchase type (local/import)") or "LOCAL").strip().upper()
        
        bill_key = (supplier_name, bill_date, reference, pm_name)
        if bill_key not in bills_map:
            supplier_id = suppliers.get(supplier_name.lower())
            supplier_suggestions = []
            if not supplier_id:
                close = difflib.get_close_matches(
                    supplier_name.lower(), list(suppliers.keys()), n=3, cutoff=0.5
                )
                supplier_suggestions = [
                    {"name": suppliers_by_key[k]["name"], "id": suppliers_by_key[k]["id"]}
                    for k in close
                ]

            # Resolve optional ledgers by name
            purch_ledger_id = None
            tax_ledger_id = None
            if purch_ledger_name:
                led = db.query(models.Ledger).filter(
                    models.Ledger.company_id == company_id,
                    models.Ledger.name.ilike(purch_ledger_name)
                ).first()
                purch_ledger_id = led.id if led else None
            if tax_ledger_name:
                led = db.query(models.Ledger).filter(
                    models.Ledger.company_id == company_id,
                    models.Ledger.name.ilike(tax_ledger_name)
                ).first()
                tax_ledger_id = led.id if led else None

            is_import_row = purchase_type == "IMPORT"
            bill_obj = {
                "id": len(parsed_bills),
                "supplier_name": supplier_name,
                "supplier_id": supplier_id,
                "supplier_suggestions": supplier_suggestions,
                "date": bill_date,
                "due_date": due_date_str,
                "bill_date": bill_date2,
                "reference": reference,
                "narration": narration,
                "payment_mode_name": pm_name,
                "payment_mode_id": payment_modes.get(pm_name.lower()),
                "purchase_ledger_name": purch_ledger_name,
                "purchase_ledger_id": purch_ledger_id,
                "input_tax_ledger_name": tax_ledger_name,
                "input_tax_ledger_id": tax_ledger_id,
                "purchase_type": purchase_type,
                "pragyapan_patra_no": str(row.get("pragyapan patra no.") or "").strip() or None if is_import_row else None,
                "lc_no": str(row.get("lc no.") or "").strip() or None if is_import_row else None,
                "import_invoice_no": str(row.get("bill no.") or "").strip() or None if is_import_row else None,
                "lines": [],
                "errors": [],
                "warnings": []
            }
            if not supplier_id:
                if supplier_suggestions:
                    bill_obj["errors"].append(f"Supplier '{supplier_name}' not found — did you mean one of the suggestions below?")
                else:
                    bill_obj["errors"].append(f"Supplier '{supplier_name}' not found")
            
            if purch_ledger_name and not purch_ledger_id:
                bill_obj["warnings"].append(f"Purchase Ledger '{purch_ledger_name}' not found — will use default if empty.")
            if tax_ledger_name and not tax_ledger_id:
                bill_obj["warnings"].append(f"Tax Ledger '{tax_ledger_name}' not found — will use default if empty.")
            
            bills_map[bill_key] = bill_obj
            parsed_bills.append(bill_obj)
        
        item_name = str(row.get("item name") or "").strip()
        item_info = items.get(item_name.lower())
        
        item_suggestions = []
        if not item_info:
            # Fuzzy match: find close item names
            close = difflib.get_close_matches(
                item_name.lower(),
                list(items.keys()),
                n=3, cutoff=0.5
            )
            item_suggestions = [
                {"name": items[k]["name"], "id": items[k]["id"]}
                for k in close
            ]

        wh_raw = str(row.get("warehouse name") or "").strip()
        dept_raw = str(row.get("department name") or "").strip()
        proj_raw = str(row.get("project name") or "").strip()
        seg_raw = str(row.get("segment name") or "").strip()

        wh_info = warehouses.get(wh_raw.lower())
        dept_info = departments.get(dept_raw.lower())
        proj_info = projects.get(proj_raw.lower())
        seg_info = segments.get(seg_raw.lower())

        # Resolve optional Duty Tax by name
        duty_tax_name = str(row.get("duty tax name") or "").strip()
        duty_tax_id = None
        if duty_tax_name:
            dt_obj = db.query(models.DutyTax).filter(
                models.DutyTax.company_id == company_id,
                models.DutyTax.name.ilike(duty_tax_name)
            ).first()
            duty_tax_id = dt_obj.id if dt_obj else None

        parsed_purchase_type = bills_map[bill_key]["purchase_type"]
        def _excel_cell(*keys):
            for k in keys:
                v = row.get(k)
                if v not in (None, "", "none"):
                    return v
            return None

        fc_rate_raw = _excel_cell("fc rate", "rate in foreign currency")
        ex_rate_raw = _excel_cell("exch. rate", "exchange rate")
        fc_rate = float(fc_rate_raw) if fc_rate_raw not in (None, "", "none") else None
        ex_rate = float(ex_rate_raw) if ex_rate_raw not in (None, "", "none") else None
        submitted_rate = float(row.get("rate") or 0)

        line = {
            "item_name": item_name,
            "item_id": item_info["id"] if item_info else None,
            "item_suggestions": item_suggestions,
            "quantity": float(row.get("quantity") or 0),
            "rate": submitted_rate,
            "discount": float(row.get("discount") or 0),
            "tax_rate": float(row.get("tax rate (%)") or 0),
            "hs_code": str(row.get("hs code") or "").strip() or None,
            "duty_tax_name": duty_tax_name or None,
            "duty_tax_id": duty_tax_id,
            "remarks": str(row.get("remarks") or "").strip() or None,
            "warehouse_name": wh_info["name"] if wh_info else wh_raw,
            "warehouse_id": wh_info["id"] if wh_info else None,
            "department_name": dept_info["name"] if dept_info else dept_raw,
            "department_id": dept_info["id"] if dept_info else None,
            "project_name": proj_info["name"] if proj_info else proj_raw,
            "project_id": proj_info["id"] if proj_info else None,
            "segment_name": seg_info["name"] if seg_info else seg_raw,
            "segment_id": seg_info["id"] if seg_info else None,
            "foreign_currency_rate": fc_rate if parsed_purchase_type == "IMPORT" else None,
            "exchange_rate": ex_rate if parsed_purchase_type == "IMPORT" else None,
        }

        # For IMPORT with both FC fields, pre-compute derived rate for display
        if parsed_purchase_type == "IMPORT" and fc_rate and ex_rate:
            line["rate"] = round(fc_rate * ex_rate, 2)

        # Validation Logic
        errors = bills_map[bill_key]["errors"]
        if not line["item_id"]:
            if item_suggestions:
                errors.append(f"Item '{item_name}' not found — did you mean one of the suggestions?")
            else:
                errors.append(f"Item '{item_name}' not found")
        
        if line["quantity"] <= 0:
            errors.append(f"Invalid quantity ({line['quantity']}) for item '{item_name}'")
        if line["rate"] <= 0:
            errors.append(f"Invalid rate ({line['rate']}) for item '{item_name}'")
        if line["discount"] < 0:
            errors.append(f"Negative discount ({line['discount']}) for item '{item_name}'")
        if line["tax_rate"] < 0 or line["tax_rate"] > 100:
            errors.append(f"Suspicious tax rate ({line['tax_rate']}%) for item '{item_name}'")

        # Dimension Warnings
        warnings = bills_map[bill_key].setdefault("warnings", [])
        if duty_tax_name and not duty_tax_id:
            warnings.append(f"Duty Tax '{duty_tax_name}' not found")
        if wh_raw and not wh_info:
            warnings.append(f"Warehouse '{wh_raw}' not found")
        if dept_raw and not dept_info:
            warnings.append(f"Department '{dept_raw}' not found")
        if proj_raw and not proj_info:
            warnings.append(f"Project '{proj_raw}' not found")
        if seg_raw and not seg_info:
            warnings.append(f"Segment '{seg_raw}' not found")
            
        bills_map[bill_key]["lines"].append(line)

    return parsed_bills



@router.post("/bills/confirm-import")
def confirm_purchase_bills_import(
    company_id: int,
    bills_in: list[dict],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    
    created_count = 0
    errors = []
    
    for bdata in bills_in:
        try:
            # Basic validation
            if not bdata.get("supplier_id"):
                errors.append(f"Bill {bdata.get('reference')}: Missing supplier ID")
                continue
            if not bdata.get("lines"):
                errors.append(f"Bill {bdata.get('reference')}: No lines found")
                continue
                
            bill_date = datetime.fromisoformat(bdata["date"]).date()
            validate_transaction_date(company, bill_date)
            
            # Create Bill
            confirm_purchase_type = (bdata.get("purchase_type") or "LOCAL").upper()
            is_import = confirm_purchase_type == "IMPORT"
            bill = models.PurchaseBill(
                company_id=company_id,
                supplier_id=bdata["supplier_id"],
                date=bill_date,
                due_date=(
                    datetime.fromisoformat(bdata["due_date"]).date()
                    if bdata.get("due_date")
                    else None
                ),
                bill_date=(
                    datetime.fromisoformat(bdata["bill_date"]).date() 
                    if bdata.get("bill_date") 
                    else bill_date
                ),
                reference=bdata.get("reference"),
                payment_mode_id=bdata.get("payment_mode_id"),
                payment_ledger_id=bdata.get("payment_ledger_id"),
                purchase_ledger_id=bdata.get("purchase_ledger_id"),
                input_tax_ledger_id=bdata.get("input_tax_ledger_id"),
                narration=bdata.get("narration"),
                department_id=bdata.get("department_id"),
                project_id=bdata.get("project_id"),
                segment_id=bdata.get("segment_id"),
                purchase_type=confirm_purchase_type,
                pragyapan_patra_no=bdata.get("pragyapan_patra_no") if is_import else None,
                lc_no=bdata.get("lc_no") if is_import else None,
                import_invoice_no=bdata.get("import_invoice_no") if is_import else None,
                apply_tds=bdata.get("apply_tds") or False,
                tds_amount=bdata.get("tds_amount"),
                tds_ledger_id=bdata.get("tds_ledger_id"),
            )
            db.add(bill)
            db.flush()
            
            for ldata in bdata["lines"]:
                # Remove display-only fields (names, suggestions) — keep only DB-safe IDs and values
                display_only = {
                    "item_name", "item_suggestions",
                    "warehouse_name",
                    "department_name", "project_name", "segment_name",
                    "duty_tax_name",
                    "supplier_suggestions", "supplier_name",
                    "purchase_ledger_name", "input_tax_ledger_name",
                    "payment_mode_name",
                }
                line_data = {k: v for k, v in ldata.items() if k not in display_only}
                # Derive local rate from FC fields when IMPORT
                resolved_rate, stored_fc_rate, stored_ex_rate = _resolve_import_line_rate(
                    confirm_purchase_type,
                    float(line_data.get("rate") or 0),
                    line_data.pop("foreign_currency_rate", None),
                    line_data.pop("exchange_rate", None),
                )
                line_data["rate"] = resolved_rate
                line = models.PurchaseBillLine(
                    bill_id=bill.id,
                    foreign_currency_rate=stored_fc_rate,
                    exchange_rate=stored_ex_rate,
                    **line_data
                )
                db.add(line)
            
            db.flush()
            # Build Voucher
            _build_purchase_voucher(
                db, 
                company_id, 
                bill, 
                bdata.get("payment_mode_id"),
                purchase_ledger_id=bill.purchase_ledger_id,
                input_tax_ledger_id=bill.input_tax_ledger_id,
                payment_ledger_id=bill.payment_ledger_id,
            )
            created_count += 1
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            errors.append(f"Error creating bill '{bdata.get('reference')}': {str(e)}")
            # For simplicity, we rollback this specific bill's addition if possible 
            # but since we are in one transaction, we might need a savepoint
            # Let's just raise error for now or manage transaction better
            raise HTTPException(status_code=400, detail=f"Import failed at bill {bdata.get('reference')}: {str(e)}")
            
    db.commit()
    return {"detail": f"Successfully imported {created_count} purchase invoices", "errors": errors}


