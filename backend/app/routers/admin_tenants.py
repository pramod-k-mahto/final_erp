# backend/app/routers/admin_tenants.py

from typing import List, Optional

import csv
import io
import json
import logging
import traceback
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.encoders import jsonable_encoder
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import delete, insert, text, update, func
from sqlalchemy import MetaData, Table
from sqlalchemy.orm import Session, selectinload, joinedload, subqueryload

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ..database import get_db
from .. import models, schemas
from ..menu_defaults import (
    ensure_menu_template_has_required_menus,
    ensure_menu_template_assignable_to_tenant,
    get_default_menu_template_id,
)
from ..auth import get_current_admin, get_billing_admin, get_support_admin, get_tech_admin


logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/admin/tenants",
    tags=["admin-tenants"],
    dependencies=[Depends(get_current_admin)],
)

from cryptography.fernet import Fernet
from ..config import get_settings
from pydantic import BaseModel

class LicensePayload(BaseModel):
    license_key: str

@router.post("/{tenant_id}/license")
def apply_license(
    tenant_id: int,
    payload: LicensePayload,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(get_current_admin)
):
    _ensure_same_tenant_or_superadmin(current_admin, tenant_id=tenant_id)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
        
    try:
        raw_key = payload.license_key.removeprefix("PRIXNA-")
        fernet = Fernet(get_settings().license_secret.encode())
        decrypted_bytes = fernet.decrypt(raw_key.encode())
        license_data = json.loads(decrypted_bytes.decode('utf-8'))
        
        tenant.license_key = payload.license_key
        tenant.expires_at = datetime.fromisoformat(license_data["expires_at"])
        tenant.max_users = license_data.get("max_users", 5)
        db.commit()
        return {"detail": "License applied successfully", "expires_at": tenant.expires_at}
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid License Key")


@router.get("/{tenant_id}/subscriptions", response_model=List[schemas.TenantSubscriptionRead])
def list_tenant_subscriptions(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)
    return (
        db.query(models.TenantSubscription)
        .filter(models.TenantSubscription.tenant_id == tenant_id)
        .order_by(models.TenantSubscription.payment_date.desc())
        .all()
    )


@router.get("/ghost/smart-report", response_model=schemas.GhostSmartReportResponse)
def ghost_smart_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    settings = db.query(models.AppSettings).get(1)
    if not settings or not settings.ghost_company_id:
        return {
            "total_sales": 0,
            "total_collections": 0,
            "total_outstanding": 0,
            "recent_transactions": []
        }

    ghost_id = settings.ghost_company_id

    # 1. Total Sales (Subscription Revenue)
    line = models.SalesInvoiceLine
    inv = models.SalesInvoice
    subtotal = (line.quantity * line.rate) - line.discount
    line_total = subtotal + (subtotal * (line.tax_rate / 100.0))
    
    total_sales = db.query(func.coalesce(func.sum(line_total), 0)).join(inv).filter(inv.company_id == ghost_id).scalar()

    # 2. Total Collections (Receipt Vouchers)
    # Get all tenant ledgers (Sundry Debtors)
    dg = db.query(models.LedgerGroup).filter(
        models.LedgerGroup.company_id == ghost_id, 
        models.LedgerGroup.name.ilike("Sundry Debtors%")
    ).first()
    
    debtor_ledger_ids = []
    if dg:
        # For simplicity, we'll take all ledgers in this group and its immediate children
        child_groups = db.query(models.LedgerGroup.id).filter(models.LedgerGroup.parent_group_id == dg.id).all()
        group_ids = [dg.id] + [r[0] for r in child_groups]
        debtor_ledgers = db.query(models.Ledger.id).filter(models.Ledger.group_id.in_(group_ids)).all()
        debtor_ledger_ids = [r[0] for r in debtor_ledgers]

    total_collections = 0
    if debtor_ledger_ids:
        total_collections = db.query(func.coalesce(func.sum(models.VoucherLine.credit), 0))\
            .join(models.Voucher)\
            .filter(
                models.Voucher.company_id == ghost_id,
                models.Voucher.voucher_type == models.VoucherType.RECEIPT,
                models.VoucherLine.ledger_id.in_(debtor_ledger_ids)
            ).scalar()

    # 3. Total Outstanding (Sundry Debtors Balance)
    total_outstanding = 0
    if debtor_ledger_ids:
        vd = db.query(func.coalesce(func.sum(models.VoucherLine.debit), 0))\
            .join(models.Voucher)\
            .filter(models.Voucher.company_id == ghost_id, models.VoucherLine.ledger_id.in_(debtor_ledger_ids)).scalar()
        vc = db.query(func.coalesce(func.sum(models.VoucherLine.credit), 0))\
            .join(models.Voucher)\
            .filter(models.Voucher.company_id == ghost_id, models.VoucherLine.ledger_id.in_(debtor_ledger_ids)).scalar()
        vo = db.query(func.coalesce(func.sum(models.Ledger.opening_balance), 0))\
            .filter(models.Ledger.company_id == ghost_id, models.Ledger.id.in_(debtor_ledger_ids)).scalar()
        total_outstanding = float(vo or 0) + float(vd or 0) - float(vc or 0)

    # 4. Recent Transactions
    from sqlalchemy.orm import joinedload
    recent_invoices = db.query(models.SalesInvoice)\
        .options(joinedload(models.SalesInvoice.customer))\
        .filter(models.SalesInvoice.company_id == ghost_id)\
        .order_by(models.SalesInvoice.date.desc(), models.SalesInvoice.id.desc())\
        .limit(10).all()
    
    recent_receipts = db.query(models.Voucher)\
        .filter(
            models.Voucher.company_id == ghost_id, 
            models.Voucher.voucher_type == models.VoucherType.RECEIPT
        )\
        .order_by(models.Voucher.voucher_date.desc(), models.Voucher.id.desc())\
        .limit(10).all()
    
    recent_txs = []
    for i in recent_invoices:
        i_total = db.query(func.coalesce(func.sum(line_total), 0)).filter(line.invoice_id == i.id).scalar()
        recent_txs.append({
            "date": i.date,
            "type": "SALE",
            "tenant_name": i.customer.name if i.customer else "Unknown",
            "amount": float(i_total or 0),
            "reference": i.reference or f"INV-{i.id}"
        })
    
    for r in recent_receipts:
        r_total = db.query(func.coalesce(func.sum(models.VoucherLine.debit), 0)).filter(models.VoucherLine.voucher_id == r.id).scalar()
        # Find the customer ledger to get the name
        customer_line = db.query(models.VoucherLine).join(models.Ledger).filter(
            models.VoucherLine.voucher_id == r.id, 
            models.Ledger.id.in_(debtor_ledger_ids)
        ).first()
        tenant_name = customer_line.ledger.name if customer_line else "Unknown"
        
        recent_txs.append({
            "date": r.voucher_date,
            "type": "PAYMENT",
            "tenant_name": tenant_name,
            "amount": float(r_total or 0),
            "reference": r.voucher_number or f"RCPT-{r.id}"
        })
    
    recent_txs.sort(key=lambda x: x["date"], reverse=True)

    return {
        "total_sales": float(total_sales or 0),
        "total_collections": float(total_collections or 0),
        "total_outstanding": total_outstanding,
        "recent_transactions": recent_txs[:10]
    }


@router.get("/ghost/reports/sales")
def get_ghost_sales_report(
    from_date: str = Query(None),
    to_date: str = Query(None),
    tenant_id: int = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    settings = db.query(models.AppSettings).get(1)
    if not settings or not settings.ghost_company_id:
        raise HTTPException(status_code=400, detail="Ghost accounting not configured")
    ghost_id = settings.ghost_company_id

    query = db.query(models.SalesInvoice).options(joinedload(models.SalesInvoice.customer))\
        .filter(models.SalesInvoice.company_id == ghost_id)
    
    if from_date:
        query = query.filter(models.SalesInvoice.date >= from_date)
    if to_date:
        query = query.filter(models.SalesInvoice.date <= to_date)
    if tenant_id:
        query = query.filter(models.SalesInvoice.customer_id == tenant_id)
    
    invoices = query.order_by(models.SalesInvoice.date.desc(), models.SalesInvoice.id.desc()).all()
    
    line = models.SalesInvoiceLine
    results = []
    for i in invoices:
        subtotal = (line.quantity * line.rate) - line.discount
        line_total = subtotal + (subtotal * (line.tax_rate / 100.0))
        i_total = db.query(func.coalesce(func.sum(line_total), 0)).filter(line.invoice_id == i.id).scalar()
        
        results.append({
            "id": i.id,
            "date": i.date,
            "reference": i.reference,
            "tenant_name": i.customer.name if i.customer else "Unknown",
            "amount": float(i_total or 0),
            "status": "PAID" if i.is_paid else "UNPAID"
        })
    
    return results


@router.get("/ghost/reports/collections")
def get_ghost_collections_report(
    from_date: str = Query(None),
    to_date: str = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    settings = db.query(models.AppSettings).get(1)
    if not settings or not settings.ghost_company_id:
        raise HTTPException(status_code=400, detail="Ghost accounting not configured")
    ghost_id = settings.ghost_company_id

    # 1. Identify "Sundry Debtors" groups for Ghost Company
    debtor_groups = db.query(models.LedgerGroup).filter(
        models.LedgerGroup.company_id == ghost_id,
        models.LedgerGroup.name.ilike("%Sundry Debtors%")
    ).all()
    debtor_group_ids = [g.id for g in debtor_groups]
    
    # Get all debtor ledgers
    debtor_ledgers = db.query(models.Ledger).filter(
        models.Ledger.company_id == ghost_id,
        models.Ledger.group_id.in_(debtor_group_ids)
    ).all()
    debtor_ledger_ids = [l.id for l in debtor_ledgers]

    query = db.query(models.Voucher).filter(
        models.Voucher.company_id == ghost_id,
        models.Voucher.voucher_type == models.VoucherType.RECEIPT
    )
    
    if from_date:
        query = query.filter(models.Voucher.voucher_date >= from_date)
    if to_date:
        query = query.filter(models.Voucher.voucher_date <= to_date)
    
    receipts = query.order_by(models.Voucher.voucher_date.desc(), models.Voucher.id.desc()).all()
    
    results = []
    for r in receipts:
        r_total = db.query(func.coalesce(func.sum(models.VoucherLine.debit), 0)).filter(models.VoucherLine.voucher_id == r.id).scalar()
        
        # Find customer line
        customer_line = db.query(models.VoucherLine).join(models.Ledger).filter(
            models.VoucherLine.voucher_id == r.id, 
            models.Ledger.id.in_(debtor_ledger_ids)
        ).first()
        
        results.append({
            "id": r.id,
            "date": r.voucher_date,
            "voucher_number": r.voucher_number,
            "tenant_name": customer_line.ledger.name if customer_line else "Unknown",
            "amount": float(r_total or 0),
            "payment_mode": r.payment_mode.name if r.payment_mode else "Cash",
            "narration": r.narration
        })
    
    return results


@router.get("/ghost/reports/debtors")
def get_ghost_debtors_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    settings = db.query(models.AppSettings).get(1)
    if not settings or not settings.ghost_company_id:
        raise HTTPException(status_code=400, detail="Ghost accounting not configured")
    ghost_id = settings.ghost_company_id

    # 1. Identify "Sundry Debtors" groups
    debtor_groups = db.query(models.LedgerGroup).filter(
        models.LedgerGroup.company_id == ghost_id,
        models.LedgerGroup.name.ilike("%Sundry Debtors%")
    ).all()
    debtor_group_ids = [g.id for g in debtor_groups]
    
    debtor_ledgers = db.query(models.Ledger).filter(
        models.Ledger.company_id == ghost_id,
        models.Ledger.group_id.in_(debtor_group_ids)
    ).all()
    
    results = []
    for l in debtor_ledgers:
        # Calculate current balance
        vo = db.query(func.coalesce(func.sum(models.Ledger.opening_balance), 0)).filter(models.Ledger.id == l.id).scalar()
        vd = db.query(func.coalesce(func.sum(models.VoucherLine.debit), 0)).filter(models.VoucherLine.ledger_id == l.id).scalar()
        vc = db.query(func.coalesce(func.sum(models.VoucherLine.credit), 0)).filter(models.VoucherLine.ledger_id == l.id).scalar()
        
        balance = float(vo or 0) + float(vd or 0) - float(vc or 0)
        
        if abs(balance) > 0.01:
            results.append({
                "ledger_id": l.id,
                "tenant_name": l.name,
                "balance": balance,
                "email": l.email,
                "phone": l.phone
            })
    
    # Sort by balance descending
    results.sort(key=lambda x: x["balance"], reverse=True)
    return results


@router.get("/payment-modes")
def list_ghost_payment_modes(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    settings = db.query(models.AppSettings).get(1)
    if not settings or not settings.ghost_company_id:
        return []
        
    return (
        db.query(models.PaymentMode)
        .filter(models.PaymentMode.company_id == settings.ghost_company_id, models.PaymentMode.is_active == True)
        .all()
    )


class PaymentModeCreateRequest(BaseModel):
    name: str


@router.post("/payment-modes")
def create_ghost_payment_mode(
    payload: PaymentModeCreateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    try:
        settings = db.query(models.AppSettings).first()
        if not settings or not settings.ghost_company_id:
            logger.error("Ghost accounting failed: settings.ghost_company_id is missing")
            raise HTTPException(status_code=400, detail="Ghost accounting not configured")
            
        ghost_company_id = settings.ghost_company_id
        ghost_tenant_id = settings.ghost_tenant_id
        
        logger.info(f"Creating ghost bank '{payload.name}' for company {ghost_company_id}")

        # 1. Ensure Ledger exists for this bank
        from ..services.ghost_accounting import _ensure_ledger
        bank_ledger = _ensure_ledger(
            db, ghost_company_id, ghost_tenant_id,
            f"BANK_{payload.name.upper().replace(' ', '_')}", payload.name,
            "Bank Accounts", models.LedgerGroupType.ASSET
        )
        
        # 2. Check if PaymentMode already exists
        pm = db.query(models.PaymentMode).filter(
            models.PaymentMode.company_id == ghost_company_id,
            models.PaymentMode.name.ilike(payload.name)
        ).first()
        
        if pm:
            logger.info(f"Bank '{payload.name}' already exists as PM ID {pm.id}")
            return pm
            
        # 3. Create PaymentMode
        try:
            pm = models.PaymentMode(
                company_id=ghost_company_id,
                tenant_id=ghost_tenant_id,
                name=payload.name,
                ledger_id=bank_ledger.id
            )
            db.add(pm)
            db.commit()
            db.refresh(pm)
        except Exception as e:
            db.rollback()
            # Check for unique constraint violation
            if "unique constraint" in str(e).lower() or "duplicate key" in str(e).lower():
                 raise HTTPException(status_code=400, detail=f"Bank '{payload.name}' already exists.")
            raise e

        logger.info(f"Successfully created ghost bank '{payload.name}' (PM ID {pm.id})")
        return pm
    except Exception as e:
        logger.error(f"Error creating ghost bank: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{tenant_id}/record-payment", response_model=schemas.TenantSubscriptionRead)
def record_tenant_payment(
    tenant_id: int,
    payload: schemas.TenantSubscriptionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_billing_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)

    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    subscription = models.TenantSubscription(
        tenant_id=tenant_id,
        plan_code=payload.plan_code,
        amount_paid=payload.amount_paid,
        period_start=payload.period_start,
        period_end=payload.period_end,
        payment_method=payload.payment_method,
        bank_name=payload.bank_name,
        reference_no=payload.reference_no,
        status=payload.status,
    )
    db.add(subscription)

    # Update tenant plan and expiry
    tenant.plan = payload.plan_code
    tenant.expires_at = payload.period_end

    db.commit()
    db.refresh(subscription)

    # Sync to Ghost Accounting
    try:
        sync_subscription_to_accounting(db, subscription)
    except Exception as e:
        logger.error(f"Ghost accounting sync failed: {e}")
        # We don't fail the primary transaction if ghost sync fails

    return subscription


BACKUP_FORMAT_VERSION = 1


def _ensure_superadmin(current_user: models.User) -> None:
    if current_user.role not in (models.UserRole.superadmin, models.UserRole.ghost_tech, models.UserRole.ghost_support, models.UserRole.ghost_billing):
        raise HTTPException(status_code=403, detail="Ghost Admin privileges required")


def _ensure_same_tenant_or_superadmin(current_user: models.User, *, tenant_id: int) -> None:
    if current_user.role in (models.UserRole.superadmin, models.UserRole.ghost_billing, models.UserRole.ghost_support, models.UserRole.ghost_tech):
        return
    if current_user.tenant_id is None or int(current_user.tenant_id) != int(tenant_id):
        raise HTTPException(status_code=403, detail="Cannot access another tenant")


def _company_backup_tables() -> list[type[models.Base]]:
    # Order matters for restore (parents before children).
    return [
        models.CompanySettings,
        models.LedgerGroup,
        models.Ledger,
        models.PaymentMode,
        models.Customer,
        models.Supplier,
        models.Item,
        models.ItemUnit,
        models.Warehouse,
        models.Voucher,
        models.PurchaseBill,
        models.PurchaseBillLine,
        models.SalesInvoice,
        models.SalesInvoiceLine,
        models.VoucherLine,
        models.VoucherAllocation,
        models.VoucherLog,
        models.SalesReturn,
        models.SalesReturnLine,
        models.PurchaseReturn,
        models.PurchaseReturnLine,
        models.StockTransfer,
        models.StockTransferLine,
        models.StockMovement,
        models.StockLedger,
    ]


def _reflect_table(db: Session, table_name: str) -> Table:
    bind = db.get_bind()
    md = MetaData()
    return Table(table_name, md, autoload_with=bind)


def _company_backup_delete_tables() -> list[type[models.Base]]:
    # Reverse-ish dependency order (children first).
    return [
        models.StockLedger,
        models.StockMovement,
        models.StockTransferLine,
        models.StockTransfer,
        models.PurchaseReturnLine,
        models.PurchaseReturn,
        models.SalesReturnLine,
        models.SalesReturn,
        models.VoucherLog,
        models.VoucherAllocation,
        models.VoucherLine,
        models.SalesInvoiceLine,
        models.SalesInvoice,
        models.PurchaseBillLine,
        models.PurchaseBill,
        models.Voucher,
        models.Warehouse,
        models.ItemUnit,
        models.Item,
        models.Supplier,
        models.Customer,
        models.PaymentMode,
        models.Ledger,
        models.LedgerGroup,
        models.CompanySettings,
    ]


def _safe_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        return None


def _safe_restore_new_company(
    db: Session,
    *,
    tenant_id: int,
    backup: dict,
    actor_user_id: int,
) -> int:
    tables = backup.get("tables") or {}
    company_data = backup.get("company") or {}
    if not isinstance(tables, dict) or not isinstance(company_data, dict):
        raise HTTPException(status_code=400, detail="Invalid backup payload")

    base_name = str(company_data.get("name") or "Restored Company").strip() or "Restored Company"
    new_name = f"{base_name} (Restored {datetime.utcnow().strftime('%Y-%m-%d %H%M%S')})"

    new_company = models.Company(
        owner_id=int(actor_user_id),
        tenant_id=int(tenant_id),
        name=new_name,
        fiscal_year_start=company_data.get("fiscal_year_start"),
        fiscal_year_end=company_data.get("fiscal_year_end"),
        address=company_data.get("address"),
        phone=company_data.get("phone"),
        pan_number=company_data.get("pan_number"),
        business_type=company_data.get("business_type"),
        country=company_data.get("country"),
        currency=company_data.get("currency"),
        inventory_valuation_method=company_data.get("inventory_valuation_method")
        or models.InventoryValuationMethod.AVERAGE,
        cost_center_mode=company_data.get("cost_center_mode"),
        cost_center_single_dimension=company_data.get("cost_center_single_dimension"),
    )
    db.add(new_company)
    db.flush()
    new_company_id = int(new_company.id)

    def _rows(name: str) -> list[dict]:
        r = tables.get(name) or []
        if not isinstance(r, list):
            raise HTTPException(status_code=400, detail=f"Invalid rows for table: {name}")
        return [dict(x) for x in r]

    id_map: dict[str, dict[int, int]] = {
        "ledger_groups": {},
        "ledgers": {},
        "payment_modes": {},
        "customers": {},
        "suppliers": {},
        "items": {},
        "warehouses": {},
        "vouchers": {},
        "purchase_bills": {},
        "sales_invoices": {},
        "sales_returns": {},
        "purchase_returns": {},
        "stock_transfers": {},
    }

    def _table(name: str) -> Table:
        return _reflect_table(db, name)

    # CompanySettings (1:1): re-create without preserving ID.
    for r in _rows(models.CompanySettings.__tablename__):
        t = _table(models.CompanySettings.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        db.execute(insert(t), [payload])
        break

    # Sales returns
    for r in _rows(models.SalesReturn.__tablename__):
        t = _table(models.SalesReturn.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "customer_id" in allowed:
            payload["customer_id"] = id_map["customers"].get(_safe_int(r.get("customer_id")) or -1)
        if "source_invoice_id" in allowed:
            old_src = _safe_int(r.get("source_invoice_id"))
            if old_src is not None:
                payload["source_invoice_id"] = id_map["sales_invoices"].get(old_src)
        if payload.get("customer_id") is None and "customer_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["sales_returns"][old_id] = new_id

    for r in _rows(models.SalesReturnLine.__tablename__):
        t = _table(models.SalesReturnLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "return_id" in allowed:
            payload["return_id"] = id_map["sales_returns"].get(_safe_int(r.get("return_id")) or -1)
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
        if "warehouse_id" in allowed:
            wh_old = _safe_int(r.get("warehouse_id"))
            if wh_old is not None:
                payload["warehouse_id"] = id_map["warehouses"].get(wh_old)
        if payload.get("return_id") is None or payload.get("item_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Purchase returns
    for r in _rows(models.PurchaseReturn.__tablename__):
        t = _table(models.PurchaseReturn.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "supplier_id" in allowed:
            payload["supplier_id"] = id_map["suppliers"].get(_safe_int(r.get("supplier_id")) or -1)
        if "source_bill_id" in allowed:
            old_src = _safe_int(r.get("source_bill_id"))
            if old_src is not None:
                payload["source_bill_id"] = id_map["purchase_bills"].get(old_src)
        if payload.get("supplier_id") is None and "supplier_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["purchase_returns"][old_id] = new_id

    for r in _rows(models.PurchaseReturnLine.__tablename__):
        t = _table(models.PurchaseReturnLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "return_id" in allowed:
            payload["return_id"] = id_map["purchase_returns"].get(_safe_int(r.get("return_id")) or -1)
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
        if payload.get("return_id") is None or payload.get("item_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Ledger groups
    for r in _rows(models.LedgerGroup.__tablename__):
        t = _table(models.LedgerGroup.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "parent_group_id" in allowed:
            payload["parent_group_id"] = None
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["ledger_groups"][old_id] = new_id

    # second pass parent_group_id remap
    for r in _rows(models.LedgerGroup.__tablename__):
        old_id = _safe_int(r.get("id"))
        parent_old = _safe_int(r.get("parent_group_id"))
        if old_id is None or parent_old is None:
            continue
        new_id = id_map["ledger_groups"].get(old_id)
        new_parent = id_map["ledger_groups"].get(parent_old)
        if new_id and new_parent:
            db.execute(text("UPDATE ledger_groups SET parent_group_id=:p WHERE id=:i"), {"p": int(new_parent), "i": int(new_id)})

    # Ledgers
    for r in _rows(models.Ledger.__tablename__):
        t = _table(models.Ledger.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "group_id" in allowed:
            payload["group_id"] = id_map["ledger_groups"].get(_safe_int(r.get("group_id")) or -1)
        if payload.get("group_id") is None and "group_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["ledgers"][old_id] = new_id

    # Update company default ledger pointers
    for field in (
        "default_purchase_ledger_id",
        "default_sales_ledger_id",
        "default_item_income_ledger_id",
        "default_item_expense_ledger_id",
        "default_input_tax_ledger_id",
        "default_output_tax_ledger_id",
        "default_item_input_tax_ledger_id",
        "default_item_output_tax_ledger_id",
    ):
        old_val = _safe_int(company_data.get(field))
        if old_val is not None and old_val in id_map["ledgers"]:
            setattr(new_company, field, int(id_map["ledgers"][old_val]))

    # Payment modes
    for r in _rows(models.PaymentMode.__tablename__):
        t = _table(models.PaymentMode.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "tenant_id" in allowed:
            payload["tenant_id"] = int(tenant_id)
        ledger_old = _safe_int(r.get("ledger_id"))
        if "ledger_id" in allowed:
            payload["ledger_id"] = id_map["ledgers"].get(ledger_old or -1)
        if payload.get("ledger_id") is None and "ledger_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["payment_modes"][old_id] = new_id

    # Customers/Suppliers
    for cls, key in ((models.Customer, "customers"), (models.Supplier, "suppliers")):
        t = _table(cls.__tablename__)
        for r in _rows(cls.__tablename__):
            old_id = _safe_int(r.get("id"))
            item = dict(r)
            item.pop("id", None)
            allowed = set(t.c.keys())
            payload = {k: v for k, v in item.items() if k in allowed}
            if "company_id" in allowed:
                payload["company_id"] = new_company_id
            ledger_old = _safe_int(r.get("ledger_id"))
            if ledger_old is not None:
                if "ledger_id" in allowed:
                    payload["ledger_id"] = id_map["ledgers"].get(ledger_old)
            ins = insert(t)
            if "id" in t.c:
                ins = ins.returning(t.c.id)
                res = db.execute(ins, [payload])
                new_id = int(res.scalar_one())
            else:
                db.execute(ins, [payload])
                new_id = int(old_id or 0)
            if old_id is not None:
                id_map[key][old_id] = new_id

    # Items
    for r in _rows(models.Item.__tablename__):
        t = _table(models.Item.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        for fk in ("income_ledger_id", "expense_ledger_id", "output_tax_ledger_id", "input_tax_ledger_id"):
            old_fk = _safe_int(r.get(fk))
            if old_fk is not None:
                if fk in allowed:
                    payload[fk] = id_map["ledgers"].get(old_fk)
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["items"][old_id] = new_id

    # Item units
    for r in _rows(models.ItemUnit.__tablename__):
        t = _table(models.ItemUnit.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        old_item = _safe_int(r.get("item_id"))
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(old_item or -1)
        if payload.get("item_id") is None and "item_id" in allowed:
            continue
        db.execute(insert(t), [payload])

    # Warehouses
    for r in _rows(models.Warehouse.__tablename__):
        t = _table(models.Warehouse.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["warehouses"][old_id] = new_id

    # Vouchers
    for r in _rows(models.Voucher.__tablename__):
        t = _table(models.Voucher.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        pm_old = _safe_int(r.get("payment_mode_id"))
        if pm_old is not None:
            if "payment_mode_id" in allowed:
                payload["payment_mode_id"] = id_map["payment_modes"].get(pm_old)
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["vouchers"][old_id] = new_id

    # Purchase bills
    for r in _rows(models.PurchaseBill.__tablename__):
        t = _table(models.PurchaseBill.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "supplier_id" in allowed:
            payload["supplier_id"] = id_map["suppliers"].get(_safe_int(r.get("supplier_id")) or -1)
        for fk in ("purchase_ledger_id", "input_tax_ledger_id"):
            if fk in allowed:
                old_fk = _safe_int(r.get(fk))
                if old_fk is not None:
                    payload[fk] = id_map["ledgers"].get(old_fk)
        if "voucher_id" in allowed:
            old_v = _safe_int(r.get("voucher_id"))
            if old_v is not None:
                payload["voucher_id"] = id_map["vouchers"].get(old_v)
        if payload.get("supplier_id") is None and "supplier_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["purchase_bills"][old_id] = new_id

    for r in _rows(models.PurchaseBillLine.__tablename__):
        t = _table(models.PurchaseBillLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "bill_id" in allowed:
            payload["bill_id"] = id_map["purchase_bills"].get(_safe_int(r.get("bill_id")) or -1)
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
        if "warehouse_id" in allowed:
            wh_old = _safe_int(r.get("warehouse_id"))
            if wh_old is not None:
                payload["warehouse_id"] = id_map["warehouses"].get(wh_old)
        if payload.get("bill_id") is None or payload.get("item_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Sales invoices
    for r in _rows(models.SalesInvoice.__tablename__):
        t = _table(models.SalesInvoice.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "customer_id" in allowed:
            payload["customer_id"] = id_map["customers"].get(_safe_int(r.get("customer_id")) or -1)
        for fk in ("sales_ledger_id", "output_tax_ledger_id"):
            if fk in allowed:
                old_fk = _safe_int(r.get(fk))
                if old_fk is not None:
                    payload[fk] = id_map["ledgers"].get(old_fk)
        if "voucher_id" in allowed:
            old_v = _safe_int(r.get("voucher_id"))
            if old_v is not None:
                payload["voucher_id"] = id_map["vouchers"].get(old_v)
        if payload.get("customer_id") is None and "customer_id" in allowed:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["sales_invoices"][old_id] = new_id

    for r in _rows(models.SalesInvoiceLine.__tablename__):
        t = _table(models.SalesInvoiceLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "invoice_id" in allowed:
            payload["invoice_id"] = id_map["sales_invoices"].get(_safe_int(r.get("invoice_id")) or -1)
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
        if "warehouse_id" in allowed:
            wh_old = _safe_int(r.get("warehouse_id"))
            if wh_old is not None:
                payload["warehouse_id"] = id_map["warehouses"].get(wh_old)
        if payload.get("invoice_id") is None or payload.get("item_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Voucher lines
    for r in _rows(models.VoucherLine.__tablename__):
        t = _table(models.VoucherLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "voucher_id" in allowed:
            payload["voucher_id"] = id_map["vouchers"].get(_safe_int(r.get("voucher_id")) or -1)
        if "ledger_id" in allowed:
            payload["ledger_id"] = id_map["ledgers"].get(_safe_int(r.get("ledger_id")) or -1)
        if "department_id" in allowed:
            payload["department_id"] = None
        if "project_id" in allowed:
            payload["project_id"] = None
        if payload.get("voucher_id") is None or payload.get("ledger_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Voucher allocations
    for r in _rows(models.VoucherAllocation.__tablename__):
        t = _table(models.VoucherAllocation.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "voucher_id" in allowed:
            payload["voucher_id"] = id_map["vouchers"].get(_safe_int(r.get("voucher_id")) or -1)
        if "party_ledger_id" in allowed:
            payload["party_ledger_id"] = id_map["ledgers"].get(_safe_int(r.get("party_ledger_id")) or -1)
        if payload.get("voucher_id") is None or payload.get("party_ledger_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Voucher logs
    for r in _rows(models.VoucherLog.__tablename__):
        t = _table(models.VoucherLog.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "tenant_id" in allowed:
            payload["tenant_id"] = int(tenant_id)
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "voucher_id" in allowed:
            payload["voucher_id"] = id_map["vouchers"].get(_safe_int(r.get("voucher_id")) or -1)
        if payload.get("voucher_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Stock transfers
    for r in _rows(models.StockTransfer.__tablename__):
        t = _table(models.StockTransfer.__tablename__)
        old_id = _safe_int(r.get("id"))
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "company_id" in allowed:
            payload["company_id"] = new_company_id
        if "from_warehouse_id" in allowed:
            payload["from_warehouse_id"] = id_map["warehouses"].get(_safe_int(r.get("from_warehouse_id")) or -1)
        if "to_warehouse_id" in allowed:
            payload["to_warehouse_id"] = id_map["warehouses"].get(_safe_int(r.get("to_warehouse_id")) or -1)
        if payload.get("from_warehouse_id") is None or payload.get("to_warehouse_id") is None:
            continue
        ins = insert(t)
        if "id" in t.c:
            ins = ins.returning(t.c.id)
            res = db.execute(ins, [payload])
            new_id = int(res.scalar_one())
        else:
            db.execute(ins, [payload])
            new_id = int(old_id or 0)
        if old_id is not None:
            id_map["stock_transfers"][old_id] = new_id

    for r in _rows(models.StockTransferLine.__tablename__):
        t = _table(models.StockTransferLine.__tablename__)
        item = dict(r)
        item.pop("id", None)
        allowed = set(t.c.keys())
        payload = {k: v for k, v in item.items() if k in allowed}
        if "transfer_id" in allowed:
            payload["transfer_id"] = id_map["stock_transfers"].get(_safe_int(r.get("transfer_id")) or -1)
        if "item_id" in allowed:
            payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
        if payload.get("transfer_id") is None or payload.get("item_id") is None:
            continue
        db.execute(insert(t), [payload])

    # Stock movements/ledger: these are usually derived, but restore them as-is with remapped ids.
    for cls in (models.StockMovement, models.StockLedger):
        t = _table(cls.__tablename__)
        for r in _rows(cls.__tablename__):
            item = dict(r)
            item.pop("id", None)
            allowed = set(t.c.keys())
            payload = {k: v for k, v in item.items() if k in allowed}
            if "company_id" in allowed:
                payload["company_id"] = new_company_id
            if "warehouse_id" in allowed:
                payload["warehouse_id"] = id_map["warehouses"].get(_safe_int(r.get("warehouse_id")) or -1)
            if "item_id" in allowed:
                payload["item_id"] = id_map["items"].get(_safe_int(r.get("item_id")) or -1)
            if payload.get("warehouse_id") is None or payload.get("item_id") is None:
                continue
            db.execute(insert(t), [payload])

    db.add(
        models.AuditLog(
            user_id=int(actor_user_id),
            tenant_id=int(tenant_id),
            action="company_restore_new",
            message=f"Restored new company_id={new_company_id} from backup (safe restore).",
        )
    )
    return new_company_id


def _export_company_table(db: Session, model_cls: type[models.Base], *, company_id: int) -> list[dict]:
    table = _reflect_table(db, str(model_cls.__tablename__))
    if "company_id" not in table.c:
        return []
    stmt = table.select().where(table.c.company_id == int(company_id))
    rows = db.execute(stmt).mappings().all()
    return jsonable_encoder(list(rows))


def _import_company_table(
    db: Session,
    model_cls: type[models.Base],
    *,
    rows: list[dict],
    company_id: int,
) -> None:
    if not rows:
        return
    table = _reflect_table(db, str(model_cls.__tablename__))
    normalized: list[dict] = []
    allowed_cols = set(table.c.keys())
    for r in rows:
        item = {k: v for k, v in dict(r).items() if k in allowed_cols}
        if "company_id" in allowed_cols:
            item["company_id"] = int(company_id)
        normalized.append(item)
    db.execute(insert(table), normalized)


def _delete_company_table(db: Session, model_cls: type[models.Base], *, company_id: int) -> None:
    table = _reflect_table(db, str(model_cls.__tablename__))

    # Special handling: the companies table has multiple FK columns pointing at ledgers.
    # We must NULL those out before deleting ledgers.
    if str(table.name) == "ledgers":
        companies = _reflect_table(db, "companies")
        if "id" in companies.c and "company_id" not in companies.c:
            # Identify all columns that look like ledger FKs and exist in this DB schema.
            ledger_fk_cols = [c for c in companies.c.keys() if c.endswith("_ledger_id")]
            if ledger_fk_cols:
                values = {c: None for c in ledger_fk_cols}
                db.execute(
                    update(companies)
                    .where(companies.c.id == int(company_id))
                    .values(**values)
                )

    if "company_id" in table.c:
        db.execute(delete(table).where(table.c.company_id == int(company_id)))
        return

    # Child tables that don't carry company_id must be deleted through their parent.
    child_parent_map: dict[str, tuple[str, str, str]] = {
        # child_table: (child_fk_col, parent_table, parent_pk_col)
        "sales_invoice_lines": ("invoice_id", "sales_invoices", "id"),
        "purchase_bill_lines": ("bill_id", "purchase_bills", "id"),
        "voucher_lines": ("voucher_id", "vouchers", "id"),
        "stock_transfer_lines": ("transfer_id", "stock_transfers", "id"),
        "sales_return_lines": ("return_id", "sales_returns", "id"),
        "purchase_return_lines": ("return_id", "purchase_returns", "id"),
    }

    mapping = child_parent_map.get(str(table.name))
    if not mapping:
        return

    child_fk_col, parent_table_name, parent_pk_col = mapping
    if child_fk_col not in table.c:
        return

    parent = _reflect_table(db, parent_table_name)
    if "company_id" not in parent.c or parent_pk_col not in parent.c:
        return

    subq = parent.select().with_only_columns(parent.c[parent_pk_col]).where(parent.c.company_id == int(company_id))
    db.execute(delete(table).where(table.c[child_fk_col].in_(subq)))


def _ensure_company_belongs_to_tenant(db: Session, *, tenant_id: int, company_id: int) -> models.Company:
    company = (
        db.query(models.Company)
        .filter(models.Company.id == int(company_id), models.Company.tenant_id == int(tenant_id))
        .first()
    )
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company


def _reset_postgres_sequence_for_table(db: Session, *, table_name: str, pk_col: str = "id") -> None:
    # If the PK column is backed by a sequence (serial/bigserial), move it to MAX(id)
    # so future inserts don't collide.
    seq = db.execute(
        text("SELECT pg_get_serial_sequence(:t, :c)"),
        {"t": table_name, "c": pk_col},
    ).scalar()
    if not seq:
        return

    db.execute(
        text(
            "SELECT setval(:seq, GREATEST(COALESCE((SELECT MAX(id) FROM "
            + table_name
            + "), 1), 1), true)"
        ),
        {"seq": seq},
    )


def _reset_postgres_sequences(db: Session) -> None:
    for model_cls in _company_backup_tables():
        try:
            table = _reflect_table(db, str(model_cls.__tablename__))
        except Exception:
            continue
        if "id" not in table.c:
            continue
        if not bool(getattr(table.c.id, "primary_key", False)):
            continue
        _reset_postgres_sequence_for_table(db, table_name=str(table.name), pk_col="id")


def _parse_xml_backup(raw: bytes) -> dict:
    try:
        root = ET.fromstring(raw)
        if root.tag != "Backup":
            raise ValueError("Root tag is not Backup")
        
        backup = {
            "format_version": int(root.get("format_version") or 0),
            "exported_at": root.get("exported_at"),
            "tenant_id": int(root.get("tenant_id") or 0),
            "company_id": int(root.get("company_id") or 0),
            "company": {},
            "tables": {},
        }

        comp_el = root.find("Company")
        if comp_el is not None:
            for child in comp_el:
                val = child.text
                if val == "None" or val == "":
                    backup["company"][child.tag] = None
                else:
                    backup["company"][child.tag] = val

        tables_el = root.find("Tables")
        if tables_el is not None:
            for table_el in tables_el:
                table_name = table_el.tag
                rows = []
                for row_el in table_el.findall("Row"):
                    row = {}
                    for cell in row_el:
                        val = cell.text
                        if val == "None" or val == "":
                            row[cell.tag] = None
                        else:
                            row[cell.tag] = val
                    rows.append(row)
                backup["tables"][table_name] = rows
        
        return backup
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid backup XML: {e}")


def _parse_excel_backup(raw: bytes) -> dict:
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")
    
    try:
        from openpyxl import load_workbook
        stream = io.BytesIO(raw)
        wb = load_workbook(stream, data_only=True)
        
        backup = {
            "format_version": 0,
            "exported_at": None,
            "tenant_id": 0,
            "company_id": 0,
            "company": {},
            "tables": {},
        }

        # Metadata
        if "Metadata" in wb.sheetnames:
            ws = wb["Metadata"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2: continue
                key, val = row[0], row[1]
                if key in ("format_version", "tenant_id", "company_id"):
                    backup[key] = int(val) if val is not None else 0
                else:
                    backup[key] = val

        # Company
        if "Company" in wb.sheetnames:
            ws = wb["Company"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                headers = rows[0]
                values = rows[1]
                for i, h in enumerate(headers):
                    if h: backup["company"][str(h)] = values[i]

        # Tables
        for sheet_name in wb.sheetnames:
            if sheet_name in ("Metadata", "Company"):
                continue
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            
            headers = rows[0]
            table_rows = []
            for r in rows[1:]:
                row_dict = {}
                for i, h in enumerate(headers):
                    if h: row_dict[str(h)] = r[i]
                table_rows.append(row_dict)
            
            # Since sheet names were truncated to 31 chars, we need to match back to table names
            matched_table = None
            for model_cls in _company_backup_tables():
                tname = model_cls.__tablename__
                if tname.startswith(sheet_name) or sheet_name.startswith(tname[:31]):
                    matched_table = tname
                    break
            
            if matched_table:
                backup["tables"][matched_table] = table_rows
            else:
                backup["tables"][sheet_name] = table_rows

        return backup
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid backup Excel: {e}")


def _get_table_headers(table_name: str, with_indicators: bool = False) -> list[str]:
    for model_cls in _company_backup_tables():
        if str(model_cls.__tablename__) == table_name:
            headers = []
            for c in model_cls.__table__.columns:
                name = c.name
                if with_indicators:
                    # Mark nullable=False as compulsory with *
                    if not c.nullable and not c.primary_key:
                        name += "*"
                headers.append(name)
            return headers
    return []


def _generate_sample_row(table_name: str) -> dict:
    for model_cls in _company_backup_tables():
        if str(model_cls.__tablename__) == table_name:
            row = {}
            for c in model_cls.__table__.columns:
                name = c.name
                # Mark indicators in keys if we want to match _get_table_headers(with_indicators=True)
                # But headers in csv are just strings.
                
                # Check column type
                from sqlalchemy import String, Integer, Numeric, Date, DateTime, Boolean, Enum
                t = c.type
                val = None
                if isinstance(t, String): val = "Sample Text"
                elif isinstance(t, Integer): val = 1
                elif isinstance(t, Numeric): val = 100.00
                elif isinstance(t, Date): val = datetime.now().date().isoformat()
                elif isinstance(t, DateTime): val = datetime.now().isoformat()
                elif isinstance(t, Boolean): val = True
                elif isinstance(t, Enum): 
                    val = t.enums[0] if t.enums else None
                
                if name == "company_id": val = 0
                if name == "tenant_id": val = 0
                
                row[name] = val
            return row
    return {}


def _clean_csv_row(raw_row: dict) -> dict:
    clean = {}
    for k, v in raw_row.items():
        if k is None: continue
        # Strip indicator suffix like '*'
        clean_k = k.strip().rstrip('*')
        if v == "" or v is None:
            clean[clean_k] = None
        else:
            clean[clean_k] = v
    return clean


def _parse_csv_backup(raw: bytes, table_name: str = None) -> dict:
    try:
        content = raw.decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        rows = [_clean_csv_row(r) for r in reader]
        return {
            "format_version": BACKUP_FORMAT_VERSION,
            "tables": {table_name: rows} if table_name else {"unknown": rows}
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid CSV: {e}")


def _parse_zip_backup(raw: bytes) -> dict:
    try:
        stream = io.BytesIO(raw)
        with zipfile.ZipFile(stream, 'r') as zf:
            backup = {
                "format_version": BACKUP_FORMAT_VERSION,
                "tables": {},
                "company": {}
            }
            for name in zf.namelist():
                if name.endswith(".csv"):
                    # Remove dir prefix if any
                    base_name = name.split("/")[-1]
                    table_name = base_name[:-4]
                    with zf.open(name) as f:
                        content = f.read().decode("utf-8")
                        reader = csv.DictReader(io.StringIO(content))
                        rows = list(reader)
                        for r in rows:
                            for k, v in r.items():
                                if v == "": r[k] = None
                        backup["tables"][table_name] = rows
                elif name == "metadata.json":
                     with zf.open(name) as f:
                        meta = json.load(f)
                        backup.update(meta)
            return backup
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid ZIP backup: {e}")


async def _get_backup_payload_from_file(file: UploadFile) -> dict:
    raw = await file.read()
    if raw is None or len(raw) == 0:
        raise HTTPException(status_code=400, detail="Empty backup file")
    if len(raw) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Backup file too large")

    filename = (file.filename or "").lower()
    
    # Try XML
    if filename.endswith(".xml") or (file.content_type and "xml" in file.content_type):
        return _parse_xml_backup(raw)
    
    # Try Excel
    if filename.endswith((".xlsx", ".xls")) or (file.content_type and ("spreadsheet" in file.content_type or "excel" in file.content_type)):
        return _parse_excel_backup(raw)

    # Try ZIP
    if filename.endswith(".zip") or (file.content_type and "zip" in file.content_type):
        return _parse_zip_backup(raw)

    # Try CSV
    if filename.endswith(".csv") or (file.content_type and "csv" in file.content_type):
        table_name = filename[:-4] if "." in filename else "unknown"
        return _parse_csv_backup(raw, table_name)

    # Fallback to JSON
    try:
        return json.loads(raw.decode("utf-8"))
    except Exception:
        # If filename says XML or Excel but JSON parse failed, it was probably intended to be that
        if filename.endswith(".xml"): return _parse_xml_backup(raw)
        if filename.endswith((".xlsx", ".xls")): return _parse_excel_backup(raw)
        raise HTTPException(status_code=400, detail="Invalid backup JSON")


@router.get("", response_model=List[schemas.TenantRead])
def list_tenants(
    q: Optional[str] = Query(None, description="Search by tenant name"),
    status: Optional[str] = Query(None),
    plan: Optional[str] = Query(None),
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    query = db.query(models.Tenant).options(
        selectinload(models.Tenant.companies),
        selectinload(models.Tenant.business_type),
        selectinload(models.Tenant.users)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(models.Tenant.name.ilike(like))

    if status:
        query = query.filter(models.Tenant.status == status)

    if plan:
        query = query.filter(models.Tenant.plan == plan)

    tenants = (
        query.order_by(models.Tenant.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return tenants


@router.get("/{tenant_id}", response_model=schemas.TenantRead)
def get_tenant(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    tenant = (
        db.query(models.Tenant)
        .options(
            selectinload(models.Tenant.business_type),
            selectinload(models.Tenant.companies),
            selectinload(models.Tenant.users)
        )
        .get(tenant_id)
    )
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    # Explicitly build response so `companies` is serialized correctly
    companies = db.query(models.Company).filter(models.Company.tenant_id == tenant_id).order_by(models.Company.id).all()
    result = schemas.TenantRead.model_validate(tenant)
    result.companies = [schemas.TenantCompanyBrief(id=c.id, name=c.name) for c in companies]
    return result


@router.post("", response_model=schemas.TenantRead, status_code=201)
def create_tenant(
    tenant_in: schemas.TenantCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    existing = (
        db.query(models.Tenant)
        .filter(models.Tenant.name == tenant_in.name)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Tenant name already exists")

    tenant = models.Tenant(
        name=tenant_in.name,
        plan=tenant_in.plan,
        status="active",
        inventory_valuation_method=tenant_in.inventory_valuation_method,
        expires_at=tenant_in.expires_at,
        business_type_id=tenant_in.business_type_id,
    )

    # menu_template_id is now mandatory in schemas.TenantCreate
    template = db.query(models.MenuTemplate).get(int(tenant_in.menu_template_id))
    if not template or not bool(getattr(template, "is_active", True)):
        raise HTTPException(status_code=400, detail="Invalid menu_template_id")
    ensure_menu_template_assignable_to_tenant(template)

    tenant.menu_template_id = int(tenant_in.menu_template_id)
    # Note: We no longer automatically inject DASHBOARD into the template
    # to maintain strict template-based isolation.

    db.add(tenant)
    db.commit()
    db.refresh(tenant)
    return tenant


@router.put("/{tenant_id}", response_model=schemas.TenantRead)
def update_tenant(
    tenant_id: int,
    tenant_in: schemas.TenantUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    if tenant_in.name is not None and tenant_in.name != tenant.name:
        existing = db.query(models.Tenant).filter(models.Tenant.name == tenant_in.name, models.Tenant.id != tenant_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Tenant name already exists")
        tenant.name = tenant_in.name
    if tenant_in.plan is not None:
        tenant.plan = tenant_in.plan
    if tenant_in.status is not None:
        tenant.status = tenant_in.status
    if tenant_in.expires_at is not None:
        tenant.expires_at = tenant_in.expires_at

    if tenant_in.inventory_valuation_method is not None:
        tenant.inventory_valuation_method = tenant_in.inventory_valuation_method

    if tenant_in.business_type_id is not None:
        tenant.business_type_id = tenant_in.business_type_id

    data = tenant_in.model_dump(exclude_unset=True)

    if "menu_template_id" in data:
        new_template_id = data["menu_template_id"]
        if new_template_id is not None:
            template = db.query(models.MenuTemplate).get(int(new_template_id))
            if not template or not bool(getattr(template, "is_active", True)):
                raise HTTPException(status_code=400, detail="Invalid menu_template_id")
            ensure_menu_template_assignable_to_tenant(template)

        tenant.menu_template_id = new_template_id
        # Note: We no longer automatically inject menus into the template on tenant update
        # to ensure that Superadmin configurations are strictly followed.

    db.commit()
    db.refresh(tenant)
    return tenant


@router.get("/{tenant_id}/document-scan-policy", response_model=schemas.TenantDocumentScanPolicyRead)
def get_tenant_document_scan_policy(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
    return schemas.TenantDocumentScanPolicyRead(
        tenant_id=int(tenant.id),
        document_scan_enabled=bool(getattr(tenant, "document_scan_enabled", True)),
        daily_document_scan_limit=getattr(tenant, "daily_document_scan_limit", None),
    )


@router.put("/{tenant_id}/document-scan-policy", response_model=schemas.TenantDocumentScanPolicyRead)
def update_tenant_document_scan_policy(
    tenant_id: int,
    payload: schemas.TenantDocumentScanPolicyUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    if payload.document_scan_enabled is not None:
        tenant.document_scan_enabled = bool(payload.document_scan_enabled)

    if payload.daily_document_scan_limit is not None:
        if int(payload.daily_document_scan_limit) < 0:
            raise HTTPException(status_code=400, detail="daily_document_scan_limit cannot be negative")
        tenant.daily_document_scan_limit = int(payload.daily_document_scan_limit)

    db.add(tenant)
    db.commit()
    db.refresh(tenant)

    return schemas.TenantDocumentScanPolicyRead(
        tenant_id=int(tenant.id),
        document_scan_enabled=bool(getattr(tenant, "document_scan_enabled", True)),
        daily_document_scan_limit=getattr(tenant, "daily_document_scan_limit", None),
    )


@router.get("/ghost/document-scan-usage", response_model=List[schemas.TenantDocumentScanUsageRow])
def ghost_document_scan_usage(
    tenant_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)

    now = datetime.utcnow()
    start_utc = datetime(now.year, now.month, now.day)
    end_utc = start_utc + timedelta(days=1)

    q = db.query(models.Tenant)
    if tenant_id is not None:
        q = q.filter(models.Tenant.id == int(tenant_id))
    tenants = q.order_by(models.Tenant.name.asc()).all()

    rows: list[schemas.TenantDocumentScanUsageRow] = []
    for t in tenants:
        used = (
            db.query(func.count(models.Document.id))
            .join(models.Company, models.Company.id == models.Document.company_id)
            .filter(
                models.Company.tenant_id == int(t.id),
                models.Document.created_at >= start_utc,
                models.Document.created_at < end_utc,
                models.Document.status.in_(
                    [models.DocumentStatus.processed, models.DocumentStatus.confirmed]
                ),
            )
            .scalar()
        )
        used_int = int(used or 0)
        limit = getattr(t, "daily_document_scan_limit", None)

        remaining: int | None = None
        if limit is not None:
            remaining = max(int(limit) - used_int, 0)

        rows.append(
            schemas.TenantDocumentScanUsageRow(
                tenant_id=int(t.id),
                tenant_name=str(t.name),
                document_scan_enabled=bool(getattr(t, "document_scan_enabled", True)),
                daily_document_scan_limit=limit,
                scans_used_today=used_int,
                scans_remaining_today=remaining,
            )
        )

    return rows


@router.delete("/{tenant_id}", status_code=204)
def delete_tenant(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_superadmin(current_user)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")
        
    # Safeguard: Ensure no companies or users exist
    if len(tenant.companies) > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete tenant '{tenant.name}' because it still has {len(tenant.companies)} active company(s). Please delete all companies first."
        )
        
    if len(tenant.users) > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete tenant '{tenant.name}' because it still has {len(tenant.users)} active user account(s). Please remove all users first."
        )

    db.delete(tenant)
    db.commit()
    return


@router.get("/{tenant_id}/companies", response_model=List[schemas.TenantCompanySummary])
def list_tenant_companies(
    tenant_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)
    tenant = db.query(models.Tenant).get(tenant_id)
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found")

    companies = (
        db.query(models.Company)
        .filter(models.Company.tenant_id == tenant_id)
        .order_by(models.Company.name.asc())
        .all()
    )
    return companies


@router.get("/{tenant_id}/companies/{company_id}/backup")
def backup_company_data(
    tenant_id: int,
    company_id: int,
    format: str = Query("json", regex="^(json|xml|excel|csv)$"),
    is_sample: bool = Query(False),
    tables: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)
    company = _ensure_company_belongs_to_tenant(db, tenant_id=tenant_id, company_id=company_id)

    exported: dict[str, list[dict]] = {}
    for model_cls in _company_backup_tables():
        tname = str(model_cls.__tablename__)
        if tables and tname not in tables:
            continue
        exported[tname] = _export_company_table(db, model_cls, company_id=company.id)

    company_info = {
        "id": company.id,
        "tenant_id": company.tenant_id,
        "name": company.name,
        "fiscal_year_start": company.fiscal_year_start,
        "fiscal_year_end": company.fiscal_year_end,
        "address": company.address,
        "phone": company.phone,
        "pan_number": company.pan_number,
        "business_type": company.business_type,
        "country": company.country,
        "currency": company.currency,
        "inventory_valuation_method": company.inventory_valuation_method,
        "cost_center_mode": company.cost_center_mode,
        "cost_center_single_dimension": company.cost_center_single_dimension,
    }

    payload = {
        "format_version": BACKUP_FORMAT_VERSION,
        "exported_at": datetime.utcnow().isoformat(),
        "tenant_id": int(tenant_id),
        "company_id": int(company_id),
        "company": jsonable_encoder(company_info),
        "tables": exported,
    }

    db.add(
        models.AuditLog(
            user_id=int(current_user.id),
            tenant_id=int(tenant_id),
            action="company_backup",
            message=f"Backed up company_id={company_id} in {format} format (sample={is_sample}).",
        )
    )
    db.commit()

    if format == "csv":
        if len(exported) > 1:
            stream = io.BytesIO()
            with zipfile.ZipFile(stream, 'w') as zf:
                zf.writestr("metadata.json", json.dumps({
                    "format_version": BACKUP_FORMAT_VERSION,
                    "exported_at": payload["exported_at"],
                    "tenant_id": tenant_id,
                    "company_id": company_id,
                    "company": payload["company"]
                }))
                for tname, rows in exported.items():
                    output = io.StringIO()
                    headers = _get_table_headers(tname, with_indicators=is_sample)
                    if not headers: continue
                    writer = csv.DictWriter(output, fieldnames=headers)
                    writer.writeheader()
                    if is_sample:
                        raw_sample = _generate_sample_row(tname)
                        # Map raw keys to headers with indicators
                        decorated_sample = {}
                        for k, v in raw_sample.items():
                            match = k + "*" if k + "*" in headers else k
                            decorated_sample[match] = v
                        writer.writerow(decorated_sample)
                    else:
                        writer.writerows(rows)
                    zf.writestr(f"{tname}.csv", output.getvalue())
            stream.seek(0)
            suffix = "_sample" if is_sample else ""
            filename = f"tenant_{tenant_id}_company_{company_id}_backup{suffix}.zip"
            return StreamingResponse(
                stream,
                media_type="application/zip",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        elif len(exported) == 1:
            tname = list(exported.keys())[0]
            rows = exported[tname]
            output = io.StringIO()
            headers = _get_table_headers(tname, with_indicators=is_sample)
            if headers:
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                if is_sample:
                    raw_sample = _generate_sample_row(tname)
                    decorated_sample = {}
                    for k, v in raw_sample.items():
                        match = k + "*" if k + "*" in headers else k
                        decorated_sample[match] = v
                    writer.writerow(decorated_sample)
                else:
                    writer.writerows(rows)
            content = output.getvalue().encode("utf-8")
            suffix = "_sample" if is_sample else ""
            filename = f"{tname}{suffix}.csv"
            return Response(
                content=content,
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        else:
            raise HTTPException(status_code=400, detail="No tables selected for CSV export")

    if format == "xml":
        root = ET.Element("Backup")
        root.set("format_version", str(BACKUP_FORMAT_VERSION))
        root.set("exported_at", payload["exported_at"])
        root.set("tenant_id", str(tenant_id))
        root.set("company_id", str(company_id))

        comp_el = ET.SubElement(root, "Company")
        for k, v in payload["company"].items():
            child = ET.SubElement(comp_el, k)
            child.text = str(v) if v is not None else ""

        tables_el = ET.SubElement(root, "Tables")
        for table_name, rows in exported.items():
            table_el = ET.SubElement(tables_el, table_name)
            if is_sample:
                # Use indicators in tags? Usually XML tags don't have * but for sample clarity it might help.
                # However, XML tags characters are restricted. Let's just use clean names for tags but add sample row.
                sample_row = _generate_sample_row(table_name)
                if sample_row:
                    row_el = ET.SubElement(table_el, "Row")
                    for k, v in sample_row.items():
                        cell = ET.SubElement(row_el, k)
                        cell.text = str(v) if v is not None else ""
            else:
                for row in rows:
                    row_el = ET.SubElement(table_el, "Row")
                    for k, v in row.items():
                        cell = ET.SubElement(row_el, k)
                        cell.text = str(v) if v is not None else ""

        content = ET.tostring(root, encoding="utf-8")
        filename = f"tenant_{tenant_id}_company_{company_id}_backup.xml"
        return Response(
            content=content,
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")
        
        wb = Workbook()
        # Metadata sheet
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        ws_meta.append(["Key", "Value"])
        ws_meta.append(["format_version", BACKUP_FORMAT_VERSION])
        ws_meta.append(["exported_at", payload["exported_at"]])
        ws_meta.append(["tenant_id", tenant_id])
        ws_meta.append(["company_id", company_id])
        
        # Company sheet
        ws_comp = wb.create_sheet(title="Company")
        comp_data = payload["company"]
        ws_comp.append(list(comp_data.keys()))
        ws_comp.append(list(comp_data.values()))

        for table_name, rows in exported.items():
            ws = wb.create_sheet(title=table_name[:31])
            headers = _get_table_headers(table_name, with_indicators=is_sample)
            if headers:
                ws.append(headers)
                if is_sample:
                    raw_sample = _generate_sample_row(table_name)
                    # Use base headers to ensure value ordering matches model columns
                    base_headers = _get_table_headers(table_name, with_indicators=False)
                    ws.append([raw_sample.get(h) for h in base_headers])
                else:
                    for r in rows:
                        ws.append([r.get(h) for h in headers])

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        suffix = "_sample" if is_sample else ""
        filename = f"tenant_{tenant_id}_company_{company_id}_backup{suffix}.xlsx"
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Default JSON
    if is_sample:
        new_tables = {}
        for t in payload["tables"]:
            new_tables[t] = [_generate_sample_row(t)]
        payload["tables"] = new_tables
    content = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    suffix = "_sample" if is_sample else ""
    filename = f"tenant_{tenant_id}_company_{company_id}_backup{suffix}.json"

    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{tenant_id}/companies/{company_id}/restore")
async def restore_company_data_overwrite(
    tenant_id: int,
    company_id: int,
    file: UploadFile = File(...),
    confirm_overwrite: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)
    if not confirm_overwrite:
        raise HTTPException(
            status_code=400,
            detail="confirm_overwrite=true is required to restore into an existing company",
        )

    _ensure_company_belongs_to_tenant(db, tenant_id=tenant_id, company_id=company_id)

    backup = await _get_backup_payload_from_file(file)

    if int(backup.get("format_version") or 0) != BACKUP_FORMAT_VERSION:
        raise HTTPException(status_code=400, detail="Unsupported backup format_version")

    if int(backup.get("tenant_id") or 0) != int(tenant_id):
        raise HTTPException(status_code=400, detail="Backup tenant_id does not match")

    if int(backup.get("company_id") or 0) != int(company_id):
        raise HTTPException(status_code=400, detail="Backup company_id does not match target company")

    tables = backup.get("tables") or {}
    if not isinstance(tables, dict):
        raise HTTPException(status_code=400, detail="Invalid backup tables")

    try:
        for model_cls in _company_backup_delete_tables():
            try:
                _delete_company_table(db, model_cls, company_id=company_id)
            except Exception as e:
                logger.exception(
                    "Overwrite restore delete failed",
                    extra={
                        "table": str(model_cls.__tablename__),
                        "tenant_id": int(tenant_id),
                        "company_id": int(company_id),
                    },
                )
                raise HTTPException(
                    status_code=500,
                    detail="Restore failed during cleanup. Please try again or contact support.",
                )

        company_data = backup.get("company") or {}
        if isinstance(company_data, dict):
            company = db.query(models.Company).get(int(company_id))
            if not company:
                raise HTTPException(status_code=404, detail="Company not found")
            for k, v in company_data.items():
                if k in ("id", "tenant_id", "owner_id", "created_at", "updated_at"):
                    continue
                if hasattr(company, k):
                    setattr(company, k, v)
            db.add(company)

        for model_cls in _company_backup_tables():
            table_name = model_cls.__tablename__
            table_rows = tables.get(table_name) or []
            if not isinstance(table_rows, list):
                raise HTTPException(status_code=400, detail=f"Invalid rows for table: {table_name}")
            try:
                _import_company_table(db, model_cls, rows=table_rows, company_id=company_id)
            except Exception as e:
                logger.exception(
                    "Overwrite restore insert failed",
                    extra={
                        "table": str(table_name),
                        "tenant_id": int(tenant_id),
                        "company_id": int(company_id),
                    },
                )
                raise HTTPException(
                    status_code=500,
                    detail="Restore failed during data import. Please try again or contact support.",
                )

        _reset_postgres_sequences(db)

        db.add(
            models.AuditLog(
                user_id=int(current_user.id),
                tenant_id=int(tenant_id),
                action="company_restore_overwrite",
                message=f"Restored company_id={company_id} from backup (overwrite).",
            )
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(
            "Overwrite restore failed with unhandled error: %s\n%s",
            str(e),
            traceback.format_exc(),
        )
        raise HTTPException(status_code=500, detail="Restore failed. Please try again or contact support.")

    return {"status": "ok", "company_id": int(company_id)}


@router.post("/{tenant_id}/companies/restore-new")
async def restore_company_data_safe_new_company(
    tenant_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin),
):
    _ensure_same_tenant_or_superadmin(current_user, tenant_id=tenant_id)
    backup = await _get_backup_payload_from_file(file)

    if int(backup.get("format_version") or 0) != BACKUP_FORMAT_VERSION:
        raise HTTPException(status_code=400, detail="Unsupported backup format_version")

    if int(backup.get("tenant_id") or 0) != int(tenant_id):
        raise HTTPException(status_code=400, detail="Backup tenant_id does not match")

    try:
        new_company_id = _safe_restore_new_company(
            db,
            tenant_id=int(tenant_id),
            backup=backup,
            actor_user_id=int(current_user.id),
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("Safe restore failed for tenant_id=%s: %s", tenant_id, e)
        raise HTTPException(status_code=500, detail="Restore failed. Please try again or contact support.")

    return {"status": "ok", "company_id": int(new_company_id)}
