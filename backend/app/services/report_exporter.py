import io
from datetime import datetime
from typing import List, Any, Dict, Optional
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ReportExporter:
    @staticmethod
    def export_to_excel(
        company_name: str,
        report_title: str,
        period_text: str,
        headers: List[str],
        data: List[List[Any]],
        summary_data: Optional[List[Dict[str, Any]]] = None,
        total_row: Optional[List[Any]] = None,
        filename: str = "report.xlsx"
    ) -> StreamingResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.sheet_view.showGridLines = False

        # --- Styles ---
        # --- Styles ---
        title_font = Font(size=20, bold=True, color="0F172A") # Slate 900
        subtitle_font = Font(size=14, color="475569") # Slate 600
        period_font = Font(size=11, italic=True, color="64748B") # Slate 500
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
        
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        heading_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        total_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
        
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(style='thin', color="CBD5E1")
        thick_side = Side(style='medium', color="1E293B")
        double_side = Side(style='double', color="1E293B")
        
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_border = Border(bottom=thick_side)
        total_border = Border(top=thick_side, bottom=double_side)

        table_cols = len(headers)
        last_col = get_column_letter(table_cols)

        # 1. Header Section
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = company_name.upper()
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = report_title
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align
        
        ws.merge_cells(f'A3:{last_col}3')
        ws['A3'] = period_text
        ws['A3'].font = period_font
        ws['A3'].alignment = center_align

        curr_row = 5
        # 2. Summary Info Cards
        if summary_data:
            for item in summary_data:
                # Merge first few columns for label to prevent overflow, or just use the last few
                label_col = max(1, table_cols - 1)
                ws.merge_cells(f'A{curr_row}:{get_column_letter(label_col)}{curr_row}')
                
                label_cell = ws.cell(row=curr_row, column=1)
                label_cell.value = str(item.get("label")).upper()
                label_cell.font = Font(bold=True, size=10, color="64748B")
                label_cell.alignment = Alignment(horizontal="right", vertical="center")
                
                value_cell = ws.cell(row=curr_row, column=table_cols)
                value_cell.value = item.get("value")
                value_cell.font = Font(bold=True, size=11, color="4F46E5") # Use theme color
                
                if isinstance(item.get("value"), (int, float)):
                    value_cell.number_format = '#,##0.00'
                    value_cell.alignment = right_align
                else:
                    value_cell.alignment = right_align
                
                curr_row += 1
            ws.append([])
            curr_row += 1

        # 3. Table Headers
        header_row_idx = ws.max_row + 1
        for col_num, title in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # 4. Data Rows
        for idx, row_data in enumerate(data):
            # Extract plain values for appending
            extracted_values = []
            for val in row_data:
                if isinstance(val, dict):
                    extracted_values.append(val.get("value", ""))
                else:
                    extracted_values.append(val)
            
            ws.append(extracted_values)
            row_idx = ws.max_row
            is_zebra = (idx % 2 == 1)
            
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                
                # Default style
                is_bold = False
                indent_level = 0
                
                if isinstance(val, dict):
                    is_bold = val.get("is_bold", False)
                    indent_level = val.get("indent", 0)
                
                if is_bold:
                    cell.font = Font(bold=True, color="1E293B", size=11)
                    # For bold rows (headings), a light slate background
                    cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                elif is_zebra:
                    cell.fill = zebra_fill
                
                cell.border = border
                
                # Check value from extracted list for alignment
                actual_val = extracted_values[col_num-1]
                if isinstance(actual_val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent_level * 2 if indent_level > 0 else 0)

        # 5. Grand Total Row
        if total_row:
            ws.append([]) # Spacer before total
            ws.append(total_row)
            total_row_idx = ws.max_row
            for col_num, val in enumerate(total_row, 1):
                cell = ws.cell(row=total_row_idx, column=col_num)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = total_border
                cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "center", vertical="center")
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
                elif val == "GRAND TOTAL" or val == "TOTAL":
                    cell.alignment = left_align

        # 6. Page Setup
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Unlimited height (many pages)
        
        # Set explicit margins (in inches)
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        
        ws.oddFooter.center.text = "Page &P of &N"
        ws.oddFooter.right.text = "Generated on &D"

        # 7. Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                # Include summary rows and data rows for width calculation
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @staticmethod
    def export_to_html(
        company_name: str,
        report_title: str,
        period_text: str,
        headers: List[str],
        data: List[List[Any]],
        summary_data: Optional[List[Dict[str, Any]]] = None,
        total_row: Optional[List[Any]] = None,
        filename: str = "report.html"
    ) -> StreamingResponse:
        html = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Inter', 'Segoe UI', system-ui, sans-serif; color: #0f172a; padding: 40px; background-color: #f1f5f9; margin: 0; line-height: 1.5; }}
                .report-container {{ background: white; padding: 50px; border-radius: 24px; box-shadow: 0 20px 25px -5px rgb(0 0 0 / 0.1), 0 8px 10px -6px rgb(0 0 0 / 0.1); max-width: 1100px; margin: 0 auto; border: 1px solid #e2e8f0; }}
                .header {{ text-align: left; margin-bottom: 40px; border-left: 5px solid #4f46e5; padding-left: 25px; }}
                h1 {{ margin: 0; color: #0f172a; font-size: 28px; font-weight: 900; text-transform: uppercase; letter-spacing: -0.02em; }}
                h2 {{ margin: 5px 0; color: #4f46e5; font-size: 18px; font-weight: 700; }}
                p.period {{ color: #64748b; font-size: 13px; font-weight: 600; margin-top: 10px; }}
                
                .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 40px; }}
                .summary-card {{ background: #f8fafc; padding: 20px; border-radius: 16px; border: 1px solid #e2e8f0; border-top: 4px solid #4f46e5; }}
                .summary-label {{ font-size: 10px; font-weight: 800; color: #64748b; text-transform: uppercase; letter-spacing: 0.1em; display: block; margin-bottom: 5px; }}
                .summary-value {{ font-size: 20px; font-weight: 900; color: #1e293b; font-family: 'JetBrains Mono', monospace; }}
                
                table {{ border-collapse: separate; border-spacing: 0; width: 100%; border-radius: 12px; overflow: hidden; border: 1px solid #e2e8f0; }}
                th {{ background-color: #1e293b !important; color: white !important; font-weight: 700; text-align: left; padding: 15px 12px; font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; border: none; }}
                td {{ padding: 12px; border-bottom: 1px solid #f1f5f9; font-size: 12px; color: #334155; }}
                tr:last-child td {{ border-bottom: none; }}
                tr:nth-child(even) {{ background-color: #fbfcfe; }}
                tr.font-bold {{ background-color: #f8fafc; }}
                
                .text-right {{ text-align: right; white-space: nowrap; font-weight: 700; font-family: 'JetBrains Mono', monospace; color: #1e293b; }}
                .font-bold {{ font-weight: 800; color: #0f172a; }}
                .total-row {{ background-color: #4f46e5 !important; color: white !important; font-weight: 900; font-size: 13px; }}
                .total-row td {{ border: none !important; padding: 18px 12px; color: white !important; }}
                .total-row .text-right {{ color: white !important; }}
                
                .footer {{ margin-top: 40px; text-align: center; font-size: 11px; color: #94a3b8; font-weight: 600; border-top: 1px solid #e2e8f0; padding-top: 20px; }}
                
                @media print {{
                    @page {{ size: portrait; margin: 15mm 10mm; }}
                    body {{ background: white; padding: 0; }}
                    .report-container {{ box-shadow: none; padding: 0; width: 100%; max-width: 100%; }}
                    table {{ page-break-inside: auto; width: 100%; font-size: 9px; }}
                    tr {{ page-break-inside: avoid; page-break-after: auto; }}
                    thead {{ display: table-header-group; }}
                    tfoot {{ display: table-footer-group; }}
                    th, td {{ border: 1px solid #cbd5e1 !important; }}
                    .no-print {{ display: none; }}
                }}
            </style>
        </head>
        <body>
            <div class="report-container">
                <div class="header">
                    <h1>{company_name}</h1>
                    <h2>{report_title}</h2>
                    <p class="period">Period: {period_text}</p>
                </div>
                
                {f'''<div class="summary-grid">
                    {"".join([f'<div class="summary-card"><span class="summary-label">{s.get("label")}</span><span class="summary-value">{s.get("value") if not isinstance(s.get("value"), (int, float)) else f"{s.get("value"):,.2f}"}</span></div>' for s in summary_data])}
                </div>''' if summary_data else ""}
                <table>
                    <thead>
                        <tr>
        """
        for h in headers:
            html += f"<th>{h}</th>"
        html += "</tr></thead><tbody>"

        for row in data:
            is_bold_row = any(isinstance(v, dict) and v.get("is_bold") for v in row)
            row_class = "font-bold" if is_bold_row else ""
            html += f"<tr class='{row_class}'>"
            for val in row:
                cell_val = val
                indent = 0
                cls = ""
                
                if isinstance(val, dict):
                    cell_val = val.get("value")
                    indent = val.get("indent", 0)
                    if val.get("is_bold"): cls += " font-bold"
                
                if isinstance(cell_val, (int, float)):
                    cls += " text-right"
                    display_val = f"{cell_val:,.2f}"
                else:
                    display_val = str(cell_val)
                
                style = f"padding-left: {indent * 20 + 8}px;" if indent > 0 else ""
                html += f"<td class='{cls}' style='{style}'>{display_val}</td>"
            html += "</tr>"

        if total_row:
            html += "<tr class='total-row'>"
            for val in total_row:
                cls = "text-right" if isinstance(val, (int, float)) else ""
                display_val = f"{val:,.2f}" if isinstance(val, (int, float)) else str(val)
                html += f"<td class='{cls}'>{display_val}</td>"
            html += "</tr>"

        html += f"""
                </tbody>
            </table>
            <div class="footer">
                Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            </div>
        </div>
        </body>
        </html>
        """
        from fastapi.responses import Response
        return Response(content=html, media_type="text/html", headers={"Content-Disposition": f"attachment; filename={filename}"})
