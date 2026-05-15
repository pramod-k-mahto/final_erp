from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
import xml.etree.ElementTree as ET

from fastapi import HTTPException, UploadFile
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from . import models


@dataclass
class ParsedTable:
    columns: list[str]
    rows: list[dict[str, Any]]


class ImportEngine:
    def __init__(self, db: Session):
        self.db = db

    def _ensure_job_access(self, *, job: models.ImportJob, current_user: models.User) -> None:
        # Enforce tenant boundary.
        if current_user.role == models.UserRole.admin:
            if current_user.tenant_id is None or int(job.tenant_id) != int(current_user.tenant_id):
                raise HTTPException(status_code=403, detail="Cannot access import job outside your tenant")
        elif current_user.role == models.UserRole.TENANT:
            if current_user.tenant_id is None or int(job.tenant_id) != int(current_user.tenant_id):
                raise HTTPException(status_code=403, detail="Cannot access import job outside your tenant")
        elif current_user.role != models.UserRole.superadmin:
            raise HTTPException(status_code=403, detail="Not enough permissions")

    def create_job(
        self,
        *,
        tenant_id: int,
        company_id: int,
        source_type: str,
        data_type: str,
        created_by: int | None,
        current_user: models.User,
    ) -> models.ImportJob:
        # Ensure company exists and tenant boundary is correct.
        company = self.db.query(models.Company).filter(models.Company.id == int(company_id)).first()
        if company is None:
            raise HTTPException(status_code=404, detail="Company not found")

        if int(company.tenant_id) != int(tenant_id):
            raise HTTPException(status_code=400, detail="tenant_id does not match company tenant")

        if current_user.role != models.UserRole.superadmin:
            if current_user.tenant_id is None or int(current_user.tenant_id) != int(tenant_id):
                raise HTTPException(status_code=403, detail="Cannot create import job for another tenant")

        job = models.ImportJob(
            tenant_id=int(tenant_id),
            company_id=int(company_id),
            source_type=str(source_type),
            data_type=str(data_type),
            status="DRAFT",
            created_by=int(created_by) if created_by is not None else None,
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def get_job(self, *, job_id: int, current_user: models.User) -> models.ImportJob:
        job = self.db.query(models.ImportJob).filter(models.ImportJob.id == int(job_id)).first()
        if job is None:
            raise HTTPException(status_code=404, detail="Import job not found")
        self._ensure_job_access(job=job, current_user=current_user)
        return job

    def _base_upload_dir(self) -> Path:
        base = Path(__file__).resolve().parents[1] / "import_uploads"
        base.mkdir(parents=True, exist_ok=True)
        return base

    _MAX_IMPORT_FILE_BYTES = 20 * 1024 * 1024  # 20 MB

    def store_upload(self, *, job: models.ImportJob, upload: UploadFile) -> models.ImportFile:
        name = str(upload.filename or "upload")
        ext = name.split(".")[-1].lower() if "." in name else ""
        if ext not in {"csv", "xlsx", "xls", "json", "xml"}:
            raise HTTPException(status_code=400, detail="Unsupported file type")

        # Sanitize: strip path components to prevent directory traversal
        safe_name = name.replace("/", "_").replace("\\", "_").replace("..", "_")
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
        stored_name = f"job_{int(job.id)}_{ts}_{safe_name}"
        dest = self._base_upload_dir() / stored_name

        data = upload.file.read()
        if len(data) == 0:
            raise HTTPException(status_code=400, detail="Empty file")
        if len(data) > self._MAX_IMPORT_FILE_BYTES:
            raise HTTPException(status_code=413, detail="File too large. Max 20 MB.")
        with open(dest, "wb") as f:
            f.write(data)

        rec = models.ImportFile(
            import_job_id=int(job.id),
            filename=name,
            file_type=ext or "unknown",
            stored_path=str(dest),
            uploaded_at=datetime.utcnow(),
        )
        self.db.add(rec)
        self.db.flush()
        return rec

    def _read_json_records(self, path: str) -> ParsedTable:
        raw = Path(path).read_text(encoding="utf-8")
        payload = json.loads(raw)

        rows: list[dict[str, Any]] = []
        if isinstance(payload, list):
            for item in payload:
                if isinstance(item, dict):
                    rows.append(item)
                else:
                    rows.append({"value": item})
        elif isinstance(payload, dict):
            # Common export patterns: {"data": [...]} or single object
            if isinstance(payload.get("data"), list):
                for item in payload["data"]:
                    if isinstance(item, dict):
                        rows.append(item)
                    else:
                        rows.append({"value": item})
            else:
                rows.append(payload)
        else:
            rows.append({"value": payload})

        cols: set[str] = set()
        for r in rows:
            cols.update([str(k) for k in r.keys()])
        return ParsedTable(columns=sorted(cols, key=lambda x: x.casefold()), rows=rows)

    def _read_csv_records(self, path: str) -> ParsedTable:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = [dict(r) for r in reader]
            columns = list(reader.fieldnames or [])
        return ParsedTable(columns=[str(c) for c in columns], rows=rows)

    def _read_excel_records(self, path: str) -> ParsedTable:
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=500, detail="openpyxl is required for Excel import")

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active

        rows_iter: Iterable[tuple[Any, ...]] = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return ParsedTable(columns=[], rows=[])

        columns = [str(c).strip() if c is not None else "" for c in header]
        columns = [c if c else f"col_{i+1}" for i, c in enumerate(columns)]

        out_rows: list[dict[str, Any]] = []
        for row in rows_iter:
            record: dict[str, Any] = {}
            for i, col in enumerate(columns):
                val = row[i] if i < len(row) else None
                # Normalize datetime/date to ISO strings for JSONB staging
                if hasattr(val, "isoformat"):
                    try:
                        val = val.isoformat()
                    except Exception:
                        pass
                record[col] = val
            # skip fully empty rows
            if all(v in (None, "") for v in record.values()):
                continue
            out_rows.append(record)

        return ParsedTable(columns=columns, rows=out_rows)

    def _read_tally_xml_records(self, path: str) -> ParsedTable:
        """Best-effort Tally XML parser.

        Produces rows in two shapes:
        - Ledger rows: {"record_type":"LEDGER", "name":..., "parent":..., "opening_balance":..., "opening_balance_type":...}
        - Voucher rows: {"record_type":"VOUCHER", "voucher_type":..., "date":..., "voucher_number":..., "narration":..., "lines":[...]}

        This is intentionally tolerant: missing fields are left as None.
        """

        try:
            tree = ET.parse(path)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Invalid XML file: {exc}")

        root = tree.getroot()

        def _txt(node: ET.Element | None) -> str | None:
            if node is None or node.text is None:
                return None
            t = str(node.text).strip()
            return t or None

        def _find_first_text(parent: ET.Element, tags: list[str]) -> str | None:
            for tag in tags:
                el = parent.find(f".//{tag}")
                val = _txt(el)
                if val:
                    return val
            return None

        def _parse_amount(s: str | None) -> float | None:
            if not s:
                return None
            try:
                # Tally often stores amounts like -1234.00
                return float(str(s).replace(",", "").strip())
            except Exception:
                return None

        rows: list[dict[str, Any]] = []

        # Tally exports frequently have repeated collections under <TALLYMESSAGE>.
        # We'll scan for any element that looks like a LEDGER or VOUCHER.
        for ledger in root.findall(".//LEDGER"):
            name = _txt(ledger.find("NAME")) or _txt(ledger)
            parent = _find_first_text(ledger, ["PARENT", "GROUP", "PARENTNAME"])
            opening = _find_first_text(ledger, ["OPENINGBALANCE", "OPENINGBAL"])
            opening_f = _parse_amount(opening)
            ob_type = None
            if opening_f is not None:
                ob_type = "DEBIT" if opening_f >= 0 else "CREDIT"
            rows.append(
                {
                    "record_type": "LEDGER",
                    "name": name,
                    "parent": parent,
                    "opening_balance": abs(opening_f) if opening_f is not None else None,
                    "opening_balance_type": ob_type,
                }
            )

        for voucher in root.findall(".//VOUCHER"):
            vtype = _find_first_text(voucher, ["VOUCHERTYPENAME", "VOUCHERTYPE", "VOUCHERTYPENAME.LIST"])
            vdate = _find_first_text(voucher, ["DATE", "VOUCHERDATE"])
            vno = _find_first_text(voucher, ["VOUCHERNUMBER", "VOUCHERNUM", "REFERENCE"])
            narration = _find_first_text(voucher, ["NARRATION"])

            lines: list[dict[str, Any]] = []

            for entry in voucher.findall(".//ALLLEDGERENTRIES.LIST"):
                ledger_name = _find_first_text(entry, ["LEDGERNAME", "LEDGER" , "ACCOUNT"])
                amount_s = _find_first_text(entry, ["AMOUNT"])
                amt = _parse_amount(amount_s) or 0.0

                # Convention: negative amount => credit, positive => debit
                debit = float(amt) if amt > 0 else 0.0
                credit = float(-amt) if amt < 0 else 0.0

                lines.append(
                    {
                        "ledger_name": ledger_name,
                        "debit": debit,
                        "credit": credit,
                    }
                )

            if not lines:
                # Some exports use different nesting; tolerate but still output row
                pass

            rows.append(
                {
                    "record_type": "VOUCHER",
                    "voucher_type": vtype,
                    "date": vdate,
                    "voucher_number": vno,
                    "narration": narration,
                    "lines": lines,
                }
            )

        cols: set[str] = set()
        for r in rows:
            cols.update([str(k) for k in r.keys()])

        return ParsedTable(columns=sorted(cols, key=lambda x: x.casefold()), rows=rows)

    def parse_file_to_table(self, *, file_rec: models.ImportFile) -> ParsedTable:
        ft = str(file_rec.file_type or "").lower()
        if ft == "csv":
            return self._read_csv_records(file_rec.stored_path)
        if ft in {"xlsx", "xls"}:
            return self._read_excel_records(file_rec.stored_path)
        if ft == "json":
            return self._read_json_records(file_rec.stored_path)
        if ft == "xml":
            return self._read_tally_xml_records(file_rec.stored_path)

        raise HTTPException(status_code=400, detail="Unsupported file type")

    def stage_rows(self, *, job: models.ImportJob, table: ParsedTable) -> int:
        # Clear any previous staging rows for re-upload.
        self.db.query(models.ImportStagingRow).filter(models.ImportStagingRow.import_job_id == int(job.id)).delete()
        self.db.flush()

        count = 0
        for idx, row in enumerate(table.rows, start=1):
            self.db.add(
                models.ImportStagingRow(
                    import_job_id=int(job.id),
                    row_no=int(idx),
                    raw_data=row,
                    mapped_data=None,
                    validation_errors=None,
                    status="PENDING",
                )
            )
            count += 1

        job.status = "UPLOADED"
        self.db.add(job)
        self.db.commit()
        return count

    def detect_columns(self, *, job: models.ImportJob) -> list[str]:
        row = (
            self.db.query(models.ImportStagingRow)
            .filter(models.ImportStagingRow.import_job_id == int(job.id))
            .order_by(models.ImportStagingRow.row_no.asc())
            .first()
        )
        if row is None:
            return []
        raw = getattr(row, "raw_data", None) or {}
        if not isinstance(raw, dict):
            return []
        return sorted([str(k) for k in raw.keys()], key=lambda x: x.casefold())

    def upsert_mapping(
        self,
        *,
        tenant_id: int,
        company_id: int,
        source_type: str,
        data_type: str,
        mapping_name: str,
        mapping_json: dict[str, Any],
    ) -> models.ImportFieldMapping:
        existing = (
            self.db.query(models.ImportFieldMapping)
            .filter(
                models.ImportFieldMapping.company_id == int(company_id),
                models.ImportFieldMapping.source_type == str(source_type),
                models.ImportFieldMapping.data_type == str(data_type),
                models.ImportFieldMapping.mapping_name == str(mapping_name),
            )
            .first()
        )
        if existing is None:
            existing = models.ImportFieldMapping(
                tenant_id=int(tenant_id),
                company_id=int(company_id),
                source_type=str(source_type),
                data_type=str(data_type),
                mapping_name=str(mapping_name),
                mapping_json=mapping_json,
            )
        else:
            existing.mapping_json = mapping_json

        self.db.add(existing)
        self.db.commit()
        self.db.refresh(existing)
        return existing

    def _find_effective_mapping(self, *, job: models.ImportJob) -> models.ImportFieldMapping | None:
        return (
            self.db.query(models.ImportFieldMapping)
            .filter(
                models.ImportFieldMapping.company_id == int(job.company_id),
                models.ImportFieldMapping.source_type == str(job.source_type),
                models.ImportFieldMapping.data_type == str(job.data_type),
            )
            .order_by(models.ImportFieldMapping.updated_at.desc(), models.ImportFieldMapping.id.desc())
            .first()
        )

    def _apply_mapping_to_row(self, *, raw: dict[str, Any], mapping_json: dict[str, Any]) -> dict[str, Any]:
        header_map: dict[str, Any]
        lines_map: dict[str, Any] | None = None
        flat_line_map: dict[str, Any] | None = None
        group_key: str | None = None

        group_key_val = mapping_json.get("group_key")
        if isinstance(group_key_val, str) and group_key_val.strip():
            group_key = group_key_val.strip()

        if isinstance(mapping_json.get("header"), dict):
            header_map = dict(mapping_json.get("header") or {})
            if isinstance(mapping_json.get("lines"), dict):
                lines_map = dict(mapping_json.get("lines") or {})
            if isinstance(mapping_json.get("line"), dict):
                flat_line_map = dict(mapping_json.get("line") or {})
        else:
            header_map = dict(mapping_json or {})

        mapped: dict[str, Any] = {}
        for target, spec in header_map.items():
            if spec is None:
                mapped[target] = None
                continue
            if isinstance(spec, str):
                mapped[target] = raw.get(spec)
                continue
            mapped[target] = spec

        if lines_map is not None:
            raw_lines = raw.get("lines")
            out_lines: list[dict[str, Any]] = []
            if isinstance(raw_lines, list):
                for rl in raw_lines:
                    if not isinstance(rl, dict):
                        continue
                    line_obj: dict[str, Any] = {}
                    for target, spec in lines_map.items():
                        if spec is None:
                            line_obj[target] = None
                        elif isinstance(spec, str):
                            line_obj[target] = rl.get(spec)
                        else:
                            line_obj[target] = spec
                    out_lines.append(line_obj)
            mapped["lines"] = out_lines

        # Flat-row mode: current raw row represents one line item.
        # Mapping JSON shape: {"header": {...}, "line": {...}, "group_key": "reference"}
        if flat_line_map is not None and "lines" not in mapped:
            line_obj: dict[str, Any] = {}
            for target, spec in flat_line_map.items():
                if spec is None:
                    line_obj[target] = None
                elif isinstance(spec, str):
                    line_obj[target] = raw.get(spec)
                else:
                    line_obj[target] = spec
            mapped["lines"] = [line_obj]

        if group_key and "group_key" not in mapped:
            mapped["group_key"] = group_key

        return mapped

    def _infer_group_ref(self, *, mapped: dict[str, Any], raw: dict[str, Any]) -> str:
        # Prefer explicit external_ref/reference in mapped, then raw.
        candidates = [
            mapped.get("external_ref"),
            mapped.get("reference"),
            mapped.get("invoice_no"),
            mapped.get("order_id"),
            raw.get("external_ref"),
            raw.get("reference"),
            raw.get("invoice_no"),
            raw.get("order_id"),
        ]
        for c in candidates:
            if c is None:
                continue
            s = str(c).strip()
            if s:
                return s
        return ""

    def _group_invoice_rows(
        self,
        *,
        job: models.ImportJob,
        rows: list[models.ImportStagingRow],
    ) -> dict[str, list[models.ImportStagingRow]]:
        mapping = self._find_effective_mapping(job=job)
        mapping_json = dict(getattr(mapping, "mapping_json", None) or {}) if mapping is not None else {}
        group_key = str(mapping_json.get("group_key") or "").strip() or "reference"

        grouped: dict[str, list[models.ImportStagingRow]] = {}
        for r in rows:
            raw = getattr(r, "raw_data", None) or {}
            mapped = getattr(r, "mapped_data", None) or {}

            # If mapping defined a group_key field, try that column in raw/mapped.
            key_val = mapped.get(group_key)
            if key_val is None and isinstance(raw, dict):
                key_val = raw.get(group_key)

            group_ref = str(key_val).strip() if key_val is not None else ""
            if not group_ref:
                group_ref = self._infer_group_ref(mapped=mapped, raw=raw)
            if not group_ref:
                # Fallback: bucket per-row (still importable but no grouping)
                group_ref = f"row_{int(r.id)}"

            grouped.setdefault(group_ref, []).append(r)

        return grouped

    def _group_orders_rows(
        self,
        *,
        job: models.ImportJob,
        rows: list[models.ImportStagingRow],
    ) -> dict[str, list[models.ImportStagingRow]]:
        mapping = self._find_effective_mapping(job=job)
        mapping_json = dict(getattr(mapping, "mapping_json", None) or {}) if mapping is not None else {}
        group_key = str(mapping_json.get("group_key") or "").strip() or "order_id"

        grouped: dict[str, list[models.ImportStagingRow]] = {}
        for r in rows:
            raw = getattr(r, "raw_data", None) or {}
            mapped = getattr(r, "mapped_data", None) or {}

            key_val = mapped.get(group_key)
            if key_val is None and isinstance(raw, dict):
                key_val = raw.get(group_key)

            group_ref = str(key_val).strip() if key_val is not None else ""
            if not group_ref:
                group_ref = self._infer_group_ref(mapped=mapped, raw=raw)
            if not group_ref:
                group_ref = f"row_{int(r.id)}"

            grouped.setdefault(group_ref, []).append(r)
        return grouped

    def _parse_float(self, v: Any) -> float | None:
        if v is None or v == "":
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).strip())
        except Exception:
            return None

    def _parse_int(self, v: Any) -> int | None:
        if v is None or v == "":
            return None
        if isinstance(v, bool):
            return None
        if isinstance(v, int):
            return int(v)
        try:
            return int(float(str(v).strip()))
        except Exception:
            return None

    def _parse_date(self, v: Any):
        if v is None or v == "":
            return None
        if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
            try:
                return v
            except Exception:
                pass
        try:
            from datetime import date

            return date.fromisoformat(str(v).strip()[0:10])
        except Exception:
            return None

    def validate_job(self, *, job: models.ImportJob, current_user: models.User) -> dict[str, Any]:
        self._ensure_job_access(job=job, current_user=current_user)

        mapping = self._find_effective_mapping(job=job)
        if mapping is None:
            raise HTTPException(status_code=400, detail="No mapping found for this job")

        mapping_json = dict(getattr(mapping, "mapping_json", None) or {})

        rows = (
            self.db.query(models.ImportStagingRow)
            .filter(models.ImportStagingRow.import_job_id == int(job.id))
            .order_by(models.ImportStagingRow.row_no.asc())
            .all()
        )

        job.status = "VALIDATING"
        self.db.add(job)
        self.db.commit()

        total = 0
        valid = 0
        error = 0

        company_id = int(job.company_id)

        # For invoices, support flat-row imports by allowing grouping.
        if str(job.data_type) in {"sales_invoices", "purchase_invoices"}:
            # First apply mapping and basic per-row checks as before.
            for r in rows:
                total += 1
                raw = getattr(r, "raw_data", None) or {}
                mapped = self._apply_mapping_to_row(raw=raw, mapping_json=mapping_json)
                r.mapped_data = mapped
                self.db.add(r)
            self.db.commit()

            grouped = self._group_invoice_rows(job=job, rows=rows)

            for group_ref, group_rows in grouped.items():
                # Validate header once (take first row as header source)
                header_row = group_rows[0]
                header_mapped = getattr(header_row, "mapped_data", None) or {}

                errors_group: dict[str, Any] = {}
                doc_date = self._parse_date(header_mapped.get("date") or header_mapped.get("invoice_date"))
                if doc_date is None:
                    errors_group["date"] = "invalid_or_missing"
                else:
                    header_mapped["date"] = doc_date
                    header_row.mapped_data = header_mapped
                    self.db.add(header_row)

                party_id_field = "customer_id" if str(job.data_type) == "sales_invoices" else "supplier_id"
                party_name_field = "customer_name" if str(job.data_type) == "sales_invoices" else "supplier_name"
                party_id = self._parse_int(header_mapped.get(party_id_field))
                party_name = header_mapped.get(party_name_field)
                if party_id is None:
                    if isinstance(party_name, str) and party_name.strip():
                        if str(job.data_type) == "sales_invoices":
                            party = (
                                self.db.query(models.Customer)
                                .filter(models.Customer.company_id == company_id, models.Customer.name == party_name.strip())
                                .order_by(models.Customer.id.asc())
                                .first()
                            )
                        else:
                            party = (
                                self.db.query(models.Supplier)
                                .filter(models.Supplier.company_id == company_id, models.Supplier.name == party_name.strip())
                                .order_by(models.Supplier.id.asc())
                                .first()
                            )
                        if party is not None:
                            header_mapped[party_id_field] = int(party.id)
                            header_row.mapped_data = header_mapped
                            self.db.add(header_row)
                        else:
                            errors_group[party_name_field] = "not_found"
                    else:
                        errors_group[party_id_field] = "required"

                # Validate each line row
                any_line_error = False
                for r in group_rows:
                    mapped = getattr(r, "mapped_data", None) or {}
                    lines = mapped.get("lines")
                    if not isinstance(lines, list) or not lines:
                        r.validation_errors = {"lines": "required"}
                        r.status = "ERROR"
                        error += 1
                        any_line_error = True
                        self.db.add(r)
                        continue

                    line = lines[0] if isinstance(lines[0], dict) else {}
                    le: dict[str, Any] = {}
                    item_id = self._parse_int(line.get("item_id"))
                    item_name = line.get("item_name")
                    if item_id is None:
                        if isinstance(item_name, str) and item_name.strip():
                            item = (
                                self.db.query(models.Item)
                                .filter(models.Item.company_id == company_id, models.Item.name == item_name.strip())
                                .order_by(models.Item.id.asc())
                                .first()
                            )
                            if item is not None:
                                item_id = int(item.id)
                                line["item_id"] = item_id
                            else:
                                le["item_name"] = "not_found"
                        else:
                            le["item_id"] = "required"

                    qty = self._parse_float(line.get("quantity"))
                    rate = self._parse_float(line.get("rate"))
                    if qty is None or qty <= 0:
                        le["quantity"] = "must_be_positive"
                    if rate is None or rate < 0:
                        le["rate"] = "must_be_non_negative"

                    disc = self._parse_float(line.get("discount"))
                    if disc is None:
                        disc = 0.0
                    tax_rate = self._parse_float(line.get("tax_rate"))
                    if tax_rate is None:
                        tax_rate = 0.0

                    line["quantity"] = float(qty or 0)
                    line["rate"] = float(rate or 0)
                    line["discount"] = float(disc)
                    line["tax_rate"] = float(tax_rate)
                    mapped["lines"] = [line]
                    r.mapped_data = mapped

                    if errors_group or le:
                        any_line_error = True
                        r.validation_errors = {**errors_group, **({"line": le} if le else {})}
                        r.status = "ERROR"
                        error += 1
                    else:
                        r.validation_errors = None
                        r.status = "VALID"
                        valid += 1

                    self.db.add(r)

                # If group-level errors exist, ensure all rows are ERROR
                if errors_group and not any_line_error:
                    for r in group_rows:
                        r.validation_errors = errors_group
                        r.status = "ERROR"
                        error += 1
                        valid = max(valid - 1, 0)
                        self.db.add(r)

            self.db.commit()

            job.status = "VALIDATED"
            self.db.add(job)
            self.db.commit()

            return {
                "job_id": int(job.id),
                "total_rows": int(total),
                "valid_rows": int(valid),
                "error_rows": int(error),
                "status": str(job.status),
            }

        # E-commerce orders (flat-row grouped)
        if str(job.data_type) == "orders":
            for r in rows:
                total += 1
                raw = getattr(r, "raw_data", None) or {}
                mapped = self._apply_mapping_to_row(raw=raw, mapping_json=mapping_json)
                r.mapped_data = mapped
                self.db.add(r)
            self.db.commit()

            grouped = self._group_orders_rows(job=job, rows=rows)

            for group_ref, group_rows in grouped.items():
                header_row = group_rows[0]
                header_mapped = getattr(header_row, "mapped_data", None) or {}

                errors_group: dict[str, Any] = {}

                doc_date = self._parse_date(header_mapped.get("date") or header_mapped.get("order_date"))
                if doc_date is None:
                    errors_group["date"] = "invalid_or_missing"
                else:
                    header_mapped["date"] = doc_date
                    header_row.mapped_data = header_mapped
                    self.db.add(header_row)

                # Resolve customer
                customer_id = self._parse_int(header_mapped.get("customer_id"))
                customer_name = header_mapped.get("customer_name") or header_mapped.get("customer")
                if customer_id is None:
                    if isinstance(customer_name, str) and customer_name.strip():
                        cust = (
                            self.db.query(models.Customer)
                            .filter(models.Customer.company_id == company_id, models.Customer.name == customer_name.strip())
                            .order_by(models.Customer.id.asc())
                            .first()
                        )
                        if cust is not None:
                            customer_id = int(cust.id)
                            header_mapped["customer_id"] = customer_id
                            header_row.mapped_data = header_mapped
                            self.db.add(header_row)
                        else:
                            errors_group["customer_name"] = "not_found"
                    else:
                        errors_group["customer_id"] = "required"

                # Validate each line row
                any_line_error = False
                for r in group_rows:
                    mapped = getattr(r, "mapped_data", None) or {}
                    lines = mapped.get("lines")
                    if not isinstance(lines, list) or not lines:
                        r.validation_errors = {"lines": "required", **errors_group}
                        r.status = "ERROR"
                        error += 1
                        any_line_error = True
                        self.db.add(r)
                        continue

                    line = lines[0] if isinstance(lines[0], dict) else {}
                    le: dict[str, Any] = {}
                    item_id = self._parse_int(line.get("item_id"))
                    item_name = line.get("item_name") or line.get("item")
                    if item_id is None:
                        if isinstance(item_name, str) and item_name.strip():
                            item = (
                                self.db.query(models.Item)
                                .filter(models.Item.company_id == company_id, models.Item.name == item_name.strip())
                                .order_by(models.Item.id.asc())
                                .first()
                            )
                            if item is not None:
                                item_id = int(item.id)
                                line["item_id"] = item_id
                            else:
                                le["item_name"] = "not_found"
                        else:
                            le["item_id"] = "required"

                    qty = self._parse_float(line.get("quantity") or line.get("qty"))
                    rate = self._parse_float(line.get("rate"))
                    if qty is None or qty <= 0:
                        le["quantity"] = "must_be_positive"
                    if rate is None or rate < 0:
                        le["rate"] = "must_be_non_negative"
                    disc = self._parse_float(line.get("discount"))
                    if disc is None:
                        disc = 0.0
                    tax_rate = self._parse_float(line.get("tax_rate"))
                    if tax_rate is None:
                        tax_rate = 0.0

                    line["quantity"] = float(qty or 0)
                    line["rate"] = float(rate or 0)
                    line["discount"] = float(disc)
                    line["tax_rate"] = float(tax_rate)
                    mapped["lines"] = [line]
                    r.mapped_data = mapped

                    if errors_group or le:
                        any_line_error = True
                        r.validation_errors = {**errors_group, **({"line": le} if le else {})}
                        r.status = "ERROR"
                        error += 1
                    else:
                        r.validation_errors = None
                        r.status = "VALID"
                        valid += 1
                    self.db.add(r)

                if errors_group and not any_line_error:
                    for r in group_rows:
                        r.validation_errors = errors_group
                        r.status = "ERROR"
                        error += 1
                        valid = max(valid - 1, 0)
                        self.db.add(r)

            self.db.commit()

            job.status = "VALIDATED"
            self.db.add(job)
            self.db.commit()

            return {
                "job_id": int(job.id),
                "total_rows": int(total),
                "valid_rows": int(valid),
                "error_rows": int(error),
                "status": str(job.status),
            }

        for r in rows:
            total += 1
            raw = getattr(r, "raw_data", None) or {}
            mapped = self._apply_mapping_to_row(raw=raw, mapping_json=mapping_json)

            errors: dict[str, Any] = {}
            dt = str(job.data_type)

            if dt == "masters_warehouses":
                name = (mapped.get("name") or "").strip() if isinstance(mapped.get("name"), str) else mapped.get("name")
                if not name:
                    errors["name"] = "required"
            elif dt == "masters_items":
                name = (mapped.get("name") or "").strip() if isinstance(mapped.get("name"), str) else mapped.get("name")
                if not name:
                    errors["name"] = "required"
            elif dt == "masters_ledgers":
                name = (mapped.get("name") or "").strip() if isinstance(mapped.get("name"), str) else mapped.get("name")
                if not name:
                    errors["name"] = "required"
                group_id = self._parse_int(mapped.get("group_id"))
                group_name = mapped.get("group_name")
                if group_id is None:
                    if isinstance(group_name, str) and group_name.strip():
                        grp = (
                            self.db.query(models.LedgerGroup)
                            .filter(
                                models.LedgerGroup.company_id == company_id,
                                models.LedgerGroup.name == group_name.strip(),
                            )
                            .order_by(models.LedgerGroup.id.asc())
                            .first()
                        )
                        if grp is not None:
                            mapped["group_id"] = int(grp.id)
                        else:
                            errors["group_name"] = "not_found"
                    else:
                        errors["group_id"] = "required"
            elif dt in {"journals", "payments_receipts"}:
                vdate = self._parse_date(mapped.get("voucher_date") or mapped.get("date"))
                if vdate is None:
                    errors["voucher_date"] = "invalid_or_missing"
                else:
                    mapped["voucher_date"] = vdate

                lines = mapped.get("lines")
                if not isinstance(lines, list) or not lines:
                    errors["lines"] = "required"
                else:
                    dr = 0.0
                    cr = 0.0
                    line_errors: list[dict[str, Any]] = []
                    for idx, ln in enumerate(lines, start=1):
                        if not isinstance(ln, dict):
                            continue
                        le: dict[str, Any] = {}
                        ledger_id = self._parse_int(ln.get("ledger_id"))
                        ledger_name = ln.get("ledger_name")
                        if ledger_id is None:
                            if isinstance(ledger_name, str) and ledger_name.strip():
                                led = (
                                    self.db.query(models.Ledger)
                                    .filter(
                                        models.Ledger.company_id == company_id,
                                        models.Ledger.name == ledger_name.strip(),
                                    )
                                    .order_by(models.Ledger.id.asc())
                                    .first()
                                )
                                if led is not None:
                                    ledger_id = int(led.id)
                                else:
                                    le["ledger_name"] = "not_found"
                            else:
                                le["ledger_id"] = "required"

                        debit = self._parse_float(ln.get("debit")) or 0.0
                        credit = self._parse_float(ln.get("credit")) or 0.0
                        if debit < 0 or credit < 0:
                            le["amount"] = "must_be_non_negative"
                        if debit > 0 and credit > 0:
                            le["amount"] = "either_debit_or_credit"

                        dr += float(debit)
                        cr += float(credit)

                        ln["ledger_id"] = ledger_id
                        ln["debit"] = float(debit)
                        ln["credit"] = float(credit)
                        if le:
                            le["line_no"] = idx
                            line_errors.append(le)

                    if abs(dr - cr) > 1e-6:
                        errors["balance"] = {"debit": dr, "credit": cr}
                    if line_errors:
                        errors["line_errors"] = line_errors

            elif dt == "opening_balances":
                ledger_id = self._parse_int(mapped.get("ledger_id"))
                ledger_name = mapped.get("ledger_name")
                if ledger_id is None:
                    if isinstance(ledger_name, str) and ledger_name.strip():
                        led = (
                            self.db.query(models.Ledger)
                            .filter(models.Ledger.company_id == company_id, models.Ledger.name == ledger_name.strip())
                            .order_by(models.Ledger.id.asc())
                            .first()
                        )
                        if led is not None:
                            ledger_id = int(led.id)
                            mapped["ledger_id"] = ledger_id
                        else:
                            errors["ledger_name"] = "not_found"
                    else:
                        errors["ledger_id"] = "required"

                amt = self._parse_float(mapped.get("opening_balance"))
                if amt is None:
                    errors["opening_balance"] = "invalid_or_missing"
                else:
                    mapped["opening_balance"] = float(amt)

                ob_type = mapped.get("opening_balance_type") or mapped.get("balance_type")
                if ob_type is None:
                    errors["opening_balance_type"] = "required"
                else:
                    if isinstance(ob_type, str):
                        obs = ob_type.strip().upper()
                        if obs not in {"DEBIT", "CREDIT"}:
                            errors["opening_balance_type"] = "must_be_DEBIT_or_CREDIT"
                        else:
                            mapped["opening_balance_type"] = obs
                    else:
                        # allow enum passthrough
                        mapped["opening_balance_type"] = ob_type

            elif dt == "stock_opening":
                item_id = self._parse_int(mapped.get("item_id"))
                item_name = mapped.get("item_name")
                if item_id is None:
                    if isinstance(item_name, str) and item_name.strip():
                        item = (
                            self.db.query(models.Item)
                            .filter(models.Item.company_id == company_id, models.Item.name == item_name.strip())
                            .order_by(models.Item.id.asc())
                            .first()
                        )
                        if item is not None:
                            item_id = int(item.id)
                            mapped["item_id"] = item_id
                        else:
                            errors["item_name"] = "not_found"
                    else:
                        errors["item_id"] = "required"

                opening_stock = self._parse_float(mapped.get("opening_stock"))
                if opening_stock is None:
                    errors["opening_stock"] = "invalid_or_missing"
                else:
                    mapped["opening_stock"] = float(opening_stock)

                opening_rate = mapped.get("opening_rate")
                if opening_rate is not None and opening_rate != "":
                    parsed_rate = self._parse_float(opening_rate)
                    if parsed_rate is None:
                        errors["opening_rate"] = "invalid"
                    else:
                        mapped["opening_rate"] = float(parsed_rate)

                opening_date = mapped.get("opening_date")
                if opening_date is not None and opening_date != "":
                    parsed_date = self._parse_date(opening_date)
                    if parsed_date is None:
                        errors["opening_date"] = "invalid"
                    else:
                        mapped["opening_date"] = parsed_date

            elif dt in {"sales_invoices", "purchase_invoices"}:
                doc_date = self._parse_date(mapped.get("date") or mapped.get("invoice_date"))
                if doc_date is None:
                    errors["date"] = "invalid_or_missing"
                else:
                    mapped["date"] = doc_date

                party_id_field = "customer_id" if dt == "sales_invoices" else "supplier_id"
                party_name_field = "customer_name" if dt == "sales_invoices" else "supplier_name"
                party_id = self._parse_int(mapped.get(party_id_field))
                party_name = mapped.get(party_name_field)

                if party_id is None:
                    if isinstance(party_name, str) and party_name.strip():
                        if dt == "sales_invoices":
                            party = (
                                self.db.query(models.Customer)
                                .filter(models.Customer.company_id == company_id, models.Customer.name == party_name.strip())
                                .order_by(models.Customer.id.asc())
                                .first()
                            )
                        else:
                            party = (
                                self.db.query(models.Supplier)
                                .filter(models.Supplier.company_id == company_id, models.Supplier.name == party_name.strip())
                                .order_by(models.Supplier.id.asc())
                                .first()
                            )
                        if party is not None:
                            mapped[party_id_field] = int(party.id)
                        else:
                            errors[party_name_field] = "not_found"
                    else:
                        errors[party_id_field] = "required"

                lines = mapped.get("lines")
                if not isinstance(lines, list) or not lines:
                    errors["lines"] = "required"
                else:
                    line_errors: list[dict[str, Any]] = []
                    for idx, ln in enumerate(lines, start=1):
                        if not isinstance(ln, dict):
                            continue
                        le: dict[str, Any] = {}
                        item_id = self._parse_int(ln.get("item_id"))
                        item_name = ln.get("item_name")
                        if item_id is None:
                            if isinstance(item_name, str) and item_name.strip():
                                item = (
                                    self.db.query(models.Item)
                                    .filter(models.Item.company_id == company_id, models.Item.name == item_name.strip())
                                    .order_by(models.Item.id.asc())
                                    .first()
                                )
                                if item is not None:
                                    item_id = int(item.id)
                                else:
                                    le["item_name"] = "not_found"
                            else:
                                le["item_id"] = "required"

                        qty = self._parse_float(ln.get("quantity"))
                        rate = self._parse_float(ln.get("rate"))
                        if qty is None or qty <= 0:
                            le["quantity"] = "must_be_positive"
                        if rate is None or rate < 0:
                            le["rate"] = "must_be_non_negative"
                        disc = self._parse_float(ln.get("discount"))
                        if disc is None:
                            disc = 0.0
                        tax_rate = self._parse_float(ln.get("tax_rate"))
                        if tax_rate is None:
                            tax_rate = 0.0

                        ln["item_id"] = item_id
                        ln["quantity"] = float(qty or 0)
                        ln["rate"] = float(rate or 0)
                        ln["discount"] = float(disc)
                        ln["tax_rate"] = float(tax_rate)
                        if le:
                            le["line_no"] = idx
                            line_errors.append(le)
                    if line_errors:
                        errors["line_errors"] = line_errors
            else:
                errors["data_type"] = "unsupported_for_validation"

            r.mapped_data = mapped
            if errors:
                r.validation_errors = errors
                r.status = "ERROR"
                error += 1
            else:
                r.validation_errors = None
                r.status = "VALID"
                valid += 1
            self.db.add(r)

        if error:
            job.status = "VALIDATED"
        else:
            job.status = "VALIDATED"
        self.db.add(job)
        self.db.commit()

        return {
            "job_id": int(job.id),
            "total_rows": int(total),
            "valid_rows": int(valid),
            "error_rows": int(error),
            "status": str(job.status),
        }

    def _idempotency_exists(self, *, company_id: int, data_type: str, external_ref: str) -> models.ImportIdempotencyKey | None:
        return (
            self.db.query(models.ImportIdempotencyKey)
            .filter(
                models.ImportIdempotencyKey.company_id == int(company_id),
                models.ImportIdempotencyKey.data_type == str(data_type),
                models.ImportIdempotencyKey.external_ref == str(external_ref),
            )
            .first()
        )

    def commit_job(self, *, job: models.ImportJob, current_user: models.User) -> dict[str, Any]:
        self._ensure_job_access(job=job, current_user=current_user)

        company_id = int(job.company_id)
        tenant_id = int(job.tenant_id)
        dt = str(job.data_type)

        rows = (
            self.db.query(models.ImportStagingRow)
            .filter(
                models.ImportStagingRow.import_job_id == int(job.id),
                models.ImportStagingRow.status == "VALID",
            )
            .order_by(models.ImportStagingRow.row_no.asc())
            .all()
        )

        if not rows:
            raise HTTPException(status_code=400, detail="No VALID rows to commit")

        job.status = "COMMITTING"
        self.db.add(job)
        self.db.commit()

        created: dict[str, Any] = {"created": [], "skipped": []}
        summary: dict[str, Any] = {"processed": 0, "created": 0, "skipped": 0}

        try:
            # For invoices/bills, create one document per group_ref (flat rows).
            if dt in {"sales_invoices", "purchase_invoices"}:
                grouped = self._group_invoice_rows(job=job, rows=rows)
                for group_ref, group_rows in grouped.items():
                    summary["processed"] += len(group_rows)

                    header_row = group_rows[0]
                    header_mapped = getattr(header_row, "mapped_data", None) or {}

                    ext_ref_s = self._infer_group_ref(mapped=header_mapped, raw=getattr(header_row, "raw_data", None) or {})
                    if not ext_ref_s:
                        ext_ref_s = str(group_ref)

                    existing = self._idempotency_exists(company_id=company_id, data_type=dt, external_ref=ext_ref_s)
                    if existing is not None:
                        for r in group_rows:
                            created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                            summary["skipped"] += 1
                        continue

                    lines: list[dict[str, Any]] = []
                    for r in group_rows:
                        mapped = getattr(r, "mapped_data", None) or {}
                        row_lines = mapped.get("lines")
                        if isinstance(row_lines, list) and row_lines and isinstance(row_lines[0], dict):
                            lines.append(row_lines[0])

                    created_entity_type: str | None = None
                    created_entity_id: int | None = None

                    if dt == "sales_invoices":
                        from .schemas import SalesInvoiceCreate, SalesInvoiceLine
                        from .routers import sales

                        inv_lines = [
                            SalesInvoiceLine(
                                item_id=int(ln.get("item_id")),
                                quantity=float(ln.get("quantity") or 0),
                                rate=float(ln.get("rate") or 0),
                                discount=float(ln.get("discount") or 0),
                                tax_rate=float(ln.get("tax_rate") or 0),
                                warehouse_id=ln.get("warehouse_id"),
                            )
                            for ln in lines
                            if ln.get("item_id") is not None
                        ]
                        payload = SalesInvoiceCreate(
                            customer_id=int(header_mapped.get("customer_id")),
                            date=header_mapped.get("date"),
                            due_date=header_mapped.get("due_date"),
                            reference=header_mapped.get("reference") or ext_ref_s or None,
                            sales_person_id=header_mapped.get("sales_person_id"),
                            sales_ledger_id=header_mapped.get("sales_ledger_id"),
                            output_tax_ledger_id=header_mapped.get("output_tax_ledger_id"),
                            lines=inv_lines,
                        )
                        obj = sales.create_invoice(
                            company_id=company_id,
                            invoice_in=payload,
                            db=self.db,
                            current_user=current_user,
                        )
                        created_entity_type = "SALES_INVOICE"
                        created_entity_id = int(obj.id)
                    else:
                        from .schemas import PurchaseBillCreate, PurchaseBillLine
                        from .routers import purchases

                        bill_lines = [
                            PurchaseBillLine(
                                item_id=int(ln.get("item_id")),
                                quantity=float(ln.get("quantity") or 0),
                                rate=float(ln.get("rate") or 0),
                                discount=float(ln.get("discount") or 0),
                                tax_rate=float(ln.get("tax_rate") or 0),
                                warehouse_id=ln.get("warehouse_id"),
                            )
                            for ln in lines
                            if ln.get("item_id") is not None
                        ]
                        payload = PurchaseBillCreate(
                            supplier_id=int(header_mapped.get("supplier_id")),
                            date=header_mapped.get("date"),
                            reference=header_mapped.get("reference") or ext_ref_s or None,
                            purchase_ledger_id=header_mapped.get("purchase_ledger_id"),
                            input_tax_ledger_id=header_mapped.get("input_tax_ledger_id"),
                            lines=bill_lines,
                        )
                        obj = purchases.create_bill(
                            company_id=company_id,
                            bill_in=payload,
                            db=self.db,
                            current_user=current_user,
                        )
                        created_entity_type = "PURCHASE_BILL"
                        created_entity_id = int(obj.id)

                    self.db.add(
                        models.ImportIdempotencyKey(
                            tenant_id=tenant_id,
                            company_id=company_id,
                            import_job_id=int(job.id),
                            data_type=dt,
                            external_ref=str(ext_ref_s),
                            created_entity_type=created_entity_type,
                            created_entity_id=created_entity_id,
                            created_at=datetime.utcnow(),
                        )
                    )
                    created["created"].append(
                        {
                            "group_ref": str(group_ref),
                            "row_ids": [int(x.id) for x in group_rows],
                            "entity_type": created_entity_type,
                            "entity_id": created_entity_id,
                            "external_ref": ext_ref_s,
                        }
                    )
                    summary["created"] += 1

                    try:
                        self.db.commit()
                    except IntegrityError:
                        self.db.rollback()
                        created["skipped"].append({"group_ref": str(group_ref), "external_ref": ext_ref_s})
                        summary["skipped"] += 1
                        summary["created"] = max(int(summary.get("created", 0)) - 1, 0)

                res = (
                    self.db.query(models.ImportResult)
                    .filter(models.ImportResult.import_job_id == int(job.id))
                    .first()
                )
                if res is None:
                    res = models.ImportResult(import_job_id=int(job.id))
                res.created_ids = created
                res.summary = summary
                res.completed_at = datetime.utcnow()
                self.db.add(res)
                job.status = "COMPLETED"
                self.db.add(job)
                self.db.commit()

                self.db.refresh(job)
                return {
                    "job_id": int(job.id),
                    "status": str(job.status),
                    "created_ids": created,
                    "summary": summary,
                }

            if dt == "orders":
                from .routers import orders
                from .schemas import SalesOrderCreate, SalesOrderLine

                grouped = self._group_orders_rows(job=job, rows=rows)
                for group_ref, group_rows in grouped.items():
                    summary["processed"] += len(group_rows)

                    header_row = group_rows[0]
                    header_mapped = getattr(header_row, "mapped_data", None) or {}

                    ext_ref_s = self._infer_group_ref(mapped=header_mapped, raw=getattr(header_row, "raw_data", None) or {})
                    if not ext_ref_s:
                        ext_ref_s = str(group_ref)

                    existing = self._idempotency_exists(company_id=company_id, data_type=dt, external_ref=ext_ref_s)
                    if existing is not None:
                        for r in group_rows:
                            created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                            summary["skipped"] += 1
                        continue

                    customer_id = int(header_mapped.get("customer_id"))
                    order_date = header_mapped.get("date")
                    due_date = header_mapped.get("due_date")
                    reference = header_mapped.get("reference") or header_mapped.get("order_number") or ext_ref_s

                    lines_raw: list[dict[str, Any]] = []
                    for r in group_rows:
                        mapped = getattr(r, "mapped_data", None) or {}
                        row_lines = mapped.get("lines")
                        if isinstance(row_lines, list) and row_lines and isinstance(row_lines[0], dict):
                            lines_raw.append(row_lines[0])

                    order_lines = [
                        SalesOrderLine(
                            item_id=int(ln.get("item_id")),
                            quantity=float(ln.get("quantity") or 0),
                            rate=float(ln.get("rate") or 0),
                            discount=float(ln.get("discount") or 0),
                            tax_rate=float(ln.get("tax_rate") or 0),
                        )
                        for ln in lines_raw
                        if ln.get("item_id") is not None
                    ]

                    payload = SalesOrderCreate(
                        customer_id=customer_id,
                        date=order_date,
                        due_date=due_date,
                        sales_person_id=header_mapped.get("sales_person_id"),
                        reference=str(reference) if reference is not None else None,
                        lines=order_lines,
                    )
                    obj = orders.create_sales_order(
                        company_id=company_id,
                        order_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )

                    self.db.add(
                        models.ImportIdempotencyKey(
                            tenant_id=tenant_id,
                            company_id=company_id,
                            import_job_id=int(job.id),
                            data_type=dt,
                            external_ref=str(ext_ref_s),
                            created_entity_type="SALES_ORDER",
                            created_entity_id=int(obj.id),
                            created_at=datetime.utcnow(),
                        )
                    )

                    created["created"].append(
                        {
                            "group_ref": str(group_ref),
                            "row_ids": [int(x.id) for x in group_rows],
                            "entity_type": "SALES_ORDER",
                            "entity_id": int(obj.id),
                            "external_ref": ext_ref_s,
                        }
                    )
                    summary["created"] += 1

                    try:
                        self.db.commit()
                    except IntegrityError:
                        self.db.rollback()
                        created["skipped"].append({"group_ref": str(group_ref), "external_ref": ext_ref_s})
                        summary["skipped"] += 1
                        summary["created"] = max(int(summary.get("created", 0)) - 1, 0)

                res = (
                    self.db.query(models.ImportResult)
                    .filter(models.ImportResult.import_job_id == int(job.id))
                    .first()
                )
                if res is None:
                    res = models.ImportResult(import_job_id=int(job.id))
                res.created_ids = created
                res.summary = summary
                res.completed_at = datetime.utcnow()
                self.db.add(res)
                job.status = "COMPLETED"
                self.db.add(job)
                self.db.commit()

                self.db.refresh(job)
                return {
                    "job_id": int(job.id),
                    "status": str(job.status),
                    "created_ids": created,
                    "summary": summary,
                }

            if dt == "opening_balances":
                from .routers import ledgers
                from .schemas import LedgerUpdate

                for r in rows:
                    summary["processed"] += 1
                    mapped = getattr(r, "mapped_data", None) or {}
                    raw = getattr(r, "raw_data", None) or {}
                    ext_ref_s = self._infer_group_ref(mapped=mapped, raw=raw)
                    if ext_ref_s:
                        existing = self._idempotency_exists(company_id=company_id, data_type=dt, external_ref=ext_ref_s)
                        if existing is not None:
                            created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                            summary["skipped"] += 1
                            continue

                    ledger_id = int(mapped.get("ledger_id"))
                    payload = LedgerUpdate(
                        opening_balance=float(mapped.get("opening_balance") or 0),
                        opening_balance_type=mapped.get("opening_balance_type"),
                    )
                    obj = ledgers.update_ledger(
                        company_id=company_id,
                        ledger_id=ledger_id,
                        ledger_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )

                    if ext_ref_s:
                        self.db.add(
                            models.ImportIdempotencyKey(
                                tenant_id=tenant_id,
                                company_id=company_id,
                                import_job_id=int(job.id),
                                data_type=dt,
                                external_ref=ext_ref_s,
                                created_entity_type="LEDGER",
                                created_entity_id=int(obj.id),
                                created_at=datetime.utcnow(),
                            )
                        )

                    created["created"].append({"row_id": int(r.id), "entity_type": "LEDGER", "entity_id": int(obj.id)})
                    summary["created"] += 1
                    self.db.commit()

                res = (
                    self.db.query(models.ImportResult)
                    .filter(models.ImportResult.import_job_id == int(job.id))
                    .first()
                )
                if res is None:
                    res = models.ImportResult(import_job_id=int(job.id))
                res.created_ids = created
                res.summary = summary
                res.completed_at = datetime.utcnow()
                self.db.add(res)
                job.status = "COMPLETED"
                self.db.add(job)
                self.db.commit()

                self.db.refresh(job)
                return {
                    "job_id": int(job.id),
                    "status": str(job.status),
                    "created_ids": created,
                    "summary": summary,
                }

            elif dt == "stock_opening":
                from .routers import inventory
                from .schemas import ItemUpdate

                for r in rows:
                    summary["processed"] += 1
                    mapped = getattr(r, "mapped_data", None) or {}
                    raw = getattr(r, "raw_data", None) or {}
                    ext_ref_s = self._infer_group_ref(mapped=mapped, raw=raw)
                    if ext_ref_s:
                        existing = self._idempotency_exists(company_id=company_id, data_type=dt, external_ref=ext_ref_s)
                        if existing is not None:
                            created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                            summary["skipped"] += 1
                            continue

                    item_id = int(mapped.get("item_id"))
                    payload = ItemUpdate(
                        opening_stock=mapped.get("opening_stock"),
                        opening_rate=mapped.get("opening_rate"),
                        opening_date=mapped.get("opening_date"),
                    )
                    obj = inventory.update_item(
                        company_id=company_id,
                        item_id=item_id,
                        item_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )

                    if ext_ref_s:
                        self.db.add(
                            models.ImportIdempotencyKey(
                                tenant_id=tenant_id,
                                company_id=company_id,
                                import_job_id=int(job.id),
                                data_type=dt,
                                external_ref=ext_ref_s,
                                created_entity_type="ITEM",
                                created_entity_id=int(obj.id),
                                created_at=datetime.utcnow(),
                            )
                        )
                    created["created"].append({"row_id": int(r.id), "entity_type": "ITEM", "entity_id": int(obj.id)})
                    summary["created"] += 1
                    self.db.commit()

                res = (
                    self.db.query(models.ImportResult)
                    .filter(models.ImportResult.import_job_id == int(job.id))
                    .first()
                )
                if res is None:
                    res = models.ImportResult(import_job_id=int(job.id))
                res.created_ids = created
                res.summary = summary
                res.completed_at = datetime.utcnow()
                self.db.add(res)
                job.status = "COMPLETED"
                self.db.add(job)
                self.db.commit()

                self.db.refresh(job)
                return {
                    "job_id": int(job.id),
                    "status": str(job.status),
                    "created_ids": created,
                    "summary": summary,
                }

            else:
                # fall back to per-row commit path implemented below
                pass

            for r in rows:
                summary["processed"] += 1
                mapped = getattr(r, "mapped_data", None) or {}
                raw = getattr(r, "raw_data", None) or {}
                ext_ref = (
                    mapped.get("external_ref")
                    or raw.get("external_ref")
                    or mapped.get("reference")
                    or raw.get("reference")
                )

                ext_ref_s = str(ext_ref).strip() if ext_ref is not None else ""

                if ext_ref_s:
                    existing = self._idempotency_exists(company_id=company_id, data_type=dt, external_ref=ext_ref_s)
                    if existing is not None:
                        created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                        summary["skipped"] += 1
                        continue

                created_entity_type: str | None = None
                created_entity_id: int | None = None

                # Important: the existing create_* functions commit internally.
                # So we do per-row commit tracking and abort on first failure.
                if dt == "masters_warehouses":
                    from .schemas import WarehouseCreate
                    from .routers import inventory

                    payload = WarehouseCreate(
                        code=mapped.get("code") or None,
                        name=str(mapped.get("name") or "").strip(),
                        is_active=bool(mapped.get("is_active", True)),
                    )
                    obj = inventory.create_warehouse(
                        company_id=company_id,
                        warehouse_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "WAREHOUSE"
                    created_entity_id = int(obj.id)

                elif dt == "masters_items":
                    from .schemas import ItemCreate
                    from .routers import inventory

                    payload = ItemCreate(**{k: v for k, v in mapped.items() if k != "lines"})
                    obj = inventory.create_item(
                        company_id=company_id,
                        item_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "ITEM"
                    created_entity_id = int(obj.id)

                elif dt == "masters_ledgers":
                    from .schemas import LedgerCreate
                    from .routers import ledgers

                    payload = LedgerCreate(**{k: v for k, v in mapped.items() if k != "lines"})
                    obj = ledgers.create_ledger(
                        company_id=company_id,
                        ledger_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "LEDGER"
                    created_entity_id = int(obj.id)

                elif dt == "journals":
                    from .schemas import VoucherCreate, VoucherLineCreate
                    from .routers import vouchers
                    from .models import VoucherType

                    lines = mapped.get("lines") or []
                    voucher_lines = [
                        VoucherLineCreate(
                            ledger_id=int(ln.get("ledger_id")),
                            debit=float(ln.get("debit") or 0),
                            credit=float(ln.get("credit") or 0),
                            department_id=ln.get("department_id"),
                            project_id=ln.get("project_id"),
                        )
                        for ln in lines
                        if isinstance(ln, dict) and ln.get("ledger_id") is not None
                    ]
                    vpayload = VoucherCreate(
                        voucher_date=mapped.get("voucher_date"),
                        voucher_date_bs=mapped.get("voucher_date_bs"),
                        voucher_type=VoucherType.JOURNAL,
                        narration=mapped.get("narration"),
                        payment_mode_id=mapped.get("payment_mode_id"),
                        lines=voucher_lines,
                    )
                    obj = vouchers.create_voucher(
                        company_id=company_id,
                        voucher_in=vpayload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "VOUCHER"
                    created_entity_id = int(obj.id)

                elif dt == "payments_receipts":
                    from .schemas import VoucherCreate, VoucherLineCreate
                    from .routers import vouchers
                    from .models import VoucherType

                    vtype = mapped.get("voucher_type")
                    if isinstance(vtype, str) and vtype.upper() == "PAYMENT":
                        voucher_type = VoucherType.PAYMENT
                    else:
                        voucher_type = VoucherType.RECEIPT

                    lines = mapped.get("lines") or []
                    voucher_lines = [
                        VoucherLineCreate(
                            ledger_id=int(ln.get("ledger_id")),
                            debit=float(ln.get("debit") or 0),
                            credit=float(ln.get("credit") or 0),
                            department_id=ln.get("department_id"),
                            project_id=ln.get("project_id"),
                        )
                        for ln in lines
                        if isinstance(ln, dict) and ln.get("ledger_id") is not None
                    ]
                    vpayload = VoucherCreate(
                        voucher_date=mapped.get("voucher_date"),
                        voucher_date_bs=mapped.get("voucher_date_bs"),
                        voucher_type=voucher_type,
                        narration=mapped.get("narration"),
                        payment_mode_id=mapped.get("payment_mode_id"),
                        lines=voucher_lines,
                    )
                    obj = vouchers.create_voucher(
                        company_id=company_id,
                        voucher_in=vpayload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "VOUCHER"
                    created_entity_id = int(obj.id)

                elif dt == "sales_invoices":
                    from .schemas import SalesInvoiceCreate, SalesInvoiceLine
                    from .routers import sales

                    lines = mapped.get("lines") or []
                    inv_lines = [
                        SalesInvoiceLine(
                            item_id=int(ln.get("item_id")),
                            quantity=float(ln.get("quantity") or 0),
                            rate=float(ln.get("rate") or 0),
                            discount=float(ln.get("discount") or 0),
                            tax_rate=float(ln.get("tax_rate") or 0),
                            warehouse_id=ln.get("warehouse_id"),
                        )
                        for ln in lines
                        if isinstance(ln, dict) and ln.get("item_id") is not None
                    ]
                    payload = SalesInvoiceCreate(
                        customer_id=int(mapped.get("customer_id")),
                        date=mapped.get("date"),
                        due_date=mapped.get("due_date"),
                        reference=mapped.get("reference") or ext_ref_s or None,
                        sales_person_id=mapped.get("sales_person_id"),
                        sales_ledger_id=mapped.get("sales_ledger_id"),
                        output_tax_ledger_id=mapped.get("output_tax_ledger_id"),
                        lines=inv_lines,
                    )
                    obj = sales.create_invoice(
                        company_id=company_id,
                        invoice_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "SALES_INVOICE"
                    created_entity_id = int(obj.id)

                elif dt == "purchase_invoices":
                    from .schemas import PurchaseBillCreate, PurchaseBillLine
                    from .routers import purchases

                    lines = mapped.get("lines") or []
                    bill_lines = [
                        PurchaseBillLine(
                            item_id=int(ln.get("item_id")),
                            quantity=float(ln.get("quantity") or 0),
                            rate=float(ln.get("rate") or 0),
                            discount=float(ln.get("discount") or 0),
                            tax_rate=float(ln.get("tax_rate") or 0),
                            warehouse_id=ln.get("warehouse_id"),
                            foreign_currency_rate=ln.get("foreign_currency_rate"),
                            exchange_rate=ln.get("exchange_rate"),
                        )
                        for ln in lines
                        if isinstance(ln, dict) and ln.get("item_id") is not None
                    ]
                    payload = PurchaseBillCreate(
                        supplier_id=int(mapped.get("supplier_id")),
                        date=mapped.get("date"),
                        due_date=mapped.get("due_date"),
                        reference=mapped.get("reference") or ext_ref_s or None,
                        purchase_ledger_id=mapped.get("purchase_ledger_id"),
                        input_tax_ledger_id=mapped.get("input_tax_ledger_id"),
                        purchase_type=mapped.get("purchase_type", "LOCAL"),
                        pragyapan_patra_no=mapped.get("pragyapan_patra_no"),
                        lc_no=mapped.get("lc_no"),
                        import_invoice_no=mapped.get("import_invoice_no"),
                        lines=bill_lines,
                    )
                    obj = purchases.create_bill(
                        company_id=company_id,
                        bill_in=payload,
                        db=self.db,
                        current_user=current_user,
                    )
                    created_entity_type = "PURCHASE_BILL"
                    created_entity_id = int(obj.id)

                else:
                    raise HTTPException(status_code=400, detail=f"Unsupported data_type for commit: {dt}")

                if ext_ref_s:
                    self.db.add(
                        models.ImportIdempotencyKey(
                            tenant_id=tenant_id,
                            company_id=company_id,
                            import_job_id=int(job.id),
                            data_type=dt,
                            external_ref=ext_ref_s,
                            created_entity_type=created_entity_type,
                            created_entity_id=created_entity_id,
                            created_at=datetime.utcnow(),
                        )
                    )

                created["created"].append(
                    {
                        "row_id": int(r.id),
                        "entity_type": created_entity_type,
                        "entity_id": created_entity_id,
                        "external_ref": ext_ref_s or None,
                    }
                )
                summary["created"] += 1

                # Commit idempotency key + result tracking changes for this row.
                try:
                    self.db.commit()
                except IntegrityError:
                    # If idempotency key conflicts (race/retry), treat as skipped.
                    self.db.rollback()
                    if ext_ref_s:
                        created["skipped"].append({"row_id": int(r.id), "external_ref": ext_ref_s})
                        summary["skipped"] += 1
                        summary["created"] = max(int(summary.get("created", 0)) - 1, 0)
                    else:
                        raise

            res = (
                self.db.query(models.ImportResult)
                .filter(models.ImportResult.import_job_id == int(job.id))
                .first()
            )
            if res is None:
                res = models.ImportResult(import_job_id=int(job.id))
            res.created_ids = created
            res.summary = summary
            res.completed_at = datetime.utcnow()
            self.db.add(res)

            job.status = "COMPLETED"
            self.db.add(job)
            self.db.commit()

        except HTTPException:
            self.db.rollback()
            job.status = "FAILED"
            self.db.add(job)
            self.db.commit()
            raise
        except Exception as exc:
            self.db.rollback()
            job.status = "FAILED"
            self.db.add(job)
            self.db.commit()
            raise HTTPException(status_code=500, detail=str(exc))

        self.db.refresh(job)
        return {
            "job_id": int(job.id),
            "status": str(job.status),
            "created_ids": created,
            "summary": summary,
        }
