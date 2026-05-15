from collections import defaultdict
from typing import Any, Sequence

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, File, UploadFile
from fastapi.responses import StreamingResponse
import io
import difflib
from sqlalchemy import func, select
from sqlalchemy.orm import Session
from sqlalchemy.exc import DataError
from datetime import datetime, date

from .. import models, schemas
from ..auth import get_current_user
from ..database import get_db
from .inventory import _compute_batch_stock, _compute_issue_unit_cost
from ..voucher_service import get_next_voucher_number
from ..stock_service import StockValuationService
from ..services import notification_service
from ..dependencies import get_company_secure, validate_transaction_date
from ..bom_helpers import explode_flat_kit_components, get_latest_bom_for_product
import logging as _logging

_logger = _logging.getLogger(__name__)

router = APIRouter(prefix="/companies/{company_id}", tags=["sales"])


@router.get("/hs-codes/{item_id}")
def get_item_hs_codes(
    company_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Returns a list of unique HS codes previously used for this item across item master, purchase bills, and sales invoices.
    """
    hs_codes = set()
    
    item = db.query(models.Item).filter(models.Item.id == item_id, models.Item.company_id == company_id).first()
    if item and item.hsn_sac_code:
        hs_codes.add(item.hsn_sac_code.strip())

    purchase_results = (
        db.query(models.PurchaseBillLine.hs_code)
        .join(models.PurchaseBill)
        .filter(
            models.PurchaseBill.company_id == company_id,
            models.PurchaseBillLine.item_id == item_id,
            models.PurchaseBillLine.hs_code.isnot(None),
            models.PurchaseBillLine.hs_code != "",
        )
        .distinct()
        .all()
    )
    for r in purchase_results:
        if r[0]:
            hs_codes.add(r[0].strip())

    sales_results = (
        db.query(models.SalesInvoiceLine.hs_code)
        .join(models.SalesInvoice)
        .filter(
            models.SalesInvoice.company_id == company_id,
            models.SalesInvoiceLine.item_id == item_id,
            models.SalesInvoiceLine.hs_code.isnot(None),
            models.SalesInvoiceLine.hs_code != "",
        )
        .distinct()
        .all()
    )
    for r in sales_results:
        if r[0]:
            hs_codes.add(r[0].strip())

    return list(hs_codes)



def _compute_sales_invoice_total_subquery():
    line = models.SalesInvoiceLine
    subtotal = (line.quantity * line.rate) - line.discount
    line_total = subtotal + (subtotal * (line.tax_rate / 100.0))
    return (
        select(func.coalesce(func.sum(line_total), 0))
        .where(line.invoice_id == models.SalesInvoice.id)
        .correlate(models.SalesInvoice)
        .scalar_subquery()
    )


def _compute_sales_invoice_paid_subquery(*, company_id: int):
    return (
        select(func.coalesce(func.sum(models.VoucherAllocation.allocated_amount), 0))
        .where(
            models.VoucherAllocation.company_id == company_id,
            models.VoucherAllocation.doc_type == models.AllocationDocType.SALES_INVOICE.value,
            models.VoucherAllocation.doc_id == models.SalesInvoice.id,
        )
        .correlate(models.SalesInvoice)
        .scalar_subquery()
    )


def _payment_status(*, total_amount: float, paid_amount: float, is_credit: bool = True) -> str:
    if not is_credit:
        return "PAID"
    if paid_amount >= total_amount - 1e-9 and total_amount > 0:
        return "PAID"
    if paid_amount > 1e-9:
        return "PARTIAL"
    return "UNPAID"


def _validate_sales_person(
    db: Session,
    *,
    company_id: int,
    sales_person_id: int | None,
) -> models.SalesPerson | None:
    if sales_person_id is None:
        return None
    return (
        db.query(models.SalesPerson)
        .filter(
            models.SalesPerson.company_id == company_id,
            models.SalesPerson.id == int(sales_person_id),
        )
        .first()
    )


def _reverse_stock_ledger(
    *,
    db: Session,
    company_id: int,
    source_type: str,
    source_id: int,
    created_by: int | None,
) -> None:
    StockLedger = models.StockLedger
    now = datetime.utcnow()
    rows = (
        db.query(StockLedger)
        .filter(
            StockLedger.company_id == company_id,
            StockLedger.source_type == source_type,
            StockLedger.source_id == source_id,
            StockLedger.reversed_at.is_(None),
        )
        .all()
    )
    if not rows:
        return

    for r in rows:
        r.reversed_at = now
        db.add(
            StockLedger(
                company_id=r.company_id,
                warehouse_id=r.warehouse_id,
                item_id=r.item_id,
                qty_delta=-float(r.qty_delta),
                unit_cost=r.unit_cost,
                source_type=r.source_type,
                source_id=r.source_id,
                source_line_id=r.source_line_id,
                posted_at=now,
                reversal_of_ledger_id=r.id,
                created_by=created_by,
            )
        )


def _get_company(db: Session, company_id: int, user: models.User) -> models.Company:
    return get_company_secure(db, company_id, user)


def _resolve_sales_default_warehouse(db: Session, *, company_id: int) -> models.Warehouse | None:
    w = (
        db.query(models.Warehouse)
        .filter(
            models.Warehouse.company_id == company_id,
            models.Warehouse.name == "Main",
            models.Warehouse.is_active.is_(True),
        )
        .first()
    )
    if w is None:
        w = (
            db.query(models.Warehouse)
            .filter(models.Warehouse.company_id == company_id, models.Warehouse.is_active.is_(True))
            .order_by(models.Warehouse.id.asc())
            .first()
        )
    return w


def _collect_kit_component_ids_for_lines(
    db: Session,
    *,
    company_id: int,
    lines: Sequence[Any],
    sell_as_kit_ids: set[int],
    invoice_date: date | None,
) -> set[int]:
    extra: set[int] = set()
    as_of = invoice_date
    for line in lines:
        if int(line.item_id) not in sell_as_kit_ids:
            continue
        bom = get_latest_bom_for_product(db, company_id=company_id, product_id=int(line.item_id), as_of=as_of)
        if bom and bom.items:
            for bi in bom.items:
                extra.add(int(bi.component_product_id))
    return extra


def _clear_sales_invoice_stock_postings(db: Session, *, company_id: int, invoice_id: int) -> None:
    db.query(models.StockMovement).filter(
        models.StockMovement.company_id == company_id,
        models.StockMovement.source_type == "SALES_INVOICE",
        models.StockMovement.source_id == invoice_id,
    ).delete()
    db.query(models.StockLedger).filter(
        models.StockLedger.company_id == company_id,
        models.StockLedger.source_type == "SALES_INVOICE",
        models.StockLedger.source_id == invoice_id,
        models.StockLedger.reversed_at.is_(None),
    ).delete()


def _apply_stock_postings_for_sales_invoice(
    db: Session,
    *,
    company_id: int,
    company: models.Company,
    invoice: models.SalesInvoice,
    current_user_id: int | None,
    is_service_map: dict[int, bool],
    is_fixed_asset_map: dict[int, bool],
    sell_as_kit_map: dict[int, bool],
    allow_negative_map: dict[int, bool] | None = None,
    fallback_rate_map: dict[int, float] | None = None,
    is_batch_tracked_map: dict[int, bool] | None = None,
    lines: list[models.SalesInvoiceLine] | None = None,
) -> None:
    """Post StockLedger + StockMovement for all lines (phantom kit = components only, same source_line_id)."""
    svc = StockValuationService(db)
    tenant_id = int(company.tenant_id)
    method = svc.get_inventory_valuation_method(tenant_id=tenant_id)
    posted_at = datetime.combine(invoice.date, datetime.min.time())
    default_wh = _resolve_sales_default_warehouse(db, company_id=company_id)

    iter_lines = lines if lines is not None else invoice.lines

    # Batch-fetch warehouses to avoid one query per line in the loop below
    _wh_ids = list({int(il.warehouse_id) for il in iter_lines if il.warehouse_id})
    if default_wh and int(default_wh.id) not in _wh_ids:
        _wh_ids.append(int(default_wh.id))
    _wh_map: dict[int, models.Warehouse] = {}
    if _wh_ids:
        _wh_rows = (
            db.query(models.Warehouse)
            .filter(
                models.Warehouse.id.in_(_wh_ids),
                models.Warehouse.company_id == company_id,
                models.Warehouse.is_active.is_(True),
            )
            .all()
        )
        _wh_map = {int(w.id): w for w in _wh_rows}

    for invoice_line in iter_lines:
        if is_service_map.get(invoice_line.item_id, False):
            continue

        effective_warehouse_id = invoice_line.warehouse_id
        if effective_warehouse_id is None:
            if default_wh is None:
                # If negative stock is allowed and batch tracking is OFF, we don't strictly require a warehouse
                is_compulsory = (is_batch_tracked_map or {}).get(invoice_line.item_id, False) or not (allow_negative_map or {}).get(invoice_line.item_id, False)
                if not is_compulsory:
                    continue
                raise HTTPException(status_code=400, detail="No active warehouse found for sales stock posting")
            effective_warehouse_id = int(default_wh.id)

        warehouse = _wh_map.get(int(effective_warehouse_id))
        if not warehouse:
            is_compulsory = (is_batch_tracked_map or {}).get(invoice_line.item_id, False) or not (allow_negative_map or {}).get(invoice_line.item_id, False)
            if not is_compulsory:
                continue
            raise HTTPException(status_code=400, detail="Invalid warehouse_id")

        is_fixed = is_fixed_asset_map.get(invoice_line.item_id, False)

        if sell_as_kit_map.get(invoice_line.item_id, False):
            bom = get_latest_bom_for_product(
                db, company_id=company_id, product_id=int(invoice_line.item_id), as_of=invoice.date
            )
            if bom is None or not bom.items:
                raise HTTPException(
                    status_code=400,
                    detail=f"Kit item requires a BOM (item_id={invoice_line.item_id}). Add a BOM or turn off Sell as kit.",
                )
            comp_map = explode_flat_kit_components(bom=bom, kit_qty=float(invoice_line.quantity))
            for comp_id, req_qty in comp_map.items():
                if float(req_qty) <= 0:
                    continue
                comp_fixed = is_fixed_asset_map.get(int(comp_id), False)
                if not comp_fixed and method == models.InventoryValuationMethod.FIFO:
                    total_cost = svc.fifo_consume(
                        tenant_id=tenant_id,
                        product_id=int(comp_id),
                        qty_out=float(req_qty),
                        ref_type="SALES",
                        ref_id=int(invoice.id),
                        allow_negative=(allow_negative_map or {}).get(int(comp_id), False),
                        fallback_rate=(fallback_rate_map or {}).get(int(comp_id), 0.0),
                    )
                    unit_cost = (total_cost / float(req_qty)) if float(req_qty) else 0.0
                else:
                    unit_cost = _compute_issue_unit_cost(
                        db=db,
                        company=company,
                        company_id=company_id,
                        item_id=int(comp_id),
                        warehouse_id=int(effective_warehouse_id),
                        as_of=posted_at,
                        qty_out=float(req_qty),
                    )
                db.add(
                    models.StockLedger(
                        company_id=company_id,
                        warehouse_id=int(effective_warehouse_id),
                        item_id=int(comp_id),
                        qty_delta=-float(req_qty),
                        unit_cost=unit_cost,
                        source_type="SALES_INVOICE",
                        source_id=invoice.id,
                        source_line_id=invoice_line.id,
                        posted_at=posted_at,
                        created_by=current_user_id,
                    )
                )
                db.add(
                    models.StockMovement(
                        company_id=company_id,
                        warehouse_id=int(effective_warehouse_id),
                        item_id=int(comp_id),
                        movement_date=invoice.date,
                        source_type="SALES_INVOICE",
                        source_id=invoice.id,
                        qty_in=0,
                        qty_out=float(req_qty),
                    )
                )
            continue

        if not is_fixed and method == models.InventoryValuationMethod.FIFO:
            total_cost = svc.fifo_consume(
                tenant_id=tenant_id,
                product_id=int(invoice_line.item_id),
                qty_out=float(invoice_line.quantity),
                ref_type="SALES",
                ref_id=int(invoice.id),
                allow_negative=(allow_negative_map or {}).get(int(invoice_line.item_id), False),
                fallback_rate=(fallback_rate_map or {}).get(int(invoice_line.item_id), 0.0),
            )
            unit_cost = (total_cost / float(invoice_line.quantity)) if float(invoice_line.quantity) else 0.0
        else:
            unit_cost = _compute_issue_unit_cost(
                db=db,
                company=company,
                company_id=company_id,
                item_id=invoice_line.item_id,
                warehouse_id=int(effective_warehouse_id),
                as_of=posted_at,
                qty_out=float(invoice_line.quantity),
            )
        db.add(
            models.StockLedger(
                company_id=company_id,
                warehouse_id=int(effective_warehouse_id),
                item_id=invoice_line.item_id,
                qty_delta=-float(invoice_line.quantity),
                unit_cost=unit_cost,
                source_type="SALES_INVOICE",
                source_id=invoice.id,
                source_line_id=invoice_line.id,
                posted_at=posted_at,
                created_by=current_user_id,
            )
        )
        db.add(
            models.StockMovement(
                company_id=company_id,
                warehouse_id=int(effective_warehouse_id),
                item_id=invoice_line.item_id,
                movement_date=invoice.date,
                source_type="SALES_INVOICE",
                source_id=invoice.id,
                qty_in=0,
                qty_out=invoice_line.quantity,
            )
        )


def _get_default_stock_ledger_id(db: Session, *, company_id: int) -> int | None:
    Ledger = models.Ledger
    LedgerGroup = models.LedgerGroup

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "CLOSING_STOCK",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "OPENING_STOCK",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    stock_group = (
        db.query(LedgerGroup)
        .filter(
            LedgerGroup.company_id == company_id,
            LedgerGroup.name == "Stock-in-Hand",
        )
        .first()
    )
    if stock_group is None:
        return None

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.group_id == stock_group.id,
            Ledger.name.in_(["Closing Stock", "Inventory", "Stock"]),
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = Ledger(
        company_id=company_id,
        group_id=stock_group.id,
        name="Closing Stock",
        code="CLOSING_STOCK",
        opening_balance=0,
        opening_balance_type=models.OpeningBalanceType.DEBIT,
        is_active=True,
    )
    db.add(ledger)
    db.flush()
    return ledger.id


def _get_default_cogs_ledger_id(db: Session, *, company_id: int) -> int | None:
    Ledger = models.Ledger
    LedgerGroup = models.LedgerGroup

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.code == "COGS",
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    direct_expenses_group = (
        db.query(LedgerGroup)
        .filter(
            LedgerGroup.company_id == company_id,
            LedgerGroup.name == "Direct Expenses",
        )
        .first()
    )
    if direct_expenses_group is None:
        return None

    ledger = (
        db.query(Ledger)
        .filter(
            Ledger.company_id == company_id,
            Ledger.group_id == direct_expenses_group.id,
            Ledger.name.in_([
                "Cost of Goods Sold",
                "COGS",
            ]),
        )
        .first()
    )
    if ledger is not None:
        return ledger.id

    ledger = Ledger(
        company_id=company_id,
        group_id=direct_expenses_group.id,
        name="Cost of Goods Sold",
        code="COGS",
        opening_balance=0,
        opening_balance_type=models.OpeningBalanceType.DEBIT,
        is_active=True,
    )
    db.add(ledger)
    db.flush()
    return ledger.id


# -------- Customers --------


@router.get("/customers", response_model=list[schemas.CustomerRead])
def list_customers(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    customers = (
        db.query(models.Customer)
        .filter(models.Customer.company_id == company_id)
        .order_by(models.Customer.name)
        .all()
    )
    return customers


@router.post("/customers", response_model=schemas.CustomerRead)
def create_customer(
    company_id: int,
    customer_in: schemas.CustomerCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    customer_data = customer_in.model_dump()

    # Auto-assign default ledger if not provided
    if customer_data.get("ledger_id") is None:
        default_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code == "CUSTOMERS",
            )
            .first()
        )
        if not default_ledger:
            raise HTTPException(
                status_code=400,
                detail="Default 'Customers' ledger not found. Please seed the default chart.",
            )
        customer_data["ledger_id"] = default_ledger.id
    customer = models.Customer(
        company_id=company_id,
        tenant_id=current_user.tenant_id,
        
        updated_by_id=current_user.id,
        **customer_data,
    )
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return customer


@router.get("/customers/{customer_id}", response_model=schemas.CustomerRead)
def get_customer(
    company_id: int,
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer


@router.put("/customers/{customer_id}", response_model=schemas.CustomerRead)
def update_customer(
    company_id: int,
    customer_id: int,
    customer_in: schemas.CustomerUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    for field, value in customer_in.model_dump(exclude_unset=True).items():
        setattr(customer, field, value)
    customer.updated_by_id = current_user.id
    db.commit()
    db.refresh(customer)
    return customer


@router.delete("/customers/{customer_id}")
def delete_customer(
    company_id: int,
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    db.delete(customer)
    db.commit()
    return {"detail": "Deleted"}
def _get_counterparty_ledger_for_payment_mode(
    db: Session,
    company_id: int,
    *,
    payment_mode_id: int | None,
    fallback_ledger_id: int,
) -> int:
    """Resolve the counterparty ledger.

    Priority:
    1) If payment_mode_id is provided: use PaymentMode.ledger_id (and validate company + active).
    2) Else: fall back to provided fallback_ledger_id (e.g. customer.ledger_id for credit sales).
    """

    if payment_mode_id is not None:
        pm = (
            db.query(models.PaymentMode)
            .filter(
                models.PaymentMode.id == payment_mode_id,
                models.PaymentMode.company_id == company_id,
                models.PaymentMode.is_active == True,
            )
            .first()
        )
        if not pm:
            raise HTTPException(status_code=400, detail="Invalid payment_mode_id")

        # Treat a payment mode named 'Credit' (or similar casing) as a true
        # credit transaction: do not use its ledger; instead, fall back to the
        # customer ledger so Sundry Debtors/Creditors are affected as expected.
        if pm.name.strip().lower() == "credit":
            return fallback_ledger_id

        return pm.ledger_id

    return fallback_ledger_id


# -------- Sales Invoices --------


def _build_sales_voucher(
    db: Session,
    company_id: int,
    invoice: models.SalesInvoice,
    payment_mode_id: int | None,
    sales_ledger_id: int | None = None,
    output_tax_ledger_id: int | None = None,
    payment_ledger_id: int | None = None,
    existing_voucher: models.Voucher | None = None,
    incentive_amounts: list[schemas.SalesPersonIncentiveAmount] | None = None,
) -> models.Voucher:
    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == invoice.customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")

    # Aggregate lines by income, tax, and COGS ledgers/dimensions
    income_totals: dict[tuple[int, int | None, int | None], float] = {}
    tax_totals: dict[tuple[int, int | None, int | None], float] = {}
    cogs_totals: dict[tuple[int | None, int | None], float] = {}

    grand_total = 0.0

    # --- Batch pre-fetch to eliminate N+1 queries in the line loop ---
    line_item_ids = list({int(line.item_id) for line in invoice.lines})
    _item_rows = (
        db.query(models.Item)
        .filter(models.Item.id.in_(line_item_ids), models.Item.company_id == company_id)
        .all()
    )
    item_map: dict[int, models.Item] = {int(i.id): i for i in _item_rows}

    line_ids = [int(line.id) for line in invoice.lines]
    _stock_rows_all = (
        db.query(models.StockLedger)
        .filter(
            models.StockLedger.source_type == "SALES_INVOICE",
            models.StockLedger.source_id == invoice.id,
            models.StockLedger.source_line_id.in_(line_ids),
            models.StockLedger.company_id == company_id,
        )
        .all()
    )
    stock_map: dict[int, list[models.StockLedger]] = defaultdict(list)
    for sl in _stock_rows_all:
        stock_map[int(sl.source_line_id)].append(sl)

    # Pre-fetch warehouses for lines to avoid per-line lazy loads
    warehouse_ids = list({int(line.warehouse_id) for line in invoice.lines if line.warehouse_id})
    _warehouse_rows = []
    if warehouse_ids:
        _warehouse_rows = (
            db.query(models.Warehouse)
            .filter(models.Warehouse.id.in_(warehouse_ids))
            .all()
        )
    warehouse_map: dict[int, models.Warehouse] = {int(w.id): w for w in _warehouse_rows}
    # -----------------------------------------------------------------

    for line in invoice.lines:
        item = item_map.get(int(line.item_id))
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")

        # Resolve dimensions for this line (Hub & Spoke fallback)
        wh = warehouse_map.get(int(line.warehouse_id)) if line.warehouse_id else None
        warehouse_dept_id = wh.department_id if wh else None
        warehouse_proj_id = wh.project_id if wh else None
        line_dept_id = line.department_id or warehouse_dept_id or invoice.department_id
        line_proj_id = line.project_id or warehouse_proj_id or invoice.project_id

        subtotal = float(line.quantity) * float(line.rate) - float(line.discount)
        tax = subtotal * float(line.tax_rate) / 100.0
        grand_total += subtotal + tax

        # Income aggregation
        # Prioritize Item's specific income ledger. Fallback to invoice header ledger.
        eff_sales_ledger_id = item.income_ledger_id or sales_ledger_id
        if not eff_sales_ledger_id:
            raise HTTPException(status_code=400, detail="Item missing income ledger")
        income_key = (eff_sales_ledger_id, line_dept_id, line_proj_id)
        income_totals[income_key] = income_totals.get(income_key, 0.0) + subtotal

        # Tax aggregation
        if tax:
            eff_output_tax_ledger_id = output_tax_ledger_id or item.output_tax_ledger_id
            if not eff_output_tax_ledger_id:
                raise HTTPException(status_code=400, detail="Item missing output tax ledger")
            tax_key = (eff_output_tax_ledger_id, line_dept_id, line_proj_id)
            tax_totals[tax_key] = tax_totals.get(tax_key, 0.0) + tax

        # COGS: Prioritize actual cost from StockLedger (Real Cost)
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False))

        stock_rows: list[models.StockLedger] = []
        if not is_fixed_asset:
            stock_rows = stock_map.get(int(line.id), [])

        line_cost = 0.0
        for stock_entry in stock_rows:
            qty = abs(float(stock_entry.qty_delta or 0))
            cost = float(stock_entry.unit_cost or 0)
            line_cost += qty * cost
        if line_cost:
            _logger.debug("Line %s COGS from ledger rows=%s total=%s", line.id, len(stock_rows), line_cost)
            cogs_key = (line_dept_id, line_proj_id)
            cogs_totals[cogs_key] = cogs_totals.get(cogs_key, 0.0) + line_cost

        # Fallback: Standard-cost COGS posting if no stock entry found
        # (e.g. if negative stock allowed and not recorded in ledger, or legacy data)
        elif (
            not line_cost
            and not bool(getattr(item, "sell_as_kit", False))
            and not bool(item.allow_negative_stock)
            and not item.category == "Service"
            and not is_fixed_asset
        ):
            if item.default_purchase_rate is not None:
                line_cost = float(line.quantity) * float(item.default_purchase_rate)
                if line_cost:
                    cogs_key = (line_dept_id, line_proj_id)
                    cogs_totals[cogs_key] = cogs_totals.get(cogs_key, 0.0) + line_cost

    # ── TDS Receivable ───────────────────────────────────────────────
    # When TDS is applied, the customer pays net of TDS (they deduct it at source).
    # Accounting:
    #   DR  Customer (Sundry Debtor)      (grand_total - tds_amount)
    #   DR  TDS Receivable ledger         (tds_amount)
    #   CR  Sales / Income                (subtotal)
    #   CR  Output VAT                    (tax)
    tds_amount_value = 0.0
    tds_receivable_ledger_id: int | None = None
    if getattr(invoice, "apply_tds", False) and getattr(invoice, "tds_amount", None):
        tds_amount_value = float(invoice.tds_amount or 0)
        tds_receivable_ledger_id = getattr(invoice, "tds_ledger_id", None)

    customer_net = grand_total - tds_amount_value

    try:
        if existing_voucher:
            voucher = existing_voucher
            # Clear existing lines and allocations to rebuild them
            db.query(models.VoucherLine).filter(models.VoucherLine.voucher_id == voucher.id).delete()
            db.query(models.VoucherAllocation).filter(models.VoucherAllocation.voucher_id == voucher.id).delete()
            
            voucher.voucher_date = invoice.date
            voucher.bill_date = invoice.bill_date
            voucher.payment_mode_id = payment_mode_id
            voucher.narration = f"Sales invoice {invoice.reference or invoice.id}"
            # Preserve voucher_number, fiscal_year, and voucher_sequence
        else:
            voucher_number, fiscal_year, next_seq = get_next_voucher_number(
                db, company_id, models.VoucherType.SALES_INVOICE, invoice.date
            )
            voucher = models.Voucher(
                company_id=company_id,
                voucher_date=invoice.date,
                voucher_type=models.VoucherType.SALES_INVOICE,
                fiscal_year=fiscal_year,
                voucher_sequence=next_seq,
                voucher_number=voucher_number,
                narration=f"Sales invoice {invoice.reference or invoice.id}",
                payment_mode_id=payment_mode_id,
                bill_date=invoice.bill_date,
            )
            db.add(voucher)
        db.flush()
    except DataError as e:
        # Surface the underlying database error so we can diagnose the real cause.
        raise HTTPException(
            status_code=400,
            detail=f"Database error while saving SALES_INVOICE voucher: {e}",
        ) from e

    payment_mode: models.PaymentMode | None = None
    if payment_mode_id is not None:
        payment_mode = (
            db.query(models.PaymentMode)
            .filter(
                models.PaymentMode.id == payment_mode_id,
                models.PaymentMode.company_id == company_id,
                models.PaymentMode.is_active == True,
            )
            .first()
        )
        if not payment_mode:
            raise HTTPException(status_code=400, detail="Invalid payment_mode_id")

    is_credit_mode = bool(
        payment_mode is not None and payment_mode.name.strip().lower() == "credit"
    )

    # Always book the receivable to the customer ledger.
    counterparty_ledger_id = customer.ledger_id

    db.add(
        models.VoucherLine(
            voucher_id=voucher.id,
            ledger_id=counterparty_ledger_id,
            debit=customer_net,
            credit=0,
            department_id=invoice.department_id,
            project_id=invoice.project_id,
        )
    )

    # DR TDS Receivable ledger (if TDS is applied)
    if tds_amount_value and tds_receivable_ledger_id:
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=tds_receivable_ledger_id,
                debit=tds_amount_value,
                credit=0,
                department_id=invoice.department_id,
                project_id=invoice.project_id,
            )
        )

    # If the invoice is marked as paid via a cash/bank payment mode, also record
    # the settlement leg so the cash/bank ledger is affected and the customer
    # ledger shows both the invoice and the receipt.
    if payment_mode is not None and not is_credit_mode:
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=counterparty_ledger_id,
                debit=0,
                credit=customer_net,
                department_id=invoice.department_id,
                project_id=invoice.project_id,
            )
        )
        eff_pm_ledger_id = payment_ledger_id or payment_mode.ledger_id
        db.add(
            models.VoucherLine(
                voucher_id=voucher.id,
                ledger_id=eff_pm_ledger_id,
                debit=customer_net,
                credit=0,
                department_id=invoice.department_id,
                project_id=invoice.project_id,
            )
        )

    # CR income
    for (ledger_id, dept_id, proj_id), amount in income_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=0,
                    credit=amount,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    # CR tax
    for (ledger_id, dept_id, proj_id), amount in tax_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=0,
                    credit=amount,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    if cogs_totals:
        cogs_ledger_id = _get_default_cogs_ledger_id(db, company_id=company_id)
        stock_ledger_id = _get_default_stock_ledger_id(db, company_id=company_id)

        if cogs_ledger_id is None:
            raise HTTPException(
                status_code=400,
                detail="COGS ledger not found and could not be created (missing 'Direct Expenses' group).",
            )
        if stock_ledger_id is None:
            raise HTTPException(
                status_code=400,
                detail="Stock/Inventory ledger not found and could not be created (missing 'Stock-in-Hand' group).",
            )

        for (dept_id, proj_id), amount in cogs_totals.items():
            if amount:
                db.add(
                    models.VoucherLine(
                        voucher_id=voucher.id,
                        ledger_id=cogs_ledger_id,
                        debit=amount,
                        credit=0,
                        department_id=dept_id,
                        project_id=proj_id,
                    )
                )
                db.add(
                    models.VoucherLine(
                        voucher_id=voucher.id,
                        ledger_id=stock_ledger_id,
                        debit=0,
                        credit=amount,
                        department_id=dept_id,
                        project_id=proj_id,
                    )
                )

    # ── Incentive Postings ───────────────────────────────────────────
    # If incentive_amounts were passed, book them.
    # Entry: DR Incentive Expense (from rule/company), CR Incentive Payable (from company)
    if incentive_amounts:
        company = db.query(models.Company).filter(models.Company.id == company_id).first()
        default_expense_ledger_id = company.default_incentive_expense_ledger_id if company else None
        default_payable_ledger_id = getattr(company, "default_incentive_payable_ledger_id", None)
        
        # If no payable ledger is configured, we try to find/create one named "Sales Incentive Payable"
        if default_payable_ledger_id is None:
            group = db.query(models.LedgerGroup).filter(
                models.LedgerGroup.company_id == company_id,
                models.LedgerGroup.name == "Current Liabilities"
            ).first()
            
            payable_ledger = db.query(models.Ledger).filter(
                models.Ledger.company_id == company_id,
                models.Ledger.name == "Sales Incentive Payable"
            ).first()
            
            if not payable_ledger and group:
                payable_ledger = models.Ledger(
                    company_id=company_id,
                    group_id=group.id,
                    name="Sales Incentive Payable",
                    is_active=True,
                    opening_balance=0,
                    opening_balance_type=models.OpeningBalanceType.CREDIT
                )
                db.add(payable_ledger)
                db.flush()
            
            if payable_ledger:
                default_payable_ledger_id = payable_ledger.id

        if default_payable_ledger_id:
            # We aggregate by expense ledger
            # For simplicity in the voucher, we match rules here too if possible
            rules = db.query(models.IncentiveRule).filter(
                models.IncentiveRule.company_id == company_id,
                models.IncentiveRule.is_active == True
            ).all()

            for inc in incentive_amounts:
                if not inc.incentive_amount:
                    continue
                
                # Resolve Expense Ledger
                expense_ledger_id = None
                # Match rules for this sales person
                for r in rules:
                    if r.sales_person_id == inc.sales_person_id:
                        # Further matches could be added (item, etc), but for now sp-based rule is primary
                        if r.ledger_id:
                            expense_ledger_id = r.ledger_id
                            break
                
                if expense_ledger_id is None:
                    expense_ledger_id = default_expense_ledger_id
                
                if expense_ledger_id:
                    # DR Expense
                    db.add(
                        models.VoucherLine(
                            voucher_id=voucher.id,
                            ledger_id=expense_ledger_id,
                            debit=inc.incentive_amount,
                            credit=0,
                            department_id=invoice.department_id,
                            project_id=invoice.project_id,
                        )
                    )
                    # CR Payable
                    db.add(
                        models.VoucherLine(
                            voucher_id=voucher.id,
                            ledger_id=default_payable_ledger_id,
                            debit=0,
                            credit=inc.incentive_amount,
                            department_id=invoice.department_id,
                            project_id=invoice.project_id,
                        )
                    )

    return voucher


@router.get("/invoices", response_model=list[schemas.SalesInvoiceRead])
def list_invoices(
    company_id: int,
    voucher_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    total_subq = _compute_sales_invoice_total_subquery().label("total_amount")
    paid_subq = _compute_sales_invoice_paid_subquery(company_id=company_id).label("paid_amount")

    query = (
        db.query(
            models.SalesInvoice,
            models.SalesPerson.name.label("sales_person_name"),
            total_subq,
            paid_subq,
            models.Voucher.payment_mode_id,
            models.PaymentMode.name.label("payment_mode_name"),
        )
        .outerjoin(
            models.SalesPerson,
            models.SalesPerson.id == models.SalesInvoice.sales_person_id,
        )
        .outerjoin(
            models.Voucher,
            models.Voucher.id == models.SalesInvoice.voucher_id,
        )
        .outerjoin(
            models.PaymentMode,
            models.PaymentMode.id == models.Voucher.payment_mode_id,
        )
        .filter(models.SalesInvoice.company_id == company_id)
    )
    if voucher_id is not None:
        query = query.filter(models.SalesInvoice.voucher_id == voucher_id)
    rows = query.order_by(models.SalesInvoice.date.desc(), models.SalesInvoice.id.desc()).all()

    results: list[schemas.SalesInvoiceRead] = []
    for inv, sp_name, total_amount, paid_amount, payment_mode_id, payment_mode_name in rows:
        total = float(total_amount or 0)
        paid = float(paid_amount or 0)
        
        is_credit = (payment_mode_name is None or payment_mode_name.strip().lower() == "credit")
        if not is_credit:
            paid = total
            outstanding = 0.0
        else:
            outstanding = max(total - paid, 0.0)
        
        # For service invoices, if header sales person is empty, use the first line's sales person if available
        effective_sp_name = sp_name
        if not effective_sp_name and inv.invoice_type == "SERVICE" and inv.lines:
            # Find the first line that has a sales person
            for line in inv.lines:
                if line.sales_person_id:
                    # We need to fetch the employee name if not already joined
                    # For simplicity, we'll try to use the relationship if loaded
                    if hasattr(line, "sales_person") and line.sales_person:
                        effective_sp_name = line.sales_person.name
                        break
        
        base = schemas.SalesInvoiceRead.model_validate(inv)
        results.append(
            base.model_copy(
                update={
                    "sales_person_name": (str(effective_sp_name) if effective_sp_name is not None else None),
                    "paid_amount": paid,
                    "outstanding_amount": outstanding,
                    "payment_status": _payment_status(total_amount=total, paid_amount=paid, is_credit=is_credit),
                    "payment_mode_id": payment_mode_id,
                }
            )
        )
    return results


@router.post("/invoices", response_model=schemas.SalesInvoiceRead)
def create_invoice(
    company_id: int,
    invoice_in: schemas.SalesInvoiceCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> schemas.SalesInvoiceRead:
    company = _get_company(db, company_id, current_user)
    validate_transaction_date(company, invoice_in.date)

    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == invoice_in.customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")

    if not invoice_in.lines:
        raise HTTPException(status_code=400, detail="Invoice must have at least one line")

    # Resolve header-level sales ledger and output tax ledger from seeded
    # defaults when not explicitly provided.
    effective_sales_ledger_id = invoice_in.sales_ledger_id
    if effective_sales_ledger_id is None:
        effective_sales_ledger_id = getattr(company, "default_sales_ledger_id", None)

    if effective_sales_ledger_id is None:
        sales_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code == "SALES",
            )
            .first()
        )
        if sales_ledger is not None:
            effective_sales_ledger_id = sales_ledger.id

    effective_output_tax_ledger_id = invoice_in.output_tax_ledger_id
    if effective_output_tax_ledger_id is None:
        output_tax_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code.in_(["OUTPUT_TAX", "OUTPUT_VAT"]),
            )
            .first()
        )
        if output_tax_ledger is not None:
            effective_output_tax_ledger_id = output_tax_ledger.id

    sp = _validate_sales_person(
        db,
        company_id=company_id,
        sales_person_id=invoice_in.sales_person_id,
    )
    if invoice_in.sales_person_id is not None and sp is None:
        raise HTTPException(status_code=400, detail="Invalid sales_person_id")

    invoice = models.SalesInvoice(
        company_id=company_id,
        customer_id=invoice_in.customer_id,
        date=invoice_in.date,
        due_date=invoice_in.due_date or invoice_in.date,
        reference=invoice_in.reference,
        sales_person_id=invoice_in.sales_person_id,
        sales_ledger_id=effective_sales_ledger_id,
        output_tax_ledger_id=effective_output_tax_ledger_id,
        payment_ledger_id=invoice_in.payment_ledger_id,
        department_id=invoice_in.department_id,
        project_id=invoice_in.project_id,
        invoice_type=invoice_in.invoice_type,
        sales_type=invoice_in.sales_type,
        narration=invoice_in.narration,
        bill_date=invoice_in.bill_date,
        apply_tds=bool(getattr(invoice_in, "apply_tds", False)),
        tds_amount=getattr(invoice_in, "tds_amount", None),
        tds_ledger_id=getattr(invoice_in, "tds_ledger_id", None),
    )
    db.add(invoice)
    db.flush()

    # Validate stock availability before creating invoice
    # Service / non-stock items with allow_negative_stock=True are excluded
    # from stock checks so they can be sold without physical inventory.
    # Kit lines validate BOM components at the line warehouse (phantom kit).
    direct_item_ids = {int(line.item_id) for line in invoice_in.lines}
    base_item_rows = (
        db.query(
            models.Item.id,
            models.Item.allow_negative_stock,
            models.Item.category,
            models.Item.name,
            models.Item.code,
            models.Item.is_fixed_asset,
            models.Item.sell_as_kit,
        )
        .filter(models.Item.company_id == company_id, models.Item.id.in_(direct_item_ids))
        .all()
    )
    sell_as_kit_pre = {int(r.id): bool(r.sell_as_kit) for r in base_item_rows}
    kit_parent_ids = {i for i, v in sell_as_kit_pre.items() if v}
    extra_component_ids = _collect_kit_component_ids_for_lines(
        db,
        company_id=company_id,
        lines=invoice_in.lines,
        sell_as_kit_ids=kit_parent_ids,
        invoice_date=invoice_in.date,
    )
    all_item_ids = direct_item_ids | extra_component_ids

    item_flags = (
        db.query(
            models.Item.id,
            models.Item.allow_negative_stock,
            models.Item.category,
            models.Item.name,
            models.Item.code,
            models.Item.is_fixed_asset,
            models.Item.sell_as_kit,
            models.Item.opening_rate,
            models.Item.default_purchase_rate,
            models.Item.is_batch_tracked,
        )
        .filter(models.Item.company_id == company_id, models.Item.id.in_(all_item_ids))
        .all()
    )
    is_batch_tracked_map: dict[int, bool] = {int(row.id): bool(row.is_batch_tracked) for row in item_flags}
    allow_negative_map: dict[int, bool] = {int(row.id): bool(row.allow_negative_stock) for row in item_flags}
    is_service_map: dict[int, bool] = {
        int(row.id): (row.category and row.category.strip().lower() == "service") for row in item_flags
    }
    is_fixed_asset_map: dict[int, bool] = {int(row.id): bool(row.is_fixed_asset) for row in item_flags}
    sell_as_kit_map: dict[int, bool] = {int(row.id): bool(row.sell_as_kit) for row in item_flags}
    fallback_rate_map: dict[int, float] = {
        int(row.id): float(row.default_purchase_rate or row.opening_rate or 0) for row in item_flags
    }
    item_info_map = {int(row.id): {"name": row.name, "code": row.code} for row in item_flags}

    validation_default_warehouse = _resolve_sales_default_warehouse(db, company_id=company_id)
    if validation_default_warehouse is None and any(
        (is_batch_tracked_map.get(int(line.item_id)) or not (allow_negative_map.get(int(line.item_id)) or is_service_map.get(int(line.item_id))))
        for line in invoice_in.lines
    ):
        raise HTTPException(status_code=400, detail="No active warehouse found for stock validation")

    pair_quantities: dict[tuple[int, int], float] = {}
    for line in invoice_in.lines:
        # Stock validation only compulsory for tracked items that don't allow negative stock
        if not is_batch_tracked_map.get(int(line.item_id)) and (allow_negative_map.get(int(line.item_id)) or is_service_map.get(int(line.item_id))):
            continue
        wh_id = line.warehouse_id
        if wh_id is None:
            wh_id = int(validation_default_warehouse.id)

        if sell_as_kit_map.get(int(line.item_id)):
            bom = get_latest_bom_for_product(
                db, company_id=company_id, product_id=int(line.item_id), as_of=invoice_in.date
            )
            if bom is None or not bom.items:
                raise HTTPException(
                    status_code=400,
                    detail=f"Kit item has no active BOM (item_id={line.item_id}). Add a BOM or disable Sell as kit.",
                )
            comps = explode_flat_kit_components(bom=bom, kit_qty=float(line.quantity))
            for comp_id, rq in comps.items():
                key = (int(comp_id), int(wh_id))
                pair_quantities[key] = pair_quantities.get(key, 0.0) + float(rq)
        else:
            key = (int(line.item_id), int(wh_id))
            pair_quantities[key] = pair_quantities.get(key, 0.0) + float(line.quantity)

    svc = StockValuationService(db)
    tenant_settings = svc.get_tenant_settings(tenant_id=int(company.tenant_id))

    if pair_quantities and not invoice_in.bypass_stock_validation and not bool(getattr(tenant_settings, "allow_negative_stock", False)):
        pairs = list(pair_quantities.keys())
        batch_results = _compute_batch_stock(
            company_id=company_id,
            pairs=pairs,
            db=db,
        )
        available_map: dict[tuple[int, int], float] = {}
        for res in batch_results:
            available_map[(res.itemId, res.warehouseId)] = float(res.quantityOnHand)

        # Preload item and warehouse names for clearer error messages
        item_ids = {item_id for (item_id, _) in pair_quantities.keys()}
        warehouse_ids = {warehouse_id for (_, warehouse_id) in pair_quantities.keys()}

        item_rows = (
            db.query(models.Item.id, models.Item.name, models.Item.code)
            .filter(models.Item.company_id == company_id, models.Item.id.in_(item_ids))
            .all()
        )
        item_name_map: dict[int, dict[str, str | None]] = {
            row.id: {"name": row.name, "code": row.code} for row in item_rows
        }

        warehouse_rows = (
            db.query(models.Warehouse.id, models.Warehouse.name)
            .filter(models.Warehouse.company_id == company_id, models.Warehouse.id.in_(warehouse_ids))
            .all()
        )
        warehouse_name_map: dict[int, str] = {row.id: row.name for row in warehouse_rows}

        for (item_id, warehouse_id), required_qty in pair_quantities.items():
            available = available_map.get((item_id, warehouse_id), 0.0)
            if required_qty > available:
                item_info = item_name_map.get(item_id, {"name": None, "code": None})
                warehouse_name = warehouse_name_map.get(warehouse_id)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "error": "INSUFFICIENT_STOCK",
                        "message": "Insufficient stock for item in selected warehouse.",
                        "details": {
                            "item_id": item_id,
                            "item_name": item_info_map.get(item_id, {}).get("name"),
                            "item_code": item_info_map.get(item_id, {}).get("code"),
                            "warehouse_id": warehouse_id,
                            "warehouse_name": warehouse_name,
                            "required_quantity": required_qty,
                            "available_quantity": available,
                        },
                    },
                )

    for line in invoice_in.lines:
        is_service = is_service_map.get(line.item_id, False)
        effective_warehouse_id = line.warehouse_id

        if not is_service and effective_warehouse_id is None:
            if validation_default_warehouse is None:
                raise HTTPException(status_code=400, detail="No active warehouse found")
            effective_warehouse_id = int(validation_default_warehouse.id)

        if not is_service:
            warehouse = (
                db.query(models.Warehouse)
                .filter(
                    models.Warehouse.id == effective_warehouse_id,
                    models.Warehouse.company_id == company_id,
                    models.Warehouse.is_active == True,
                )
                .first()
            )
            if not warehouse:
                raise HTTPException(status_code=400, detail="Invalid warehouse_id")
        else:
            effective_warehouse_id = None

        invoice_line = models.SalesInvoiceLine(
            invoice_id=invoice.id,
            item_id=line.item_id,
            quantity=line.quantity,
            rate=line.rate,
            discount=line.discount,
            tax_rate=line.tax_rate,
            hs_code=line.hs_code,
            warehouse_id=effective_warehouse_id,
            sales_person_id=line.sales_person_id,
            department_id=line.department_id or invoice_in.department_id,
            project_id=line.project_id or invoice_in.project_id,
            ref_no=line.ref_no,
            remarks=line.remarks,
        )
        db.add(invoice_line)
        db.flush()

    inv_lines = (
        db.query(models.SalesInvoiceLine)
        .filter(models.SalesInvoiceLine.invoice_id == invoice.id)
        .order_by(models.SalesInvoiceLine.id.asc())
        .all()
    )
    _apply_stock_postings_for_sales_invoice(
        db,
        company_id=company_id,
        company=company,
        invoice=invoice,
        current_user_id=current_user.id,
        is_service_map=is_service_map,
        is_fixed_asset_map=is_fixed_asset_map,
        sell_as_kit_map=sell_as_kit_map,
        allow_negative_map=allow_negative_map,
        fallback_rate_map=fallback_rate_map,
        is_batch_tracked_map=is_batch_tracked_map,
        lines=inv_lines,
    )

    db.expire(invoice, ["lines"])

    voucher = _build_sales_voucher(
        db,
        company_id,
        invoice,
        invoice_in.payment_mode_id,
        sales_ledger_id=invoice.sales_ledger_id,
        output_tax_ledger_id=invoice.output_tax_ledger_id,
        payment_ledger_id=invoice.payment_ledger_id,
        incentive_amounts=invoice_in.sales_person_incentive_amounts,
    )

    # Link the created voucher back to the invoice so voucher_id is populated
    invoice.voucher_id = voucher.id

    # ── Save Sales Person Incentives ──────────────────────────────────
    if invoice_in.sales_person_incentive_amounts:
        for inc in invoice_in.sales_person_incentive_amounts:
            db.add(
                models.SalesInvoiceIncentive(
                    company_id=company_id,
                    invoice_id=invoice.id,
                    sales_person_id=inc.sales_person_id,
                    incentive_amount=inc.incentive_amount,
                    is_manual=inc.is_manual,
                    post_method=inc.post_method,
                )
            )
        db.flush()

    db.commit()
    db.refresh(invoice)

    # Trigger notification
    background_tasks.add_task(notification_service.notify_order_placed, db, invoice.id)

    total_amount = float(
        db.query(_compute_sales_invoice_total_subquery())
        .select_from(models.SalesInvoice)
        .filter(models.SalesInvoice.id == invoice.id)
        .scalar()
        or 0
    )
    paid_amount = float(
        db.query(func.coalesce(func.sum(models.VoucherAllocation.allocated_amount), 0))
        .filter(
            models.VoucherAllocation.company_id == company_id,
            models.VoucherAllocation.doc_type == models.AllocationDocType.SALES_INVOICE.value,
            models.VoucherAllocation.doc_id == invoice.id,
        )
        .scalar()
        or 0
    )

    pm_name = None
    if invoice_in.payment_mode_id:
        pm = db.query(models.PaymentMode).filter(models.PaymentMode.id == invoice_in.payment_mode_id).first()
        if pm:
            pm_name = pm.name
    
    is_credit = (pm_name is None or pm_name.strip().lower() == "credit")
    if not is_credit:
        paid_amount = total_amount
        outstanding = 0.0
    else:
        outstanding = max(total_amount - paid_amount, 0.0)

    sp_name = getattr(getattr(invoice, "sales_person", None), "name", None)
    base = schemas.SalesInvoiceRead.model_validate(invoice)
    return base.model_copy(
        update={
            "sales_person_name": (str(sp_name) if sp_name is not None else None),
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding,
            "payment_status": _payment_status(total_amount=total_amount, paid_amount=paid_amount, is_credit=is_credit),
            "payment_mode_id": invoice_in.payment_mode_id,
        }
    )


@router.get("/reports/sales-by-person", response_model=list[schemas.SalesByPersonReportRow])
def sales_by_person_report(
    company_id: int,
    sales_person_id: int | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    line = models.SalesInvoiceLine
    subtotal = (line.quantity * line.rate) - line.discount
    line_total = subtotal + (subtotal * (line.tax_rate / 100.0))

    inv_totals = (
        select(
            line.invoice_id.label("invoice_id"),
            func.coalesce(func.sum(line_total), 0).label("total_amount"),
        )
        .group_by(line.invoice_id)
        .subquery()
    )

    paid_totals = (
        select(
            models.VoucherAllocation.doc_id.label("invoice_id"),
            func.coalesce(func.sum(models.VoucherAllocation.allocated_amount), 0).label("paid_amount"),
        )
        .where(
            models.VoucherAllocation.company_id == company_id,
            models.VoucherAllocation.doc_type == models.AllocationDocType.SALES_INVOICE.value,
        )
        .group_by(models.VoucherAllocation.doc_id)
        .subquery()
    )

    q = (
        db.query(
            models.SalesInvoice.sales_person_id.label("sales_person_id"),
            models.SalesPerson.name.label("sales_person_name"),
            func.count(models.SalesInvoice.id).label("invoice_count"),
            func.coalesce(func.sum(inv_totals.c.total_amount), 0).label("total_sales_amount"),
            func.coalesce(
                func.sum(inv_totals.c.total_amount - func.coalesce(paid_totals.c.paid_amount, 0)),
                0,
            ).label("outstanding_amount"),
        )
        .join(inv_totals, inv_totals.c.invoice_id == models.SalesInvoice.id)
        .outerjoin(paid_totals, paid_totals.c.invoice_id == models.SalesInvoice.id)
        .outerjoin(models.SalesPerson, models.SalesPerson.id == models.SalesInvoice.sales_person_id)
        .filter(models.SalesInvoice.company_id == company_id)
    )

    if from_date is not None:
        q = q.filter(models.SalesInvoice.date >= from_date)
    if to_date is not None:
        q = q.filter(models.SalesInvoice.date <= to_date)
    if sales_person_id is not None:
        q = q.filter(models.SalesInvoice.sales_person_id == int(sales_person_id))
    q = q.group_by(models.SalesInvoice.sales_person_id, models.SalesPerson.name)
    q = q.order_by(func.coalesce(models.SalesPerson.name, "").asc())

    rows = q.all()
    out: list[schemas.SalesByPersonReportRow] = []
    for r in rows:
        out.append(
            schemas.SalesByPersonReportRow(
                sales_person_id=(int(r.sales_person_id) if r.sales_person_id is not None else None),
                sales_person_name=(str(r.sales_person_name) if r.sales_person_name is not None else None),
                invoice_count=int(r.invoice_count or 0),
                total_sales_amount=float(r.total_sales_amount or 0),
                outstanding_amount=float(r.outstanding_amount or 0),
            )
        )
    return out


@router.get("/reports/restaurant-summary", response_model=schemas.RestaurantSummaryResponse)
def restaurant_summary_report(
    company_id: int,
    from_date: date | None = None,
    to_date: date | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    line = models.SalesInvoiceLine
    subtotal = (line.quantity * line.rate) - line.discount
    line_total = subtotal + (subtotal * (line.tax_rate / 100.0))

    # Aggregate lines to get total amount and item count per invoice
    inv_stats = (
        db.query(
            line.invoice_id.label("invoice_id"),
            func.coalesce(func.sum(line_total), 0).label("total_amount"),
            func.coalesce(func.sum(line.quantity), 0).label("total_items"),
        )
        .group_by(line.invoice_id)
        .subquery()
    )

    # Summary by Order Type (stored in 'reference')
    by_type_q = (
        db.query(
            models.SalesInvoice.reference.label("order_type"),
            func.count(models.SalesInvoice.id).label("invoice_count"),
            func.sum(inv_stats.c.total_amount).label("total_amount"),
            func.sum(inv_stats.c.total_items).label("total_items"),
        )
        .join(inv_stats, inv_stats.c.invoice_id == models.SalesInvoice.id)
        .filter(models.SalesInvoice.company_id == company_id)
    )
    if from_date:
        by_type_q = by_type_q.filter(models.SalesInvoice.date >= from_date)
    if to_date:
        by_type_q = by_type_q.filter(models.SalesInvoice.date <= to_date)
    
    summary_by_type = by_type_q.group_by(models.SalesInvoice.reference).all()

    # Summary by Table (stored in 'custom_reference')
    by_table_q = (
        db.query(
            models.SalesInvoice.custom_reference.label("table_number"),
            func.count(models.SalesInvoice.id).label("invoice_count"),
            func.sum(inv_stats.c.total_amount).label("total_amount"),
            func.sum(inv_stats.c.total_items).label("total_items"),
        )
        .join(inv_stats, inv_stats.c.invoice_id == models.SalesInvoice.id)
        .filter(models.SalesInvoice.company_id == company_id)
        .filter(models.SalesInvoice.reference == "Dine-in") # Only tables for Dine-in
    )
    if from_date:
        by_table_q = by_table_q.filter(models.SalesInvoice.date >= from_date)
    if to_date:
        by_table_q = by_table_q.filter(models.SalesInvoice.date <= to_date)
    
    summary_by_table = by_table_q.group_by(models.SalesInvoice.custom_reference).all()

    # Overall totals
    totals_q = (
        db.query(
            func.sum(inv_stats.c.total_amount).label("total_sales"),
            func.count(models.SalesInvoice.id).label("total_orders"),
        )
        .join(inv_stats, inv_stats.c.invoice_id == models.SalesInvoice.id)
        .filter(models.SalesInvoice.company_id == company_id)
    )
    if from_date:
        totals_q = totals_q.filter(models.SalesInvoice.date >= from_date)
    if to_date:
        totals_q = totals_q.filter(models.SalesInvoice.date <= to_date)
    
    overall = totals_q.first()
    total_sales = overall.total_sales if overall and overall.total_sales else 0
    total_orders = overall.total_orders if overall and overall.total_orders else 0

    return {
        "summary_by_type": [
            {
                "order_type": r.order_type or "Unknown",
                "invoice_count": r.invoice_count,
                "total_amount": float(r.total_amount or 0),
                "total_items": float(r.total_items or 0),
            }
            for r in summary_by_type
        ],
        "summary_by_table": [
            {
                "order_type": "Dine-in",
                "table_number": r.table_number or "Unknown",
                "invoice_count": r.invoice_count,
                "total_amount": float(r.total_amount or 0),
                "total_items": float(r.total_items or 0),
            }
            for r in summary_by_table
        ],
        "total_sales": float(total_sales),
        "total_orders": int(total_orders),
    }


@router.get("/invoices/by-reference/{reference}", response_model=schemas.SalesInvoiceRead)
def get_invoice_by_reference(
    company_id: int,
    reference: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    total_subq = _compute_sales_invoice_total_subquery().label("total_amount")
    paid_subq = _compute_sales_invoice_paid_subquery(company_id=company_id).label("paid_amount")
    row = (
        db.query(
            models.SalesInvoice,
            models.SalesPerson.name.label("sales_person_name"),
            total_subq,
            paid_subq,
            models.Voucher.payment_mode_id,
            models.PaymentMode.name.label("payment_mode_name"),
        )
        .outerjoin(
            models.SalesPerson,
            models.SalesPerson.id == models.SalesInvoice.sales_person_id,
        )
        .outerjoin(
            models.Voucher,
            models.Voucher.id == models.SalesInvoice.voucher_id,
        )
        .outerjoin(
            models.PaymentMode,
            models.PaymentMode.id == models.Voucher.payment_mode_id,
        )
        .filter(
            models.SalesInvoice.reference == reference,
            models.SalesInvoice.company_id == company_id,
        )
        .order_by(models.SalesInvoice.date.desc())
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Invoice not found with that reference")
    invoice, sp_name, total_amount, paid_amount, payment_mode_id, payment_mode_name = row
    total = float(total_amount or 0)
    paid = float(paid_amount or 0)
    
    is_credit = (payment_mode_name is None or payment_mode_name.strip().lower() == "credit")
    if not is_credit:
        paid = total
        outstanding = 0.0
    else:
        # Calculate FIFO outstanding for this invoice
        customer = db.query(models.Customer).filter(models.Customer.id == invoice.customer_id).first()
        lid = customer.ledger_id if customer else None
        
        if lid:
            # 1. Total closing balance
            l = db.query(models.Ledger).filter(models.Ledger.id == lid).first()
            ob = float(l.opening_balance or 0) if l else 0.0
            if l and l.opening_balance_type == models.OpeningBalanceType.CREDIT:
                ob = -ob
                
            vl_sum = (
                db.query(
                    func.sum(models.VoucherLine.debit).label("total_debit"),
                    func.sum(models.VoucherLine.credit).label("total_credit"),
                )
                .join(models.Voucher, models.Voucher.id == models.VoucherLine.voucher_id)
                .filter(
                    models.VoucherLine.ledger_id == lid,
                    models.Voucher.company_id == company_id
                )
                .first()
            )
            closing_balance = ob + float(vl_sum.total_debit or 0) - float(vl_sum.total_credit or 0)
            
            # 2. Total Billed for this ledger
            inv_total_subq = _compute_sales_invoice_total_subquery().label("total_amount")
            all_invoices = (
                db.query(models.SalesInvoice, inv_total_subq, _compute_sales_invoice_paid_subquery(company_id=company_id).label("explicit_paid_amount"))
                .join(models.Customer, models.Customer.id == models.SalesInvoice.customer_id)
                .filter(models.SalesInvoice.company_id == company_id, models.Customer.ledger_id == lid)
                .order_by(models.SalesInvoice.date.asc(), models.SalesInvoice.id.asc())
                .all()
            )
            
            total_billed = sum(float(x.total_amount or 0) for x in all_invoices)
            total_explicit = sum(float(x.explicit_paid_amount or 0) for x in all_invoices)
            
            unallocated_pool = max(0.0, (total_billed + ob) - closing_balance - total_explicit)
            
            # 3. FIFO cascade
            paid = float(paid_amount or 0)
            outstanding = max(total - paid, 0.0)
            for inv, inv_tot, inv_exp in all_invoices:
                i_tot = float(inv_tot or 0)
                i_exp = float(inv_exp or 0)
                remain = max(0.0, i_tot - i_exp)
                
                applied = 0.0
                if remain > 0 and unallocated_pool > 0:
                    applied = min(remain, unallocated_pool)
                    unallocated_pool -= applied
                
                if inv.id == invoice.id:
                    paid = i_exp + applied
                    outstanding = max(0.0, i_tot - paid)
                    break
        else:
            outstanding = max(total - paid, 0.0)
    
    base = schemas.SalesInvoiceRead.model_validate(invoice)
    return base.model_copy(
        update={
            "sales_person_name": (str(sp_name) if sp_name is not None else None),
            "paid_amount": paid,
            "outstanding_amount": outstanding,
            "payment_status": _payment_status(total_amount=total, paid_amount=paid, is_credit=is_credit),
            "payment_mode_id": payment_mode_id,
        }
    )

@router.get("/invoices/{invoice_id}", response_model=schemas.SalesInvoiceRead)
def get_invoice(
    company_id: int,
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    total_subq = _compute_sales_invoice_total_subquery().label("total_amount")
    paid_subq = _compute_sales_invoice_paid_subquery(company_id=company_id).label("paid_amount")
    row = (
        db.query(
            models.SalesInvoice,
            models.SalesPerson.name.label("sales_person_name"),
            total_subq,
            paid_subq,
            models.Voucher.payment_mode_id,
            models.PaymentMode.name.label("payment_mode_name"),
        )
        .outerjoin(
            models.SalesPerson,
            models.SalesPerson.id == models.SalesInvoice.sales_person_id,
        )
        .outerjoin(
            models.Voucher,
            models.Voucher.id == models.SalesInvoice.voucher_id,
        )
        .outerjoin(
            models.PaymentMode,
            models.PaymentMode.id == models.Voucher.payment_mode_id,
        )
        .filter(
            models.SalesInvoice.id == invoice_id,
            models.SalesInvoice.company_id == company_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Invoice not found")
    invoice, sp_name, total_amount, paid_amount, payment_mode_id, payment_mode_name = row
    total = float(total_amount or 0)
    paid = float(paid_amount or 0)
    
    is_credit = (payment_mode_name is None or payment_mode_name.strip().lower() == "credit")
    if not is_credit:
        paid = total
        outstanding = 0.0
    else:
        outstanding = max(total - paid, 0.0)
    
    base = schemas.SalesInvoiceRead.model_validate(invoice)
    return base.model_copy(
        update={
            "sales_person_name": (str(sp_name) if sp_name is not None else None),
            "paid_amount": paid,
            "outstanding_amount": outstanding,
            "payment_status": _payment_status(total_amount=total, paid_amount=paid, is_credit=is_credit),
            "payment_mode_id": payment_mode_id,
        }
    )



def recharge_stock_from_sales_invoice(db: Session, company_id: int, invoice: models.SalesInvoice, current_user_id: int | None = None):
    """Clear prior SALES_INVOICE stock rows and repost from current lines (phantom kits = components)."""
    _clear_sales_invoice_stock_postings(db, company_id=company_id, invoice_id=int(invoice.id))

    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company:
        return

    inv_lines = (
        db.query(models.SalesInvoiceLine)
        .filter(models.SalesInvoiceLine.invoice_id == invoice.id)
        .order_by(models.SalesInvoiceLine.id.asc())
        .all()
    )
    if not inv_lines:
        return

    direct_item_ids = {int(l.item_id) for l in inv_lines}
    base_rows = (
        db.query(models.Item.id, models.Item.sell_as_kit)
        .filter(models.Item.company_id == company_id, models.Item.id.in_(direct_item_ids))
        .all()
    )
    sell_as_kit_pre = {int(r.id): bool(r.sell_as_kit) for r in base_rows}
    kit_parent_ids = {i for i, v in sell_as_kit_pre.items() if v}
    extra_component_ids = _collect_kit_component_ids_for_lines(
        db,
        company_id=company_id,
        lines=inv_lines,
        sell_as_kit_ids=kit_parent_ids,
        invoice_date=invoice.date,
    )
    all_item_ids = direct_item_ids | extra_component_ids

    item_flags = (
        db.query(
            models.Item.id,
            models.Item.category,
            models.Item.is_fixed_asset,
            models.Item.sell_as_kit,
        )
        .filter(models.Item.company_id == company_id, models.Item.id.in_(all_item_ids))
        .all()
    )
    is_service_map = {
        int(r.id): (r.category and r.category.strip().lower() == "service") for r in item_flags
    }
    is_fixed_asset_map = {int(r.id): bool(r.is_fixed_asset) for r in item_flags}
    sell_as_kit_map = {int(r.id): bool(r.sell_as_kit) for r in item_flags}

    _apply_stock_postings_for_sales_invoice(
        db,
        company_id=company_id,
        company=company,
        invoice=invoice,
        current_user_id=current_user_id,
        is_service_map=is_service_map,
        is_fixed_asset_map=is_fixed_asset_map,
        sell_as_kit_map=sell_as_kit_map,
        lines=inv_lines,
    )


@router.put("/invoices/{invoice_id}", response_model=schemas.SalesInvoiceRead)
def update_invoice(
    company_id: int,
    invoice_id: int,
    invoice_in: schemas.SalesInvoiceUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    print(f"DEBUG: update_invoice - id={invoice_id}, incentives={invoice_in.sales_person_incentive_amounts}")
    company = _get_company(db, company_id, current_user)
    if invoice_in.date:
        validate_transaction_date(company, invoice_in.date)
    invoice = (
        db.query(models.SalesInvoice)
        .filter(
            models.SalesInvoice.id == invoice_id,
            models.SalesInvoice.company_id == company_id,
        )
        .first()
    )
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    _logger.debug("update_invoice %s", invoice_id)
    if invoice_in.lines and len(invoice_in.lines) > 0:
        _logger.debug("First line ref: %s", invoice_in.lines[0].ref_no)
        _logger.debug("First line sp: %s", invoice_in.lines[0].sales_person_id)

    old_date = invoice.date

    update_data = invoice_in.model_dump(
        exclude_unset=True,
        exclude={"lines"},
    )

    if "sales_person_id" in update_data:
        sp = _validate_sales_person(
            db,
            company_id=company_id,
            sales_person_id=invoice_in.sales_person_id,
        )
        if invoice_in.sales_person_id is not None and sp is None:
            raise HTTPException(status_code=400, detail="Invalid sales_person_id")

    for field, value in update_data.items():
        # Do not overwrite non-nullable fields with None; treat None as "no change".
        if field in {"customer_id", "date"} and value is None:
            continue
        # For header-level ledgers, explicitly preserve existing values when None is sent.
        if field in {"sales_ledger_id", "output_tax_ledger_id"} and value is None:
            continue
        setattr(invoice, field, value)

    explicit_due_date_sent = "due_date" in update_data
    if explicit_due_date_sent and invoice_in.due_date is None:
        invoice.due_date = invoice.date
    elif ("date" in update_data) and not explicit_due_date_sent:
        if invoice.due_date is None or invoice.due_date == old_date:
            invoice.due_date = invoice.date
    elif invoice.due_date is None:
        invoice.due_date = invoice.date

    if invoice_in.lines is not None:
        # Delete old invoice lines (we'll recreate them)
        db.query(models.SalesInvoiceLine).filter(
            models.SalesInvoiceLine.invoice_id == invoice.id
        ).delete()

        # Validate stock (kit lines = BOM components at line warehouse; default warehouse = Main else first active).
        direct_item_ids = {int(line.item_id) for line in invoice_in.lines}
        validation_default_warehouse = _resolve_sales_default_warehouse(db, company_id=company_id)
        allow_negative_map: dict[int, bool] = {}
        is_service_map: dict[int, bool] = {}
        is_fixed_asset_map: dict[int, bool] = {}
        sell_as_kit_map: dict[int, bool] = {}
        item_info_map: dict[int, dict[str, str | None]] = {}
        pair_quantities: dict[tuple[int, int], float] = {}

        if direct_item_ids:
            base_item_rows = (
                db.query(
                    models.Item.id,
                    models.Item.allow_negative_stock,
                    models.Item.category,
                    models.Item.name,
                    models.Item.code,
                    models.Item.is_fixed_asset,
                    models.Item.sell_as_kit,
                )
                .filter(models.Item.company_id == company_id, models.Item.id.in_(direct_item_ids))
                .all()
            )
            sell_as_kit_pre = {int(r.id): bool(r.sell_as_kit) for r in base_item_rows}
            kit_parent_ids = {i for i, v in sell_as_kit_pre.items() if v}
            extra_component_ids = _collect_kit_component_ids_for_lines(
                db,
                company_id=company_id,
                lines=invoice_in.lines,
                sell_as_kit_ids=kit_parent_ids,
                invoice_date=invoice.date,
            )
            all_item_ids = direct_item_ids | extra_component_ids

            item_flags = (
                db.query(
                    models.Item.id,
                    models.Item.allow_negative_stock,
                    models.Item.category,
                    models.Item.name,
                    models.Item.code,
                    models.Item.is_fixed_asset,
                    models.Item.sell_as_kit,
                )
                .filter(models.Item.company_id == company_id, models.Item.id.in_(all_item_ids))
                .all()
            )
            allow_negative_map = {int(row.id): bool(row.allow_negative_stock) for row in item_flags}
            is_service_map = {
                int(row.id): (row.category and row.category.strip().lower() == "service") for row in item_flags
            }
            is_fixed_asset_map = {int(row.id): bool(row.is_fixed_asset) for row in item_flags}
            sell_as_kit_map = {int(row.id): bool(row.sell_as_kit) for row in item_flags}
            item_info_map = {int(row.id): {"name": row.name, "code": row.code} for row in item_flags}

            if validation_default_warehouse is None and any(
                not (allow_negative_map.get(int(line.item_id)) or is_service_map.get(int(line.item_id)))
                for line in invoice_in.lines
            ):
                raise HTTPException(status_code=400, detail="No active warehouse found for stock validation")

            for line in invoice_in.lines:
                if allow_negative_map.get(int(line.item_id)) or is_service_map.get(int(line.item_id)):
                    continue
                wh_id = line.warehouse_id
                if wh_id is None:
                    wh_id = int(validation_default_warehouse.id)

                if sell_as_kit_map.get(int(line.item_id)):
                    bom = get_latest_bom_for_product(
                        db, company_id=company_id, product_id=int(line.item_id), as_of=invoice.date
                    )
                    if bom is None or not bom.items:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Kit item has no active BOM (item_id={line.item_id}). Add a BOM or disable Sell as kit.",
                        )
                    comps = explode_flat_kit_components(bom=bom, kit_qty=float(line.quantity))
                    for comp_id, rq in comps.items():
                        key = (int(comp_id), int(wh_id))
                        pair_quantities[key] = pair_quantities.get(key, 0.0) + float(rq)
                else:
                    key = (int(line.item_id), int(wh_id))
                    pair_quantities[key] = pair_quantities.get(key, 0.0) + float(line.quantity)

        svc = StockValuationService(db)
        tenant_settings = svc.get_tenant_settings(tenant_id=int(company.tenant_id))
        if (
            pair_quantities
            and not bool(getattr(invoice_in, "bypass_stock_validation", False))
            and not bool(getattr(tenant_settings, "allow_negative_stock", False))
        ):
            pairs = list(pair_quantities.keys())
            batch_results = _compute_batch_stock(
                company_id=company_id,
                pairs=pairs,
                db=db,
            )
            available_map: dict[tuple[int, int], float] = {}
            for res in batch_results:
                available_map[(res.itemId, res.warehouseId)] = float(res.quantityOnHand)

            item_ids = {item_id for (item_id, _) in pair_quantities.keys()}
            warehouse_ids = {warehouse_id for (_, warehouse_id) in pair_quantities.keys()}

            item_rows = (
                db.query(models.Item.id, models.Item.name, models.Item.code)
                .filter(models.Item.company_id == company_id, models.Item.id.in_(item_ids))
                .all()
            )
            item_name_map: dict[int, dict[str, str | None]] = {
                row.id: {"name": row.name, "code": row.code} for row in item_rows
            }

            warehouse_rows = (
                db.query(models.Warehouse.id, models.Warehouse.name)
                .filter(models.Warehouse.company_id == company_id, models.Warehouse.id.in_(warehouse_ids))
                .all()
            )
            warehouse_name_map: dict[int, str] = {row.id: row.name for row in warehouse_rows}

            for (item_id, warehouse_id), required_qty in pair_quantities.items():
                available = available_map.get((item_id, warehouse_id), 0.0)
                allow_negative = allow_negative_map.get(item_id, False)

                if required_qty > available and not allow_negative:
                    item_info = item_name_map.get(item_id, {"name": None, "code": None})
                    warehouse_name = warehouse_name_map.get(warehouse_id)
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "error": "INSUFFICIENT_STOCK",
                            "message": "Insufficient stock for item in selected warehouse.",
                            "details": {
                                "item_id": item_id,
                                "item_name": item_info_map.get(item_id, {}).get("name") or item_info.get("name"),
                                "item_code": item_info_map.get(item_id, {}).get("code") or item_info.get("code"),
                                "warehouse_id": warehouse_id,
                                "warehouse_name": warehouse_name,
                                "required_quantity": required_qty,
                                "available_quantity": available,
                            },
                        },
                    )

        for line in invoice_in.lines:
            is_service = is_service_map.get(line.item_id, False)
            effective_warehouse_id = line.warehouse_id

            if not is_service and effective_warehouse_id is None:
                if validation_default_warehouse is None:
                    raise HTTPException(status_code=400, detail="No active warehouse found")
                effective_warehouse_id = int(validation_default_warehouse.id)

            if not is_service:
                warehouse = (
                    db.query(models.Warehouse)
                    .filter(
                        models.Warehouse.id == effective_warehouse_id,
                        models.Warehouse.company_id == company_id,
                        models.Warehouse.is_active == True,
                    )
                    .first()
                )
                if not warehouse:
                    raise HTTPException(status_code=400, detail="Invalid warehouse_id")
            else:
                effective_warehouse_id = None

            invoice_line = models.SalesInvoiceLine(
                invoice_id=invoice.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=line.rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
                warehouse_id=effective_warehouse_id,
                sales_person_id=line.sales_person_id,
                remarks=line.remarks,
                department_id=line.department_id or invoice.department_id,
                project_id=line.project_id or invoice.project_id,
                ref_no=line.ref_no,
            )
            db.add(invoice_line)
    db.flush()
    
    # Refresh stock entries (Cost fix: ensure unit_cost is populated even if lines didn't change)
    recharge_stock_from_sales_invoice(db, company_id, invoice, current_user.id)
    db.flush()

    # Rebuild or create the linked voucher so that accounting entries and
    # payment mode stay in sync with the invoice.
    existing_voucher: models.Voucher | None = None
    if invoice.voucher_id is not None:
        existing_voucher = (
            db.query(models.Voucher)
            .filter(
                models.Voucher.id == invoice.voucher_id,
                models.Voucher.company_id == company_id,
            )
            .first()
        )

    # Decide which payment_mode_id to apply to the rebuilt voucher:
    # - If the client explicitly sent payment_mode_id in this update
    #   payload (even if null), honor that.
    # - Otherwise, preserve the existing voucher.payment_mode_id when
    #   present so updates that do not touch payment mode keep it.
    explicit_payment_mode_sent = "payment_mode_id" in update_data
    effective_payment_mode_id: int | None = None
    if explicit_payment_mode_sent:
        effective_payment_mode_id = invoice_in.payment_mode_id
    elif existing_voucher is not None:
        effective_payment_mode_id = existing_voucher.payment_mode_id
    
    explicit_payment_ledger_sent = "payment_ledger_id" in update_data
    effective_payment_ledger_id: int | None = None
    if explicit_payment_ledger_sent:
        effective_payment_ledger_id = invoice_in.payment_ledger_id
    else:
        effective_payment_ledger_id = invoice.payment_ledger_id

    # Ensure header-level ledgers are resolved using the same rules as on
    # create so voucher lines remain consistent with invoice headers.
    if invoice.sales_ledger_id is None:
        sales_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code == "SALES",
            )
            .first()
        )
        if sales_ledger is not None:
            invoice.sales_ledger_id = sales_ledger.id

    if invoice.output_tax_ledger_id is None:
        output_tax_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code.in_(["OUTPUT_TAX", "OUTPUT_VAT"]),
            )
            .first()
        )
        if output_tax_ledger is not None:
            invoice.output_tax_ledger_id = output_tax_ledger.id

    # Rebuild or create the linked voucher so that accounting entries and
    # payment mode stay in sync with the invoice.
    if invoice.lines:
        # Ensure relationships are fresh
        db.refresh(invoice)
        voucher = _build_sales_voucher(
            db,
            company_id,
            invoice,
            effective_payment_mode_id,
            sales_ledger_id=invoice.sales_ledger_id,
            output_tax_ledger_id=invoice.output_tax_ledger_id,
            payment_ledger_id=effective_payment_ledger_id,
            existing_voucher=existing_voucher,
            incentive_amounts=invoice_in.sales_person_incentive_amounts,
        )
        invoice.voucher_id = voucher.id
    elif existing_voucher is not None:
        invoice.voucher_id = None
        db.delete(existing_voucher)
        db.flush()

    # ── Update Sales Person Incentives ────────────────────────────────
    if invoice_in.sales_person_incentive_amounts is not None:
        db.query(models.SalesInvoiceIncentive).filter(
            models.SalesInvoiceIncentive.invoice_id == invoice.id
        ).delete()
        
        for inc in invoice_in.sales_person_incentive_amounts:
            db.add(
                models.SalesInvoiceIncentive(
                    company_id=company_id,
                    invoice_id=invoice.id,
                    sales_person_id=inc.sales_person_id,
                    incentive_amount=inc.incentive_amount,
                    is_manual=inc.is_manual,
                    post_method=inc.post_method,
                )
            )
        db.flush()

    db.commit()
    db.refresh(invoice)

    total_amount = float(
        db.query(_compute_sales_invoice_total_subquery())
        .select_from(models.SalesInvoice)
        .filter(models.SalesInvoice.id == invoice.id)
        .scalar()
        or 0
    )
    paid_amount = float(
        db.query(func.coalesce(func.sum(models.VoucherAllocation.allocated_amount), 0))
        .filter(
            models.VoucherAllocation.company_id == company_id,
            models.VoucherAllocation.doc_type == models.AllocationDocType.SALES_INVOICE.value,
            models.VoucherAllocation.doc_id == invoice.id,
        )
        .scalar()
        or 0
    )

    pm_name = None
    if effective_payment_mode_id:
        pm = db.query(models.PaymentMode).filter(models.PaymentMode.id == effective_payment_mode_id).first()
        if pm:
            pm_name = pm.name
    
    is_credit = (pm_name is None or pm_name.strip().lower() == "credit")
    if not is_credit:
        paid_amount = total_amount
        outstanding = 0.0
    else:
        outstanding = max(total_amount - paid_amount, 0.0)

    sp_name = getattr(getattr(invoice, "sales_person", None), "name", None)
    base = schemas.SalesInvoiceRead.model_validate(invoice)
    return base.model_copy(
        update={
            "sales_person_name": (str(sp_name) if sp_name is not None else None),
            "paid_amount": paid_amount,
            "outstanding_amount": outstanding,
            "payment_status": _payment_status(total_amount=total_amount, paid_amount=paid_amount, is_credit=is_credit),
            "payment_mode_id": effective_payment_mode_id,
        }
    )


@router.delete("/invoices/{invoice_id}")
def delete_invoice(
    company_id: int,
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    invoice = (
        db.query(models.SalesInvoice)
        .filter(
            models.SalesInvoice.id == invoice_id,
            models.SalesInvoice.company_id == company_id,
        )
        .first()
    )
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    linked_order = (
        db.query(models.SalesOrder)
        .filter(
            models.SalesOrder.company_id == company_id,
            models.SalesOrder.converted_to_invoice_id == invoice.id,
        )
        .first()
    )

    # Cascade behavior: if an invoice was created by converting a sales order,
    # unlink it so the invoice can be deleted.
    if linked_order is not None:
        linked_order.converted_to_invoice_id = None
        db.flush()

    _reverse_stock_ledger(
        db=db,
        company_id=company_id,
        source_type="SALES_INVOICE",
        source_id=invoice.id,
        created_by=current_user.id,
    )

    db.query(models.StockMovement).filter(
        models.StockMovement.company_id == company_id,
        models.StockMovement.source_type == "SALES_INVOICE",
        models.StockMovement.source_id == invoice.id,
    ).delete()

    # If the invoice created a voucher, delete it as part of invoice deletion.
    if invoice.voucher_id is not None:
        voucher = (
            db.query(models.Voucher)
            .filter(
                models.Voucher.id == invoice.voucher_id,
                models.Voucher.company_id == company_id,
            )
            .first()
        )
        invoice.voucher_id = None
        db.flush()
        if voucher is not None:
            db.delete(voucher)
            db.flush()

    db.delete(invoice)
    db.commit()
    return {"detail": "Deleted"}


# -------- Sales Returns --------


def _build_sales_return_voucher(
    db: Session,
    company_id: int,
    sales_return: models.SalesReturn,
    payment_mode_id: int | None,
    sales_return_ledger_id: int | None = None,
    output_tax_return_ledger_id: int | None = None,
    payment_ledger_id: int | None = None,
) -> models.Voucher:
    customer = (
        db.query(models.Customer)
        .filter(
            models.Customer.id == sales_return.customer_id,
            models.Customer.company_id == company_id,
        )
        .first()
    )
    if not customer:
        raise HTTPException(status_code=400, detail="Customer not found")

    # Aggregate lines by income, tax, and dimensions
    income_totals: dict[tuple[int, int | None, int | None], float] = {}
    tax_totals: dict[tuple[int, int | None, int | None], float] = {}
    grand_total = 0.0

    for line in sales_return.lines:
        item = (
            db.query(models.Item)
            .filter(
                models.Item.id == line.item_id,
                models.Item.company_id == company_id,
            )
            .first()
        )
        if not item:
            raise HTTPException(status_code=400, detail=f"Item {line.item_id} not found")
        
        # Resolve dimensions for this line (Hub & Spoke fallback)
        warehouse_dept_id = line.warehouse.department_id if line.warehouse else None
        warehouse_proj_id = line.warehouse.project_id if line.warehouse else None
        line_dept_id = line.department_id or warehouse_dept_id or sales_return.department_id
        line_proj_id = line.project_id or warehouse_proj_id or sales_return.project_id
        
        subtotal = float(line.quantity) * float(line.rate) - float(line.discount)
        tax = subtotal * float(line.tax_rate) / 100.0
        grand_total += subtotal + tax

        # Income aggregation
        is_fixed_asset = bool(getattr(item, "is_fixed_asset", False))
        if is_fixed_asset and item.income_ledger_id:
            eff_sales_ledger_id = item.income_ledger_id
        else:
            eff_sales_ledger_id = sales_return_ledger_id or item.income_ledger_id
            
        if not eff_sales_ledger_id:
            raise HTTPException(status_code=400, detail="Item missing income ledger")
        income_key = (eff_sales_ledger_id, line_dept_id, line_proj_id)
        income_totals[income_key] = income_totals.get(income_key, 0.0) + subtotal

        # Tax aggregation
        if tax:
            eff_output_tax_ledger_id = output_tax_return_ledger_id or item.output_tax_ledger_id
            if not eff_output_tax_ledger_id:
                raise HTTPException(status_code=400, detail="Item missing output tax ledger")
            tax_key = (eff_output_tax_ledger_id, line_dept_id, line_proj_id)
            tax_totals[tax_key] = tax_totals.get(tax_key, 0.0) + tax

    voucher_number, fiscal_year, next_seq = get_next_voucher_number(
        db, company_id, models.VoucherType.SALES_RETURN, sales_return.date
    )
    voucher = models.Voucher(
        company_id=company_id,
        voucher_date=sales_return.date,
        voucher_type=models.VoucherType.SALES_RETURN,
        fiscal_year=fiscal_year,
        voucher_sequence=next_seq,
        voucher_number=voucher_number,
        narration=f"Sales return {sales_return.reference or sales_return.id}",
        payment_mode_id=payment_mode_id,
    )
    db.add(voucher)
    db.flush()

    pm_ledger_id = payment_ledger_id
    if pm_ledger_id is None:
        pm_ledger_id = _get_counterparty_ledger_for_payment_mode(
            db,
            company_id,
            payment_mode_id=payment_mode_id,
            fallback_ledger_id=customer.ledger_id,
        )

    db.add(
        models.VoucherLine(
            voucher_id=voucher.id,
            ledger_id=pm_ledger_id,
            debit=0,
            credit=grand_total,
            department_id=sales_return.department_id,
            project_id=sales_return.project_id,
        )
    )

    # DR income
    for (ledger_id, dept_id, proj_id), amount in income_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=amount,
                    credit=0,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    # DR tax
    for (ledger_id, dept_id, proj_id), amount in tax_totals.items():
        if amount:
            db.add(
                models.VoucherLine(
                    voucher_id=voucher.id,
                    ledger_id=ledger_id,
                    debit=amount,
                    credit=0,
                    department_id=dept_id,
                    project_id=proj_id,
                )
            )

    return voucher


@router.post("/returns", response_model=schemas.SalesReturnRead)
def create_sales_return(
    company_id: int,
    return_in: schemas.SalesReturnCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    validate_transaction_date(company, return_in.date)

    # Resolve header-level sales return and output tax return ledgers from
    # seeded defaults when not explicitly provided. This mirrors the behavior
    # used for sales invoices so frontends can safely send null values.
    effective_sales_return_ledger_id = return_in.sales_return_ledger_id
    if effective_sales_return_ledger_id is None:
        # Prefer a dedicated SALES_RETURN ledger when present, otherwise fall
        # back to the main SALES ledger.
        sales_return_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code.in_(["SALES_RETURN", "SALES"]),
            )
            .order_by(models.Ledger.code.desc())
            .first()
        )
        if sales_return_ledger is not None:
            effective_sales_return_ledger_id = sales_return_ledger.id

    effective_output_tax_return_ledger_id = return_in.output_tax_return_ledger_id
    if effective_output_tax_return_ledger_id is None:
        output_tax_return_ledger = (
            db.query(models.Ledger)
            .filter(
                models.Ledger.company_id == company_id,
                models.Ledger.code.in_(["OUTPUT_TAX", "OUTPUT_VAT"]),
            )
            .first()
        )
        if output_tax_return_ledger is not None:
            effective_output_tax_return_ledger_id = output_tax_return_ledger.id

    sales_return = models.SalesReturn(
        company_id=company_id,
        customer_id=return_in.customer_id,
        date=return_in.date,
        reference=return_in.reference,
        sales_return_ledger_id=effective_sales_return_ledger_id,
        output_tax_return_ledger_id=effective_output_tax_return_ledger_id,
        source_invoice_id=return_in.source_invoice_id,
        department_id=return_in.department_id,
        project_id=return_in.project_id,
        payment_ledger_id=return_in.payment_ledger_id,
    )
    db.add(sales_return)
    db.flush()

    validation_default_warehouse = _resolve_sales_default_warehouse(db, company_id=company_id)

    svc = StockValuationService(db)
    method = svc.get_inventory_valuation_method(tenant_id=int(company.tenant_id))

    item_ids = {line.item_id for line in return_in.lines}
    item_settings = (
        db.query(models.Item.id, models.Item.category, models.Item.is_fixed_asset)
        .filter(models.Item.company_id == company_id, models.Item.id.in_(item_ids))
        .all()
    )
    is_service_map = {row.id: (row.category and row.category.strip().lower() == "service") for row in item_settings}
    is_fixed_asset_map = {row.id: bool(row.is_fixed_asset) for row in item_settings}


    for line in return_in.lines:
        is_service = is_service_map.get(line.item_id, False)
        is_fixed_asset = is_fixed_asset_map.get(line.item_id, False)
        effective_warehouse_id = line.warehouse_id
        if effective_warehouse_id is None:
            if validation_default_warehouse is None:
                raise HTTPException(status_code=400, detail="No active warehouse found")
            effective_warehouse_id = int(validation_default_warehouse.id)

        warehouse = (
            db.query(models.Warehouse)
            .filter(
                models.Warehouse.id == effective_warehouse_id,
                models.Warehouse.company_id == company_id,
                models.Warehouse.is_active == True,
            )
            .first()
        )
        if not warehouse:
            raise HTTPException(status_code=400, detail="Invalid warehouse_id")

        db.add(
            models.SalesReturnLine(
                return_id=sales_return.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=line.rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
                warehouse_id=effective_warehouse_id,
                department_id=line.department_id or return_in.department_id,
                project_id=line.project_id or return_in.project_id,
            )
        )

        # Re-introduce inventory into stock ledger.
        posted_at = datetime.combine(sales_return.date, datetime.min.time())

        # Determine return cost: original sale unit cost when available.
        return_unit_cost = None
        if return_in.source_invoice_id is not None:
            src_cost = (
                db.query(func.avg(models.StockLedger.unit_cost))
                .filter(
                    models.StockLedger.company_id == company_id,
                    models.StockLedger.source_type == "SALES_INVOICE",
                    models.StockLedger.source_id == int(return_in.source_invoice_id),
                    models.StockLedger.item_id == int(line.item_id),
                    models.StockLedger.reversed_at.is_(None),
                    models.StockLedger.unit_cost.is_not(None),
                )
                .scalar()
            )
            if src_cost is not None:
                return_unit_cost = float(src_cost)

        if return_unit_cost is None:
            # Fallback: approximate using current issue unit cost for one unit.
            return_unit_cost = _compute_issue_unit_cost(
                db=db,
                company=company,
                company_id=company_id,
                item_id=int(line.item_id),
                warehouse_id=int(effective_warehouse_id),
                as_of=posted_at,
                qty_out=1.0,
            )

        if not is_service:
            db.add(
                models.StockLedger(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    qty_delta=float(line.quantity),
                    unit_cost=float(return_unit_cost),
                    source_type="SALES_RETURN",
                    source_id=sales_return.id,
                    source_line_id=None,
                    posted_at=posted_at,
                    created_by=current_user.id,
                )
            )

            if not is_fixed_asset and method == models.InventoryValuationMethod.FIFO:
                svc.fifo_add_batch(
                    tenant_id=int(company.tenant_id),
                    product_id=int(line.item_id),
                    qty_in=float(line.quantity),
                    rate=float(return_unit_cost),
                    ref_type="SALES_RETURN",
                    ref_id=int(sales_return.id),
                    created_at=posted_at,
                )
            db.add(
                models.StockMovement(
                    company_id=company_id,
                    warehouse_id=effective_warehouse_id,
                    item_id=line.item_id,
                    movement_date=sales_return.date,
                    source_type="SALES_RETURN",
                    source_id=sales_return.id,
                    qty_in=line.quantity,
                    qty_out=0,
                )
            )
    voucher = _build_sales_return_voucher(
        db,
        company_id,
        sales_return,
        return_in.payment_mode_id,
        sales_return_ledger_id=sales_return.sales_return_ledger_id,
        output_tax_return_ledger_id=sales_return.output_tax_return_ledger_id,
        payment_ledger_id=sales_return.payment_ledger_id,
    )
    sales_return.voucher_id = voucher.id

    db.commit()
    db.refresh(sales_return)
    return sales_return


@router.get("/returns", response_model=list[schemas.SalesReturnRead])
def list_sales_returns(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    returns = (
        db.query(models.SalesReturn)
        .filter(models.SalesReturn.company_id == company_id)
        .order_by(models.SalesReturn.date.desc(), models.SalesReturn.id.desc())
        .all()
    )
    return returns


@router.get("/returns/{return_id}", response_model=schemas.SalesReturnRead)
def get_sales_return(
    company_id: int,
    return_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    sales_return = (
        db.query(models.SalesReturn)
        .filter(
            models.SalesReturn.id == return_id,
            models.SalesReturn.company_id == company_id,
        )
        .first()
    )
    if not sales_return:
        raise HTTPException(status_code=404, detail="Sales return not found")
    return sales_return


@router.post("/invoices/{invoice_id}/create-return", response_model=schemas.SalesReturnRead)
def create_sales_return_from_invoice(
    company_id: int,
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    invoice = (
        db.query(models.SalesInvoice)
        .filter(
            models.SalesInvoice.id == invoice_id,
            models.SalesInvoice.company_id == company_id,
        )
        .first()
    )
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    sales_return = models.SalesReturn(
        company_id=company_id,
        customer_id=invoice.customer_id,
        date=invoice.date,
        reference=None,
        source_invoice_id=invoice.id,
    )
    db.add(sales_return)
    db.flush()

    validation_default_warehouse = _resolve_sales_default_warehouse(db, company_id=company_id)

    for line in invoice.lines:
        effective_warehouse_id = line.warehouse_id
        if effective_warehouse_id is None:
            if validation_default_warehouse is None:
                raise HTTPException(status_code=400, detail="No active warehouse found")
            effective_warehouse_id = int(validation_default_warehouse.id)

        warehouse = (
            db.query(models.Warehouse)
            .filter(
                models.Warehouse.id == effective_warehouse_id,
                models.Warehouse.company_id == company_id,
                models.Warehouse.is_active == True,
            )
            .first()
        )
        if not warehouse:
            raise HTTPException(status_code=400, detail="Invalid warehouse_id")

        db.add(
            models.SalesReturnLine(
                return_id=sales_return.id,
                item_id=line.item_id,
                quantity=line.quantity,
                rate=line.rate,
                discount=line.discount,
                tax_rate=line.tax_rate,
                hs_code=line.hs_code,
                warehouse_id=effective_warehouse_id,
            )
        )

        db.add(
            models.StockMovement(
                company_id=company_id,
                warehouse_id=effective_warehouse_id,
                item_id=line.item_id,
                movement_date=sales_return.date,
                source_type="SALES_RETURN",
                source_id=sales_return.id,
                qty_in=line.quantity,
                qty_out=0,
            )
        )

    _build_sales_return_voucher(db, company_id, sales_return, None)

    db.commit()
    db.refresh(sales_return)
    return sales_return


@router.get("/invoices/export-template")
def export_sales_invoice_template(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Invoices"
    
    headers = [
        "Customer Name", "Date (YYYY-MM-DD)", "Bill Date (YYYY-MM-DD)", "Due Date (YYYY-MM-DD)",
        "Invoice Reference", "Payment Mode", "Sales Ledger", "Output Tax Ledger", "Sales Person", "Sales Type (LOCAL/EXPORT)",
        "Item Name", "Quantity", "Rate", "Discount", "Tax Rate (%)", "HS Code",
        "Line Sales Person", "Remarks",
        "Warehouse Name", "Department Name", "Project Name", "Segment Name",
        "Narration"
    ]
    ws.append(headers)
    
    # Sample row
    ws.append([
        "Cash Customer", "2023-01-01", "", "2023-01-15",
        "INV-101", "Cash", "", "", "",
        "Item A", "5", "1000", "50", "13", "",
        "", "",
        "Main Warehouse", "Sales", "", "", "LOCAL",
        "Standard sales transaction"
    ])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_invoice_template.xlsx"}
    )


@router.post("/invoices/parse-excel")
async def parse_sales_invoices_excel(
    company_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import load_workbook
    from datetime import date, datetime
    
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Empty or invalid file")
    
    headers = [str(c).strip().lower() for c in rows[0] if c is not None]
    data_rows = rows[1:]
    
    # Pre-fetch lookups
    customer_objs = db.query(models.Customer).filter(models.Customer.company_id == company_id).all()
    customers = {c.name.lower(): c.id for c in customer_objs}
    customers_by_key = {c.name.lower(): {"id": c.id, "name": c.name} for c in customer_objs}
    
    item_objs = db.query(models.Item).filter(models.Item.company_id == company_id).all()
    items = {i.name.lower(): {"id": i.id, "name": i.name} for i in item_objs}
    
    warehouses = {w.name.lower(): {"id": w.id, "name": w.name} for w in db.query(models.Warehouse).filter(models.Warehouse.company_id == company_id).all()}
    payment_modes = {pm.name.lower(): pm.id for pm in db.query(models.PaymentMode).filter(models.PaymentMode.company_id == company_id).all()}
    departments = {d.name.lower(): {"id": d.id, "name": d.name} for d in db.query(models.Department).filter(models.Department.company_id == company_id).all()}
    projects = {p.name.lower(): {"id": p.id, "name": p.name} for p in db.query(models.Project).filter(models.Project.company_id == company_id).all()}
    segments = {s.name.lower(): {"id": s.id, "name": s.name} for s in db.query(models.Segment).filter(models.Segment.company_id == company_id).all()}
    sales_persons = {sp.name.lower(): sp.id for sp in db.query(models.SalesPerson).filter(models.SalesPerson.company_id == company_id).all()}

    parsed_invoices = []
    invoices_map = {}
    
    for idx, row_data in enumerate(data_rows, start=2):
        row = dict(zip(headers, row_data))
        
        customer_name = str(row.get("customer name") or "").strip()
        if not customer_name:
            continue
            
        inv_date_val = row.get("date (yyyy-mm-dd)")
        inv_date = date.today().isoformat()
        if inv_date_val:
            if isinstance(inv_date_val, (date, datetime)):
                inv_date = (inv_date_val.date() if isinstance(inv_date_val, datetime) else inv_date_val).isoformat()
            else:
                try:
                    inv_date = datetime.strptime(str(inv_date_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass

        bill_date_val = row.get("bill date (yyyy-mm-dd)")
        bill_date = None
        if bill_date_val:
            if isinstance(bill_date_val, (date, datetime)):
                bill_date = (bill_date_val.date() if isinstance(bill_date_val, datetime) else bill_date_val).isoformat()
            else:
                try:
                    bill_date = datetime.strptime(str(bill_date_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass

        due_date_val = row.get("due date (yyyy-mm-dd)")
        due_date = None
        if due_date_val:
            if isinstance(due_date_val, (date, datetime)):
                due_date = (due_date_val.date() if isinstance(due_date_val, datetime) else due_date_val).isoformat()
            else:
                try:
                    due_date = datetime.strptime(str(due_date_val).strip(), "%Y-%m-%d").date().isoformat()
                except:
                    pass
        
        reference = str(row.get("invoice reference") or "").strip()
        narration = str(row.get("narration") or "").strip()
        pm_name = str(row.get("payment mode") or "").strip()
        sales_ledger_name = str(row.get("sales ledger") or "").strip()
        tax_ledger_name = str(row.get("output tax ledger") or "").strip()
        sales_person_name = str(row.get("sales person") or "").strip()
        
        inv_key = (customer_name, inv_date, reference, pm_name)
        if inv_key not in invoices_map:
            customer_id = customers.get(customer_name.lower())
            customer_suggestions = []
            if not customer_id:
                close = difflib.get_close_matches(customer_name.lower(), list(customers.keys()), n=3, cutoff=0.5)
                customer_suggestions = [
                    {"name": customers_by_key[k]["name"], "id": customers_by_key[k]["id"]}
                    for k in close
                ]

            # Resolve ledgers
            sales_ledger_id = None
            if sales_ledger_name:
                led = db.query(models.Ledger).filter(models.Ledger.company_id == company_id, models.Ledger.name.ilike(sales_ledger_name)).first()
                sales_ledger_id = led.id if led else None
            
            tax_ledger_id = None
            if tax_ledger_name:
                led = db.query(models.Ledger).filter(models.Ledger.company_id == company_id, models.Ledger.name.ilike(tax_ledger_name)).first()
                tax_ledger_id = led.id if led else None

            inv_obj = {
                "id": len(parsed_invoices),
                "customer_name": customer_name,
                "customer_id": customer_id,
                "customer_suggestions": customer_suggestions,
                "date": inv_date,
                "bill_date": bill_date,
                "due_date": due_date,
                "reference": reference,
                "narration": narration,
                "payment_mode_name": pm_name,
                "payment_mode_id": payment_modes.get(pm_name.lower()),
                "sales_ledger_name": sales_ledger_name,
                "sales_ledger_id": sales_ledger_id,
                "output_tax_ledger_name": tax_ledger_name,
                "output_tax_ledger_id": tax_ledger_id,
                "sales_type": str(row.get("sales type (local/export)") or "LOCAL").strip().upper(),
                "sales_person_id": sales_persons.get(sales_person_name.lower()),
                "sales_person_name": sales_person_name,
                "lines": [],
                "errors": [],
                "warnings": []
            }
            if not customer_id:
                if customer_suggestions:
                    inv_obj["errors"].append(f"Customer '{customer_name}' not found — did you mean one of the suggestions below?")
                else:
                    inv_obj["errors"].append(f"Customer '{customer_name}' not found")
            
            if sales_ledger_name and not sales_ledger_id:
                inv_obj["warnings"].append(f"Sales Ledger '{sales_ledger_name}' not found — will use default if empty.")
            if tax_ledger_name and not tax_ledger_id:
                inv_obj["warnings"].append(f"Tax Ledger '{tax_ledger_name}' not found — will use default if empty.")
            if sales_person_name and not inv_obj["sales_person_id"]:
                inv_obj["warnings"].append(f"Sales Person '{sales_person_name}' not found.")
            
            invoices_map[inv_key] = inv_obj
            parsed_invoices.append(inv_obj)
        
        item_name = str(row.get("item name") or "").strip()
        item_info = items.get(item_name.lower())
        
        item_suggestions = []
        if not item_info:
            close = difflib.get_close_matches(item_name.lower(), list(items.keys()), n=3, cutoff=0.5)
            item_suggestions = [
                {"name": items[k]["name"], "id": items[k]["id"]}
                for k in close
            ]

        wh_raw = str(row.get("warehouse name") or "").strip()
        dept_raw = str(row.get("department name") or "").strip()
        proj_raw = str(row.get("project name") or "").strip()
        seg_raw = str(row.get("segment name") or "").strip()
        
        wh_info = warehouses.get(wh_raw.lower())
        dept_info = departments.get(dept_raw.lower())
        proj_info = projects.get(proj_raw.lower())
        seg_info = segments.get(seg_raw.lower())
        
        line_sales_person_name = str(row.get("line sales person") or "").strip()

        line = {
            "item_name": item_name,
            "item_id": item_info["id"] if item_info else None,
            "item_suggestions": item_suggestions,
            "quantity": float(row.get("quantity") or 0),
            "rate": float(row.get("rate") or 0),
            "discount": float(row.get("discount") or 0),
            "tax_rate": float(row.get("tax rate (%)") or 0),
            "hs_code": str(row.get("hs code") or "").strip() or None,
            "remarks": str(row.get("remarks") or "").strip() or None,
            "sales_person_id": sales_persons.get(line_sales_person_name.lower()),
            "sales_person_name": line_sales_person_name,
            "warehouse_name": wh_info["name"] if wh_info else wh_raw,
            "warehouse_id": wh_info["id"] if wh_info else None,
            "department_name": dept_info["name"] if dept_info else dept_raw,
            "department_id": dept_info["id"] if dept_info else None,
            "project_name": proj_info["name"] if proj_info else proj_raw,
            "project_id": proj_info["id"] if proj_info else None,
            "segment_name": seg_info["name"] if seg_info else seg_raw,
            "segment_id": seg_info["id"] if seg_info else None,
        }
        
        # Validation Logic
        errors = invoices_map[inv_key]["errors"]
        if not line["item_id"]:
            if item_suggestions:
                errors.append(f"Item '{item_name}' not found — did you mean one of the suggestions?")
            else:
                errors.append(f"Item '{item_name}' not found")
        
        if line["quantity"] <= 0:
            errors.append(f"Invalid quantity ({line['quantity']}) for item '{item_name}'")
        if line["rate"] <= 0:
            errors.append(f"Invalid rate ({line['rate']}) for item '{item_name}'")
        if line["discount"] < 0:
            errors.append(f"Negative discount ({line['discount']}) for item '{item_name}'")
        if line["tax_rate"] < 0 or line["tax_rate"] > 100:
            errors.append(f"Suspicious tax rate ({line['tax_rate']}%) for item '{item_name}'")

        # Dimension Warnings
        warnings = invoices_map[inv_key]["warnings"]
        if wh_raw and not wh_info:
            warnings.append(f"Warehouse '{wh_raw}' not found")
        if dept_raw and not dept_info:
            warnings.append(f"Department '{dept_raw}' not found")
        if proj_raw and not proj_info:
            warnings.append(f"Project '{proj_raw}' not found")
        if seg_raw and not seg_info:
            warnings.append(f"Segment '{seg_raw}' not found")
        
        invoices_map[inv_key]["lines"].append(line)

    return parsed_invoices


@router.post("/invoices/confirm-import")
def confirm_sales_invoices_import(
    company_id: int,
    invoices_in: list[dict],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    
    created_count = 0
    
    for idata in invoices_in:
        try:
            if not idata.get("customer_id"):
                continue
            
            inv_date = datetime.fromisoformat(idata["date"]).date()
            validate_transaction_date(company, inv_date)
            
            invoice = models.SalesInvoice(
                company_id=company_id,
                customer_id=idata["customer_id"],
                date=inv_date,
                due_date=(
                    datetime.fromisoformat(idata["due_date"]).date() 
                    if idata.get("due_date") 
                    else inv_date
                ),
                bill_date=(
                    datetime.fromisoformat(idata["bill_date"]).date() 
                    if idata.get("bill_date") 
                    else None
                ),
                reference=idata.get("reference"),
                
                payment_ledger_id=idata.get("payment_ledger_id"),
                sales_ledger_id=idata.get("sales_ledger_id"),
                output_tax_ledger_id=idata.get("output_tax_ledger_id"),
                sales_person_id=idata.get("sales_person_id"),
                narration=idata.get("narration"),
                department_id=idata.get("department_id"),
                project_id=idata.get("project_id"),
                segment_id=idata.get("segment_id"),
                sales_type=idata.get("sales_type", "LOCAL"),
                apply_tds=idata.get("apply_tds") or False,
                tds_amount=idata.get("tds_amount"),
                tds_ledger_id=idata.get("tds_ledger_id"),
                
            )
            db.add(invoice)
            db.flush()
            
            # Helper maps for stock postings
            is_service_map = {}
            is_fixed_asset_map = {}
            sell_as_kit_map = {}
            
            for ldata in idata["lines"]:
                item = db.query(models.Item).filter(models.Item.id == ldata["item_id"]).first()
                if item:
                    is_service_map[item.id] = (item.category and item.category.strip().lower() == "service")
                    is_fixed_asset_map[item.id] = bool(getattr(item, "is_fixed_asset", False))
                    sell_as_kit_map[item.id] = bool(getattr(item, "sell_as_kit", False))
                
                display_only = {
                    "item_name", "item_suggestions",
                    "warehouse_name", "department_name", "project_name", "segment_name",
                    "customer_suggestions", "customer_name",
                    "sales_ledger_name", "output_tax_ledger_name", "sales_person_name",
                    "payment_mode_name"
                }
                line_data = {k: v for k, v in ldata.items() if k not in display_only}
                line = models.SalesInvoiceLine(
                    invoice_id=invoice.id,
                    **line_data
                )
                db.add(line)
            
            db.flush()
            
            # Post Stock
            db.refresh(invoice, ["lines"])
            inv_lines = invoice.lines
            _apply_stock_postings_for_sales_invoice(
                db, 
                company_id=company_id, 
                company=company, 
                invoice=invoice, 
                current_user_id=current_user.id,
                is_service_map=is_service_map,
                is_fixed_asset_map=is_fixed_asset_map,
                sell_as_kit_map=sell_as_kit_map,
                lines=inv_lines
            )
            
            # Build Voucher
            _build_sales_voucher(
                db, 
                company_id, 
                invoice, 
                idata.get("payment_mode_id"),
                sales_ledger_id=invoice.sales_ledger_id,
                output_tax_ledger_id=invoice.output_tax_ledger_id,
                payment_ledger_id=invoice.payment_ledger_id,
            )
            created_count += 1
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=400, detail=str(e))
            
    db.commit()
    return {"detail": f"Successfully imported {created_count} sales invoices"}
