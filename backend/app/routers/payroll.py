from __future__ import annotations

from datetime import date, datetime, timedelta
from datetime import time as dtime

import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas
from ..auth import get_current_user
from ..database import get_db
from .admin_logs import log_event
from ..payroll_service import (
    ingest_attendance_logs as ingest_attendance_logs_service,
    upsert_attendance_from_raw_logs,
    compute_payroll_run,
    build_payroll_voucher_payload,
    preview_formula_amount,
    apply_designation_template,
)


from ..dependencies import get_company_secure

router = APIRouter(prefix="/payroll/companies/{company_id}", tags=["payroll"])


def _parse_hhmm(value: str) -> dtime:
    s = str(value or "").strip()
    try:
        hh, mm = s.split(":", 1)
        return dtime(int(hh), int(mm), 0)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid time format; expected HH:MM")


def _get_company(db: Session, company_id: int, user: models.User) -> models.Company:
    return get_company_secure(db, company_id, user)


def _require_admin_or_superadmin(user: models.User) -> None:
    if user.role not in (models.UserRole.admin, models.UserRole.superadmin):
        raise HTTPException(status_code=403, detail="Admin privileges required")




def _build_payslip_read(
    db: Session,
    *,
    company_id: int,
    slip: models.PayrollPayslip,
) -> schemas.PayslipRead:
    lines = (
        db.query(models.PayrollPayslipLine)
        .filter(models.PayrollPayslipLine.company_id == company_id, models.PayrollPayslipLine.payslip_id == int(slip.id))
        .order_by(models.PayrollPayslipLine.id.asc())
        .all()
    )
    return schemas.PayslipRead(
        id=int(slip.id),
        company_id=int(slip.company_id),
        payroll_run_id=int(slip.payroll_run_id),
        employee_id=int(slip.employee_id),
        payable_days=float(slip.payable_days or 0),
        absent_days=float(slip.absent_days or 0),
        late_minutes=int(slip.late_minutes or 0),
        overtime_minutes=int(slip.overtime_minutes or 0),
        earnings_total=float(slip.earnings_total or 0),
        deductions_total=float(slip.deductions_total or 0),
        tds_amount=float(getattr(slip, "tds_amount", 0) or 0),
        net_pay=float(slip.net_pay or 0),
        is_manual_override=bool(slip.is_manual_override),
        override_reason=getattr(slip, "override_reason", None),
        lines=[
            schemas.PayslipLineRead(
                payhead_id=int(l.payhead_id),
                type=l.type,
                amount=float(l.amount or 0),
            )
            for l in lines
        ],
    )


@router.get("/employees", response_model=list[schemas.EmployeeRead])
def list_employees(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id)
        .order_by(models.Employee.id.asc())
        .all()
    )


@router.get("/employee-types", response_model=list[schemas.EmployeeTypeRead])
def list_employee_types(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.EmployeeType)
        .filter(models.EmployeeType.company_id == company_id)
        .order_by(models.EmployeeType.id.asc())
        .all()
    )


@router.post("/employee-types", response_model=schemas.EmployeeTypeRead)
def create_employee_type(
    company_id: int,
    payload: schemas.EmployeeTypeCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    data = payload.model_dump()
    if data.get("code") == "":
        data["code"] = None
    row = models.EmployeeType(company_id=company_id, **data)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/employee-types/{type_id}", response_model=schemas.EmployeeTypeRead)
def update_employee_type(
    company_id: int,
    type_id: int,
    payload: schemas.EmployeeTypeUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.EmployeeType)
        .filter(models.EmployeeType.company_id == company_id, models.EmployeeType.id == type_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Employee Type not found")
    
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] == "":
        data["code"] = None
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/employee-types/{type_id}")
def delete_employee_type(
    company_id: int,
    type_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.EmployeeType)
        .filter(models.EmployeeType.company_id == company_id, models.EmployeeType.id == type_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Employee Type not found")
    
    # Check if used by any employee
    usage = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.employee_type_id == type_id)
        .first()
    )
    if usage:
        raise HTTPException(status_code=400, detail="Cannot delete Employee Type as it is associated with one or more employees.")

    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.get("/designations", response_model=list[schemas.DesignationRead])
def list_designations(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id)
        .order_by(models.PayrollDesignation.id.asc())
        .all()
    )


@router.post("/designations", response_model=schemas.DesignationRead)
def create_designation(
    company_id: int,
    payload: schemas.DesignationCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    data = payload.model_dump()
    if data.get("code") == "":
        data["code"] = None
    row = models.PayrollDesignation(company_id=company_id, **data)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/designations/{designation_id}", response_model=schemas.DesignationRead)
def update_designation(
    company_id: int,
    designation_id: int,
    payload: schemas.DesignationUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == designation_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Designation not found")
    
    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] == "":
        data["code"] = None
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/designations/{designation_id}")
def delete_designation(
    company_id: int,
    designation_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == designation_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Designation not found")
    
    # Check if used by any employee
    usage = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.designation_id == designation_id)
        .first()
    )
    if usage:
        raise HTTPException(status_code=400, detail="Cannot delete Designation as it is associated with one or more employees.")

    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}



@router.get("/designations/{designation_id}/template", response_model=list[schemas.DesignationTemplateLineRead])
def list_designation_template_lines(
    company_id: int,
    designation_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    designation = (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == designation_id)
        .first()
    )
    if not designation:
        raise HTTPException(status_code=404, detail="Designation not found")
    return (
        db.query(models.DesignationTemplateLine)
        .filter(
            models.DesignationTemplateLine.company_id == company_id,
            models.DesignationTemplateLine.designation_id == designation_id,
        )
        .order_by(models.DesignationTemplateLine.sort_order.asc(), models.DesignationTemplateLine.id.asc())
        .all()
    )


@router.post("/designations/{designation_id}/template", response_model=schemas.DesignationTemplateLineRead)
def add_designation_template_line(
    company_id: int,
    designation_id: int,
    payload: schemas.DesignationTemplateLineCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    designation = (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == designation_id)
        .first()
    )
    if not designation:
        raise HTTPException(status_code=404, detail="Designation not found")
    ph = (
        db.query(models.PayrollPayhead)
        .filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.id == int(payload.payhead_id))
        .first()
    )
    if ph is None:
        raise HTTPException(status_code=400, detail="Invalid payhead_id")

    existing = (
        db.query(models.DesignationTemplateLine)
        .filter(
            models.DesignationTemplateLine.designation_id == designation_id,
            models.DesignationTemplateLine.payhead_id == int(payload.payhead_id),
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="A template line for this payhead already exists")

    row = models.DesignationTemplateLine(
        company_id=company_id,
        designation_id=designation_id,
        payhead_id=int(payload.payhead_id),
        amount=payload.amount,
        rate=payload.rate,
        formula=payload.formula,
        sort_order=payload.sort_order,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/designations/{designation_id}/template/{line_id}", response_model=schemas.DesignationTemplateLineRead)
def update_designation_template_line(
    company_id: int,
    designation_id: int,
    line_id: int,
    payload: schemas.DesignationTemplateLineUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.DesignationTemplateLine)
        .filter(
            models.DesignationTemplateLine.company_id == company_id,
            models.DesignationTemplateLine.designation_id == designation_id,
            models.DesignationTemplateLine.id == line_id,
        )
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Template line not found")
    data = payload.model_dump(exclude_unset=True)
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/designations/{designation_id}/template/{line_id}")
def delete_designation_template_line(
    company_id: int,
    designation_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.DesignationTemplateLine)
        .filter(
            models.DesignationTemplateLine.company_id == company_id,
            models.DesignationTemplateLine.designation_id == designation_id,
            models.DesignationTemplateLine.id == line_id,
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Template line not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.post("/designations/{designation_id}/apply-template")
def apply_designation_template_to_employees(
    company_id: int,
    designation_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Apply the designation template to all active employees with this designation (replaces their active pay structure)."""
    _get_company(db, company_id, current_user)
    designation = (
        db.query(models.PayrollDesignation)
        .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == designation_id)
        .first()
    )
    if not designation:
        raise HTTPException(status_code=404, detail="Designation not found")

    employees = (
        db.query(models.Employee)
        .filter(
            models.Employee.company_id == company_id,
            models.Employee.designation_id == designation_id,
            models.Employee.is_active == True,
        )
        .all()
    )

    applied_to: list[int] = []
    for emp in employees:
        apply_designation_template(
            db,
            company_id=company_id,
            employee_id=int(emp.id),
            designation=designation,
            effective_from=date.today(),
        )
        applied_to.append(int(emp.id))

    db.commit()
    return {"detail": f"Template applied to {len(applied_to)} employee(s)", "employee_ids": applied_to}


@router.get("/employees/{employee_id}/extra-payheads", response_model=list[schemas.EmployeeExtraPayheadRead])
def list_employee_extra_payheads(
    company_id: int,
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """List all extra pay heads for a specific employee (additive on top of designation template)."""
    _get_company(db, company_id, current_user)
    return (
        db.query(models.EmployeeExtraPayhead)
        .filter(
            models.EmployeeExtraPayhead.company_id == company_id,
            models.EmployeeExtraPayhead.employee_id == employee_id,
        )
        .order_by(models.EmployeeExtraPayhead.sort_order.asc(), models.EmployeeExtraPayhead.id.asc())
        .all()
    )


@router.post("/employees/{employee_id}/extra-payheads", response_model=schemas.EmployeeExtraPayheadRead)
def add_employee_extra_payhead(
    company_id: int,
    employee_id: int,
    payload: schemas.EmployeeExtraPayheadCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Add an extra pay head to a specific employee (on top of the designation template)."""
    _get_company(db, company_id, current_user)
    emp = db.query(models.Employee).filter(models.Employee.company_id == company_id, models.Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    existing = db.query(models.EmployeeExtraPayhead).filter(
        models.EmployeeExtraPayhead.employee_id == employee_id,
        models.EmployeeExtraPayhead.payhead_id == payload.payhead_id,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Extra payhead already exists for this employee; use PUT to update")
    line = models.EmployeeExtraPayhead(
        company_id=company_id,
        employee_id=employee_id,
        **payload.model_dump(),
    )
    db.add(line)
    db.commit()
    db.refresh(line)
    return line


@router.put("/employees/{employee_id}/extra-payheads/{line_id}", response_model=schemas.EmployeeExtraPayheadRead)
def update_employee_extra_payhead(
    company_id: int,
    employee_id: int,
    line_id: int,
    payload: schemas.EmployeeExtraPayheadUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    line = db.query(models.EmployeeExtraPayhead).filter(
        models.EmployeeExtraPayhead.company_id == company_id,
        models.EmployeeExtraPayhead.employee_id == employee_id,
        models.EmployeeExtraPayhead.id == line_id,
    ).first()
    if not line:
        raise HTTPException(status_code=404, detail="Extra payhead not found")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(line, k, v)
    db.commit()
    db.refresh(line)
    return line


@router.delete("/employees/{employee_id}/extra-payheads/{line_id}", status_code=204)
def delete_employee_extra_payhead(
    company_id: int,
    employee_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    line = db.query(models.EmployeeExtraPayhead).filter(
        models.EmployeeExtraPayhead.company_id == company_id,
        models.EmployeeExtraPayhead.employee_id == employee_id,
        models.EmployeeExtraPayhead.id == line_id,
    ).first()
    if not line:
        raise HTTPException(status_code=404, detail="Extra payhead not found")
    db.delete(line)
    db.commit()


@router.post("/employees", response_model=schemas.EmployeeRead)
def create_employee(
    company_id: int,
    payload: schemas.EmployeeCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    data = payload.model_dump()
    if data.get("code"):
        existing_code = (
            db.query(models.Employee)
            .filter(models.Employee.company_id == company_id, models.Employee.code == data["code"])
            .first()
        )
        if existing_code:
            raise HTTPException(status_code=400, detail="Duplicate: An employee with this ID/Code already exists.")
    # Validate employee_type_id if provided
    if data.get("employee_type_id"):
        etype = (
            db.query(models.EmployeeType)
            .filter(models.EmployeeType.company_id == company_id, models.EmployeeType.id == data["employee_type_id"])
            .first()
        )
        if not etype:
            raise HTTPException(status_code=400, detail="Invalid employee_type_id")

    designation = None
    if data.get("designation_id"):
        designation = (
            db.query(models.PayrollDesignation)
            .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == data["designation_id"])
            .first()
        )
        if designation is None:
            raise HTTPException(status_code=400, detail="Invalid designation_id")
        if designation and data.get("base_monthly_salary") is None and designation.base_monthly_salary is not None:
            data["base_monthly_salary"] = designation.base_monthly_salary

    emp = models.Employee(company_id=company_id, **data)
    db.add(emp)
    db.flush()

    if designation and designation.template_lines:
        apply_designation_template(
            db,
            company_id=company_id,
            employee_id=int(emp.id),
            designation=designation,
            effective_from=data.get("join_date") or date.today(),
        )

    db.commit()
    db.refresh(emp)
    return emp


@router.get("/shifts", response_model=list[schemas.PayrollShiftRead])
def list_shifts(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.PayrollShift)
        .filter(models.PayrollShift.company_id == company_id)
        .order_by(models.PayrollShift.id.asc())
        .all()
    )


@router.post("/shifts", response_model=schemas.PayrollShiftRead)
def create_shift(
    company_id: int,
    payload: schemas.PayrollShiftCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = models.PayrollShift(
        company_id=company_id,
        code=str(payload.code).strip().upper(),
        name=str(payload.name).strip(),
        start_time=_parse_hhmm(payload.start_time),
        end_time=_parse_hhmm(payload.end_time),
        expected_work_minutes=int(payload.expected_work_minutes or 0),
        grace_minutes=int(payload.grace_minutes or 0),
        allow_night_shift=bool(payload.allow_night_shift),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/pay-structures", response_model=list[schemas.EmployeePayStructureRead])
def list_pay_structures(
    company_id: int,
    employee_id: int | None = Query(None),
    is_active: bool | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    q = db.query(models.EmployeePayStructure).filter(models.EmployeePayStructure.company_id == company_id)
    if employee_id is not None:
        q = q.filter(models.EmployeePayStructure.employee_id == int(employee_id))
    if is_active is not None:
        q = q.filter(models.EmployeePayStructure.is_active == bool(is_active))
    rows = q.order_by(models.EmployeePayStructure.effective_from.desc(), models.EmployeePayStructure.id.desc()).all()

    result: list[schemas.EmployeePayStructureRead] = []
    for s in rows:
        lines = (
            db.query(models.EmployeePayStructureLine)
            .filter(models.EmployeePayStructureLine.company_id == company_id, models.EmployeePayStructureLine.structure_id == int(s.id))
            .order_by(models.EmployeePayStructureLine.id.asc())
            .all()
        )
        result.append(
            schemas.EmployeePayStructureRead(
                id=int(s.id),
                company_id=int(s.company_id),
                employee_id=int(s.employee_id),
                effective_from=s.effective_from,
                effective_to=s.effective_to,
                is_active=bool(s.is_active),
                created_at=s.created_at,
                lines=[
                    schemas.EmployeePayStructureLineRead(
                        id=int(l.id),
                        company_id=int(l.company_id),
                        structure_id=int(l.structure_id),
                        payhead_id=int(l.payhead_id),
                        amount=float(l.amount) if l.amount is not None else None,
                        rate=float(l.rate) if l.rate is not None else None,
                        formula=l.formula,
                        created_at=l.created_at,
                    )
                    for l in lines
                ],
            )
        )
    return result


@router.post("/pay-structures", response_model=schemas.EmployeePayStructureRead)
def create_pay_structure(
    company_id: int,
    payload: schemas.EmployeePayStructureCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    emp = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.id == int(payload.employee_id))
        .first()
    )
    if emp is None:
        raise HTTPException(status_code=400, detail="Invalid employee_id")

    if payload.effective_to is not None and payload.effective_to < payload.effective_from:
        raise HTTPException(status_code=400, detail="effective_to must be >= effective_from")

    structure = models.EmployeePayStructure(
        company_id=company_id,
        employee_id=int(payload.employee_id),
        effective_from=payload.effective_from,
        effective_to=payload.effective_to,
        is_active=bool(payload.is_active),
    )
    db.add(structure)
    db.flush()

    created_lines: list[models.EmployeePayStructureLine] = []
    for line in payload.lines or []:
        ph = (
            db.query(models.PayrollPayhead)
            .filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.id == int(line.payhead_id))
            .first()
        )
        if ph is None:
            raise HTTPException(status_code=400, detail="Invalid payhead_id")
        created_lines.append(
            models.EmployeePayStructureLine(
                company_id=company_id,
                structure_id=int(structure.id),
                payhead_id=int(line.payhead_id),
                amount=line.amount,
                rate=line.rate,
                formula=line.formula,
            )
        )

    for l in created_lines:
        db.add(l)

    db.commit()
    db.refresh(structure)

    lines = (
        db.query(models.EmployeePayStructureLine)
        .filter(models.EmployeePayStructureLine.company_id == company_id, models.EmployeePayStructureLine.structure_id == int(structure.id))
        .order_by(models.EmployeePayStructureLine.id.asc())
        .all()
    )

    return schemas.EmployeePayStructureRead(
        id=int(structure.id),
        company_id=int(structure.company_id),
        employee_id=int(structure.employee_id),
        effective_from=structure.effective_from,
        effective_to=structure.effective_to,
        is_active=bool(structure.is_active),
        created_at=structure.created_at,
        lines=[
            schemas.EmployeePayStructureLineRead(
                id=int(l.id),
                company_id=int(l.company_id),
                structure_id=int(l.structure_id),
                payhead_id=int(l.payhead_id),
                amount=float(l.amount) if l.amount is not None else None,
                rate=float(l.rate) if l.rate is not None else None,
                formula=l.formula,
                created_at=l.created_at,
            )
            for l in lines
        ],
    )


@router.get("/pay-structures/{structure_id}", response_model=schemas.EmployeePayStructureRead)
def get_pay_structure(
    company_id: int,
    structure_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    s = (
        db.query(models.EmployeePayStructure)
        .filter(models.EmployeePayStructure.company_id == company_id, models.EmployeePayStructure.id == int(structure_id))
        .first()
    )
    if s is None:
        raise HTTPException(status_code=404, detail="Pay structure not found")

    lines = (
        db.query(models.EmployeePayStructureLine)
        .filter(models.EmployeePayStructureLine.company_id == company_id, models.EmployeePayStructureLine.structure_id == int(s.id))
        .order_by(models.EmployeePayStructureLine.id.asc())
        .all()
    )

    return schemas.EmployeePayStructureRead(
        id=int(s.id),
        company_id=int(s.company_id),
        employee_id=int(s.employee_id),
        effective_from=s.effective_from,
        effective_to=s.effective_to,
        is_active=bool(s.is_active),
        created_at=s.created_at,
        lines=[
            schemas.EmployeePayStructureLineRead(
                id=int(l.id),
                company_id=int(l.company_id),
                structure_id=int(l.structure_id),
                payhead_id=int(l.payhead_id),
                amount=float(l.amount) if l.amount is not None else None,
                rate=float(l.rate) if l.rate is not None else None,
                formula=l.formula,
                created_at=l.created_at,
            )
            for l in lines
        ],
    )


@router.put("/pay-structures/{structure_id}", response_model=schemas.EmployeePayStructureRead)
def update_pay_structure(
    company_id: int,
    structure_id: int,
    payload: schemas.EmployeePayStructureUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    s = (
        db.query(models.EmployeePayStructure)
        .filter(models.EmployeePayStructure.company_id == company_id, models.EmployeePayStructure.id == int(structure_id))
        .with_for_update()
        .first()
    )
    if s is None:
        raise HTTPException(status_code=404, detail="Pay structure not found")

    data = payload.model_dump(exclude_unset=True)
    if "effective_from" in data and data["effective_from"] is not None:
        s.effective_from = data["effective_from"]
    if "effective_to" in data:
        s.effective_to = data.get("effective_to")
    if s.effective_to is not None and s.effective_to < s.effective_from:
        raise HTTPException(status_code=400, detail="effective_to must be >= effective_from")
    if "is_active" in data and data["is_active"] is not None:
        s.is_active = bool(data["is_active"])
    db.add(s)
    db.commit()
    db.refresh(s)
    return get_pay_structure(company_id=company_id, structure_id=int(s.id), db=db, current_user=current_user)


@router.delete("/pay-structures/{structure_id}")
def delete_pay_structure(
    company_id: int,
    structure_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    s = (
        db.query(models.EmployeePayStructure)
        .filter(models.EmployeePayStructure.company_id == company_id, models.EmployeePayStructure.id == int(structure_id))
        .first()
    )
    if s is None:
        raise HTTPException(status_code=404, detail="Pay structure not found")
    db.delete(s)
    db.commit()
    return {"detail": "Deleted"}


@router.post("/pay-structures/{structure_id}/lines", response_model=schemas.EmployeePayStructureRead)
def add_pay_structure_line(
    company_id: int,
    structure_id: int,
    payload: schemas.EmployeePayStructureLineCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    s = (
        db.query(models.EmployeePayStructure)
        .filter(models.EmployeePayStructure.company_id == company_id, models.EmployeePayStructure.id == int(structure_id))
        .first()
    )
    if s is None:
        raise HTTPException(status_code=404, detail="Pay structure not found")
    ph = (
        db.query(models.PayrollPayhead)
        .filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.id == int(payload.payhead_id))
        .first()
    )
    if ph is None:
        raise HTTPException(status_code=400, detail="Invalid payhead_id")

    row = models.EmployeePayStructureLine(
        company_id=company_id,
        structure_id=int(s.id),
        payhead_id=int(payload.payhead_id),
        amount=payload.amount,
        rate=payload.rate,
        formula=payload.formula,
    )
    db.add(row)
    db.commit()
    return get_pay_structure(company_id=company_id, structure_id=int(s.id), db=db, current_user=current_user)


@router.put("/pay-structures/{structure_id}/lines/{line_id}", response_model=schemas.EmployeePayStructureRead)
def update_pay_structure_line(
    company_id: int,
    structure_id: int,
    line_id: int,
    payload: schemas.EmployeePayStructureLineUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.EmployeePayStructureLine)
        .filter(
            models.EmployeePayStructureLine.company_id == company_id,
            models.EmployeePayStructureLine.structure_id == int(structure_id),
            models.EmployeePayStructureLine.id == int(line_id),
        )
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Structure line not found")
    data = payload.model_dump(exclude_unset=True)
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    return get_pay_structure(company_id=company_id, structure_id=int(structure_id), db=db, current_user=current_user)


@router.delete("/pay-structures/{structure_id}/lines/{line_id}")
def delete_pay_structure_line(
    company_id: int,
    structure_id: int,
    line_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.EmployeePayStructureLine)
        .filter(
            models.EmployeePayStructureLine.company_id == company_id,
            models.EmployeePayStructureLine.structure_id == int(structure_id),
            models.EmployeePayStructureLine.id == int(line_id),
        )
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Structure line not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.post("/formula/preview", response_model=schemas.PayrollFormulaPreviewResponse)
def preview_formula(
    company_id: int,
    payload: schemas.PayrollFormulaPreviewRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    amount, vars_used = preview_formula_amount(
        db,
        company_id=company_id,
        formula=payload.formula,
        employee_id=payload.employee_id,
        structure_id=payload.structure_id,
        payable_days=payload.payable_days,
        absent_days=payload.absent_days,
        late_minutes=payload.late_minutes,
        overtime_minutes=payload.overtime_minutes,
        worked_minutes=payload.worked_minutes,
        variables=payload.variables,
    )
    return schemas.PayrollFormulaPreviewResponse(amount=float(amount), variables=vars_used)


@router.put("/shifts/{shift_id}", response_model=schemas.PayrollShiftRead)
def update_shift(
    company_id: int,
    shift_id: int,
    payload: schemas.PayrollShiftUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.PayrollShift)
        .filter(models.PayrollShift.company_id == company_id, models.PayrollShift.id == int(shift_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Shift not found")

    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"] is not None:
        row.code = str(data["code"]).strip().upper()
    if "name" in data and data["name"] is not None:
        row.name = str(data["name"]).strip()
    if "start_time" in data and data["start_time"] is not None:
        row.start_time = _parse_hhmm(str(data["start_time"]))
    if "end_time" in data and data["end_time"] is not None:
        row.end_time = _parse_hhmm(str(data["end_time"]))
    if "expected_work_minutes" in data and data["expected_work_minutes"] is not None:
        row.expected_work_minutes = int(data["expected_work_minutes"])
    if "grace_minutes" in data and data["grace_minutes"] is not None:
        row.grace_minutes = int(data["grace_minutes"])
    if "allow_night_shift" in data and data["allow_night_shift"] is not None:
        row.allow_night_shift = bool(data["allow_night_shift"])

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/shifts/{shift_id}")
def delete_shift(
    company_id: int,
    shift_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.PayrollShift)
        .filter(models.PayrollShift.company_id == company_id, models.PayrollShift.id == int(shift_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Shift not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.get("/shift-assignments", response_model=list[schemas.EmployeeShiftAssignmentRead])
def list_shift_assignments(
    company_id: int,
    employee_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    q = db.query(models.EmployeeShiftAssignment).filter(models.EmployeeShiftAssignment.company_id == company_id)
    if employee_id is not None:
        q = q.filter(models.EmployeeShiftAssignment.employee_id == int(employee_id))
    return q.order_by(models.EmployeeShiftAssignment.id.asc()).all()


@router.post("/shift-assignments", response_model=schemas.EmployeeShiftAssignmentRead)
def create_shift_assignment(
    company_id: int,
    payload: schemas.EmployeeShiftAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    emp = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.id == int(payload.employee_id))
        .first()
    )
    if emp is None:
        raise HTTPException(status_code=400, detail="Invalid employee_id")
    shift = (
        db.query(models.PayrollShift)
        .filter(models.PayrollShift.company_id == company_id, models.PayrollShift.id == int(payload.shift_id))
        .first()
    )
    if shift is None:
        raise HTTPException(status_code=400, detail="Invalid shift_id")

    row = models.EmployeeShiftAssignment(company_id=company_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/shift-assignments/{assignment_id}")
def delete_shift_assignment(
    company_id: int,
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.EmployeeShiftAssignment)
        .filter(models.EmployeeShiftAssignment.company_id == company_id, models.EmployeeShiftAssignment.id == int(assignment_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Shift assignment not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.put("/employees/{employee_id}", response_model=schemas.EmployeeRead)
def update_employee(
    company_id: int,
    employee_id: int,
    payload: schemas.EmployeeUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    emp = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.id == employee_id)
        .first()
    )
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    data = payload.model_dump(exclude_unset=True)
    if "code" in data and data["code"]:
        existing_code = (
            db.query(models.Employee)
            .filter(
                models.Employee.company_id == company_id,
                models.Employee.code == data["code"],
                models.Employee.id != employee_id,
            )
            .first()
        )
        if existing_code:
            raise HTTPException(status_code=400, detail="Duplicate: An employee with this ID/Code already exists.")
    
    # Validate employee_type_id if provided
    if "employee_type_id" in data and data["employee_type_id"] is not None:
         etype = (
            db.query(models.EmployeeType)
            .filter(models.EmployeeType.company_id == company_id, models.EmployeeType.id == data["employee_type_id"])
            .first()
        )
         if not etype:
            raise HTTPException(status_code=400, detail="Invalid employee_type_id")

    previous_designation_id = emp.designation_id
    designation = None
    should_apply_designation_template = False
    if "designation_id" in data and data["designation_id"] is not None:
        designation = (
            db.query(models.PayrollDesignation)
            .filter(models.PayrollDesignation.company_id == company_id, models.PayrollDesignation.id == data["designation_id"])
            .first()
        )
        if designation is None:
            raise HTTPException(status_code=400, detail="Invalid designation_id")
        should_apply_designation_template = int(data["designation_id"]) != int(previous_designation_id or 0)

    for k, v in (data or {}).items():
        setattr(emp, k, v)

    if designation and should_apply_designation_template:
        if "base_monthly_salary" not in data and emp.base_monthly_salary is None and designation.base_monthly_salary is not None:
            emp.base_monthly_salary = designation.base_monthly_salary
        db.flush()
        if designation.template_lines:
            apply_designation_template(
                db,
                company_id=company_id,
                employee_id=employee_id,
                designation=designation,
                effective_from=date.today(),
            )

    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


@router.get("/payheads", response_model=list[schemas.PayrollPayheadRead])
def list_payheads(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.PayrollPayhead)
        .filter(models.PayrollPayhead.company_id == company_id)
        .order_by(models.PayrollPayhead.sort_order.asc(), models.PayrollPayhead.id.asc())
        .all()
    )


@router.post("/payheads", response_model=schemas.PayrollPayheadRead)
def create_payhead(
    company_id: int,
    payload: schemas.PayrollPayheadCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = models.PayrollPayhead(company_id=company_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/device-users", response_model=list[schemas.BiometricDeviceUserRead])
def list_device_users(
    company_id: int,
    device_id: int | None = Query(None),
    employee_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    q = db.query(models.BiometricDeviceUser).filter(models.BiometricDeviceUser.company_id == company_id)
    if device_id is not None:
        q = q.filter(models.BiometricDeviceUser.device_id == int(device_id))
    if employee_id is not None:
        q = q.filter(models.BiometricDeviceUser.employee_id == int(employee_id))
    return q.order_by(models.BiometricDeviceUser.id.asc()).all()


@router.post("/device-users", response_model=schemas.BiometricDeviceUserRead)
def create_device_user(
    company_id: int,
    payload: schemas.BiometricDeviceUserCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    device = (
        db.query(models.BiometricDevice)
        .filter(models.BiometricDevice.company_id == company_id, models.BiometricDevice.id == int(payload.device_id))
        .first()
    )
    if device is None:
        raise HTTPException(status_code=400, detail="Invalid device_id")

    if payload.employee_id is not None:
        emp = (
            db.query(models.Employee)
            .filter(models.Employee.company_id == company_id, models.Employee.id == int(payload.employee_id))
            .first()
        )
        if emp is None:
            raise HTTPException(status_code=400, detail="Invalid employee_id")

    row = models.BiometricDeviceUser(company_id=company_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/device-users/{device_user_id}", response_model=schemas.BiometricDeviceUserRead)
def update_device_user(
    company_id: int,
    device_user_id: int,
    payload: schemas.BiometricDeviceUserUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.BiometricDeviceUser)
        .filter(models.BiometricDeviceUser.company_id == company_id, models.BiometricDeviceUser.id == int(device_user_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Device user not found")

    data = payload.model_dump(exclude_unset=True)
    if "employee_id" in data and data.get("employee_id") is not None:
        emp = (
            db.query(models.Employee)
            .filter(models.Employee.company_id == company_id, models.Employee.id == int(data["employee_id"]))
            .first()
        )
        if emp is None:
            raise HTTPException(status_code=400, detail="Invalid employee_id")

    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/device-users/{device_user_id}")
def delete_device_user(
    company_id: int,
    device_user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.BiometricDeviceUser)
        .filter(models.BiometricDeviceUser.company_id == company_id, models.BiometricDeviceUser.id == int(device_user_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Device user not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.put("/payheads/{payhead_id}", response_model=schemas.PayrollPayheadRead)
def update_payhead(
    company_id: int,
    payhead_id: int,
    payload: schemas.PayrollPayheadUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.PayrollPayhead)
        .filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.id == payhead_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Payhead not found")
    data = payload.model_dump(exclude_unset=True)
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/devices", response_model=list[schemas.BiometricDeviceRead])
def list_devices(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.BiometricDevice)
        .filter(models.BiometricDevice.company_id == company_id)
        .order_by(models.BiometricDevice.id.asc())
        .all()
    )


@router.post("/devices", response_model=schemas.BiometricDeviceRead)
def create_device(
    company_id: int,
    payload: schemas.BiometricDeviceCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = models.BiometricDevice(company_id=company_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/attendance/logs/ingest")
def ingest_attendance_logs(
    company_id: int,
    payload: schemas.AttendanceLogsIngestRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    result = ingest_attendance_logs_service(db, company_id=company_id, payload=payload)
    db.commit()
    return result


@router.post("/attendance/import/csv", response_model=schemas.AttendanceCsvImportResponse)
async def import_attendance_csv(
    company_id: int,
    file: UploadFile = File(...),
    device_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except Exception:
        text = content.decode("utf-8", errors="ignore")

    reader = csv.DictReader(io.StringIO(text))
    logs: list[schemas.AttendanceLogIngestItem] = []

    for row in reader:
        code = (row.get("device_user_code") or row.get("user") or row.get("code") or "").strip()
        ts = (row.get("event_ts") or row.get("timestamp") or row.get("datetime") or "").strip()
        if not code or not ts:
            continue
        try:
            event_ts = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        except Exception:
            continue
        logs.append(
            schemas.AttendanceLogIngestItem(
                device_id=device_id,
                device_user_code=code,
                event_ts=event_ts,
                event_type=(row.get("event_type") or None),
                payload=row,
            )
        )

    payload = schemas.AttendanceLogsIngestRequest(source="CSV", logs=logs)
    result = ingest_attendance_logs_service(db, company_id=company_id, payload=payload)
    db.commit()
    return schemas.AttendanceCsvImportResponse(
        inserted=int(result.get("inserted", 0)),
        deduped=int(result.get("deduped", 0)),
        unmapped_device_users=list(result.get("unmapped_device_users", []) or []),
    )


@router.get("/leave/types", response_model=list[schemas.LeaveTypeRead])
def list_leave_types(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.LeaveType)
        .filter(models.LeaveType.company_id == company_id)
        .order_by(models.LeaveType.id.asc())
        .all()
    )


@router.post("/leave/types", response_model=schemas.LeaveTypeRead)
def create_leave_type(
    company_id: int,
    payload: schemas.LeaveTypeCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = models.LeaveType(company_id=company_id, **payload.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/leave/types/{leave_type_id}", response_model=schemas.LeaveTypeRead)
def update_leave_type(
    company_id: int,
    leave_type_id: int,
    payload: schemas.LeaveTypeCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.LeaveType)
        .filter(models.LeaveType.company_id == company_id, models.LeaveType.id == int(leave_type_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Leave type not found")

    data = payload.model_dump(exclude_unset=True)
    for k, v in (data or {}).items():
        setattr(row, k, v)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/leave/types/{leave_type_id}")
def delete_leave_type(
    company_id: int,
    leave_type_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.LeaveType)
        .filter(models.LeaveType.company_id == company_id, models.LeaveType.id == int(leave_type_id))
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Leave type not found")
    db.delete(row)
    db.commit()
    return {"detail": "Deleted"}


@router.get("/leave/requests", response_model=list[schemas.LeaveRequestRead])
def list_leave_requests(
    company_id: int,
    employee_id: int | None = Query(None),
    status: models.LeaveRequestStatus | None = Query(None),
    start: date | None = Query(None),
    end: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    q = db.query(models.LeaveRequest).filter(models.LeaveRequest.company_id == company_id)
    if employee_id is not None:
        q = q.filter(models.LeaveRequest.employee_id == int(employee_id))
    if status is not None:
        q = q.filter(models.LeaveRequest.status == status)
    if start is not None:
        q = q.filter(models.LeaveRequest.end_date >= start)
    if end is not None:
        q = q.filter(models.LeaveRequest.start_date <= end)
    return q.order_by(models.LeaveRequest.id.desc()).all()


@router.post("/leave/requests", response_model=schemas.LeaveRequestRead)
def create_leave_request(
    company_id: int,
    payload: schemas.LeaveRequestCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="end_date must be >= start_date")

    emp = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.id == int(payload.employee_id))
        .first()
    )
    if emp is None:
        raise HTTPException(status_code=400, detail="Invalid employee_id")

    lt = (
        db.query(models.LeaveType)
        .filter(models.LeaveType.company_id == company_id, models.LeaveType.id == int(payload.leave_type_id))
        .first()
    )
    if lt is None:
        raise HTTPException(status_code=400, detail="Invalid leave_type_id")

    days = float((payload.end_date - payload.start_date).days + 1)
    row = models.LeaveRequest(
        company_id=company_id,
        employee_id=int(payload.employee_id),
        leave_type_id=int(payload.leave_type_id),
        start_date=payload.start_date,
        end_date=payload.end_date,
        days=days,
        status=models.LeaveRequestStatus.PENDING,
        reason=payload.reason,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/leave/requests/{leave_request_id}/approve", response_model=schemas.LeaveRequestRead)
def approve_leave_request(
    company_id: int,
    leave_request_id: int,
    payload: schemas.LeaveRequestDecision,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.LeaveRequest)
        .filter(models.LeaveRequest.company_id == company_id, models.LeaveRequest.id == int(leave_request_id))
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if row.status != models.LeaveRequestStatus.PENDING:
        raise HTTPException(status_code=409, detail="Leave request is not pending")

    row.status = models.LeaveRequestStatus.APPROVED
    row.approved_by_user_id = int(current_user.id)
    row.approved_at = datetime.utcnow()
    if payload.reason is not None:
        row.reason = payload.reason
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/leave/requests/{leave_request_id}/reject", response_model=schemas.LeaveRequestRead)
def reject_leave_request(
    company_id: int,
    leave_request_id: int,
    payload: schemas.LeaveRequestDecision,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.LeaveRequest)
        .filter(models.LeaveRequest.company_id == company_id, models.LeaveRequest.id == int(leave_request_id))
        .with_for_update()
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Leave request not found")
    if row.status != models.LeaveRequestStatus.PENDING:
        raise HTTPException(status_code=409, detail="Leave request is not pending")

    row.status = models.LeaveRequestStatus.REJECTED
    row.approved_by_user_id = int(current_user.id)
    row.approved_at = datetime.utcnow()
    if payload.reason is not None:
        row.reason = payload.reason
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/attendance/daily", response_model=list[schemas.AttendanceDailyRead])
def list_attendance_daily(
    company_id: int,
    start: date = Query(...),
    end: date = Query(...),
    employee_id: int | None = Query(None),
    status: models.AttendanceStatus | None = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    if end < start:
        raise HTTPException(status_code=400, detail="end must be >= start")

    q = (
        db.query(models.AttendanceDaily)
        .filter(
            models.AttendanceDaily.company_id == company_id,
            models.AttendanceDaily.work_date >= start,
            models.AttendanceDaily.work_date <= end,
        )
        .order_by(models.AttendanceDaily.work_date.asc(), models.AttendanceDaily.employee_id.asc())
    )
    if employee_id is not None:
        q = q.filter(models.AttendanceDaily.employee_id == int(employee_id))
    if status is not None:
        q = q.filter(models.AttendanceDaily.status == status)

    return q.all()


@router.put("/attendance/daily/{employee_id}/{work_date}", response_model=schemas.AttendanceDailyRead)
def manual_fix_attendance_daily(
    company_id: int,
    employee_id: int,
    work_date: date,
    payload: schemas.AttendanceDailyManualUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    row = (
        db.query(models.AttendanceDaily)
        .filter(
            models.AttendanceDaily.company_id == company_id,
            models.AttendanceDaily.employee_id == int(employee_id),
            models.AttendanceDaily.work_date == work_date,
        )
        .with_for_update()
        .first()
    )
    if row is None:
        row = models.AttendanceDaily(company_id=company_id, employee_id=int(employee_id), work_date=work_date)
        db.add(row)
        db.flush()

    if payload.first_in is not None:
        row.first_in = payload.first_in
    if payload.last_out is not None:
        row.last_out = payload.last_out

    if row.first_in is not None and row.last_out is not None and row.last_out > row.first_in:
        diff = row.last_out - row.first_in
        row.worked_minutes = max(0, int(diff.total_seconds() // 60))
    else:
        row.worked_minutes = int(getattr(row, "worked_minutes", 0) or 0)

    if payload.status is not None:
        row.status = payload.status
    row.is_manual = True
    row.manual_reason = str(payload.manual_reason or "").strip() or None

    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.post("/attendance/daily/recompute")
def recompute_attendance(
    company_id: int,
    start: date = Query(...),
    end: date = Query(...),
    employee_ids: list[int] | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    if end < start:
        raise HTTPException(status_code=400, detail="end must be >= start")

    updated = 0
    cur = start
    while cur <= end:
        updated += upsert_attendance_from_raw_logs(db, company_id=company_id, work_date=cur, employee_ids=employee_ids)
        cur = cur + timedelta(days=1)

    db.commit()
    return {"updated": int(updated)}


@router.get("/runs", response_model=list[schemas.PayrollRunRead])
def list_payroll_runs(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    return (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id)
        .order_by(models.PayrollRun.period_year.desc(), models.PayrollRun.period_month.desc())
        .all()
    )


@router.post("/runs", response_model=schemas.PayrollRunRead)
def create_payroll_run(
    company_id: int,
    payload: schemas.PayrollRunCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    year = int(payload.period_year)
    month = int(payload.period_month)
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="period_month must be 1..12")

    period_start = date(year, month, 1)
    if month == 12:
        period_end = date(year, 12, 31)
    else:
        period_end = date(year, month + 1, 1).fromordinal(date(year, month + 1, 1).toordinal() - 1)

    existing = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.period_year == year, models.PayrollRun.period_month == month)
        .first()
    )
    if existing is not None:
        return existing

    run = models.PayrollRun(
        company_id=company_id,
        period_year=year,
        period_month=month,
        period_start=period_start,
        period_end=period_end,
        status=models.PayrollRunStatus.DRAFT,
        locked=False,
        created_by_user_id=int(current_user.id),
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    return run


@router.post("/runs/{run_id}/compute", response_model=schemas.PayrollRunComputeResponse)
def compute_run(
    company_id: int,
    run_id: int,
    payload: schemas.PayrollRunComputeRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == run_id)
        .first()
    )
    if not run:
        raise HTTPException(status_code=404, detail="Payroll run not found")

    summary = compute_payroll_run(
        db,
        company_id=company_id,
        run=run,
        employee_ids=payload.employee_ids,
        recompute_attendance=bool(payload.recompute_attendance),
    )
    db.commit()
    db.refresh(run)
    return schemas.PayrollRunComputeResponse(
        run_id=int(run.id),
        status=run.status,
        employees_processed=int(summary.employees_processed),
    )


@router.get("/runs/{run_id}/payslips", response_model=list[schemas.PayslipRead])
def list_payslips(
    company_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    slips = (
        db.query(models.PayrollPayslip)
        .filter(models.PayrollPayslip.company_id == company_id, models.PayrollPayslip.payroll_run_id == run_id)
        .order_by(models.PayrollPayslip.employee_id.asc())
        .all()
    )
    out: list[schemas.PayslipRead] = []
    for s in slips:
        out.append(_build_payslip_read(db, company_id=company_id, slip=s))
    return out


@router.post("/runs/{run_id}/unlock", response_model=schemas.PayrollRunRead)
def unlock_payroll_run(
    company_id: int,
    run_id: int,
    payload: schemas.PayrollRunUnlockRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    _require_admin_or_superadmin(current_user)

    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == int(run_id))
        .with_for_update()
        .first()
    )
    if run is None:
        raise HTTPException(status_code=404, detail="Payroll run not found")

    run.locked = False
    db.add(run)

    log_event(
        db,
        user_id=int(current_user.id),
        tenant_id=int(company.tenant_id),
        action="payroll_run_unlock",
        message=f"Unlocked payroll run {int(run.id)} for {int(run.period_year):04d}-{int(run.period_month):02d}. Reason: {str(payload.reason or '').strip()}" ,
    )

    db.commit()
    db.refresh(run)
    return run


@router.post(
    "/runs/{run_id}/payslips/{employee_id}/override",
    response_model=schemas.PayslipRead,
)
def override_payslip(
    company_id: int,
    run_id: int,
    employee_id: int,
    payload: schemas.PayrollPayslipOverrideRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)

    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == int(run_id))
        .with_for_update()
        .first()
    )
    if run is None:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if bool(run.locked):
        raise HTTPException(status_code=409, detail="Payroll run is locked")

    slip = (
        db.query(models.PayrollPayslip)
        .filter(
            models.PayrollPayslip.company_id == company_id,
            models.PayrollPayslip.payroll_run_id == int(run_id),
            models.PayrollPayslip.employee_id == int(employee_id),
        )
        .with_for_update()
        .first()
    )
    if slip is None:
        raise HTTPException(status_code=404, detail="Payslip not found")

    old_snapshot = _build_payslip_read(db, company_id=company_id, slip=slip).model_dump()

    if payload.payable_days is not None:
        slip.payable_days = float(payload.payable_days)
    if payload.absent_days is not None:
        slip.absent_days = float(payload.absent_days)
    if payload.late_minutes is not None:
        slip.late_minutes = int(payload.late_minutes)
    if payload.overtime_minutes is not None:
        slip.overtime_minutes = int(payload.overtime_minutes)

    # Replace payslip lines if provided
    if payload.lines is not None and len(payload.lines) > 0:
        db.query(models.PayrollPayslipLine).filter(models.PayrollPayslipLine.payslip_id == int(slip.id)).delete()

        earnings_total = 0.0
        deductions_total = 0.0
        for l in payload.lines:
            ph = (
                db.query(models.PayrollPayhead)
                .filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.id == int(l.payhead_id))
                .first()
            )
            if ph is None:
                raise HTTPException(status_code=400, detail="Invalid payhead_id")

            amt = float(l.amount or 0)
            if ph.type == models.PayrollPayheadType.EARNING:
                earnings_total += amt
            else:
                deductions_total += amt

            db.add(
                models.PayrollPayslipLine(
                    company_id=company_id,
                    payslip_id=int(slip.id),
                    payhead_id=int(ph.id),
                    type=ph.type,
                    amount=amt,
                )
            )

        slip.earnings_total = float(round(earnings_total, 2))
        slip.deductions_total = float(round(deductions_total, 2))
        slip.net_pay = float(round(earnings_total - deductions_total, 2))

    slip.is_manual_override = True
    slip.override_reason = str(payload.override_reason or "").strip()
    db.add(slip)

    new_snapshot = _build_payslip_read(db, company_id=company_id, slip=slip).model_dump()

    db.add(
        models.PayrollOverrideLog(
            company_id=company_id,
            payroll_run_id=int(run.id),
            payslip_id=int(slip.id),
            actor_user_id=int(current_user.id),
            reason=str(payload.override_reason or "").strip(),
            diff_json={"before": old_snapshot, "after": new_snapshot},
        )
    )

    db.commit()
    db.refresh(slip)
    return _build_payslip_read(db, company_id=company_id, slip=slip)


@router.get(
    "/runs/{run_id}/payslips/{employee_id}/export",
    response_model=schemas.PayslipExportResponse,
)
def export_payslip_json(
    company_id: int,
    run_id: int,
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == int(run_id))
        .first()
    )
    if run is None:
        raise HTTPException(status_code=404, detail="Payroll run not found")

    emp = (
        db.query(models.Employee)
        .filter(models.Employee.company_id == company_id, models.Employee.id == int(employee_id))
        .first()
    )
    if emp is None:
        raise HTTPException(status_code=404, detail="Employee not found")

    slip = (
        db.query(models.PayrollPayslip)
        .filter(
            models.PayrollPayslip.company_id == company_id,
            models.PayrollPayslip.payroll_run_id == int(run_id),
            models.PayrollPayslip.employee_id == int(employee_id),
        )
        .first()
    )
    if slip is None:
        raise HTTPException(status_code=404, detail="Payslip not found")

    return schemas.PayslipExportResponse(
        company_id=int(company_id),
        payroll_run=schemas.PayrollRunRead.model_validate(run),
        employee=schemas.EmployeeRead.model_validate(emp),
        payslip=_build_payslip_read(db, company_id=company_id, slip=slip),
    )


@router.post("/runs/{run_id}/approve", response_model=schemas.PayrollRunRead)
def approve_run(
    company_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == run_id)
        .with_for_update()
        .first()
    )
    if not run:
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if run.status not in (models.PayrollRunStatus.COMPUTED, models.PayrollRunStatus.DRAFT):
        raise HTTPException(status_code=409, detail="Payroll run cannot be approved in current state")

    run.status = models.PayrollRunStatus.APPROVED
    run.locked = True
    run.approved_at = datetime.utcnow()
    run.approved_by_user_id = int(current_user.id)
    db.add(run)
    db.commit()
    db.refresh(run)
    return run


@router.post("/runs/{run_id}/post-voucher")
def post_voucher(
    company_id: int,
    run_id: int,
    req: schemas.PayrollRunPostRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = (
        db.query(models.PayrollRun)
        .filter(models.PayrollRun.company_id == company_id, models.PayrollRun.id == run_id)
        .with_for_update()
        .first()
    )
    if not run:
        raise HTTPException(status_code=404, detail="Payroll run not found")

    from .vouchers import _create_voucher_impl

    voucher_payload = build_payroll_voucher_payload(db, company_id=company_id, run=run, post_date=req.post_date)
    created = _create_voucher_impl(company_id, voucher_payload, db, current_user)

    run.voucher_id = int(created.id)
    run.status = models.PayrollRunStatus.POSTED
    run.locked = True
    run.posted_at = datetime.utcnow()
    db.add(run)
    db.commit()

    return {
        "run_id": int(run.id),
        "status": str(run.status.value),
        "locked": bool(run.locked),
        "voucher_id": int(created.id),
        "voucher_number": created.voucher_number,
    }
@router.get("/runs/{run_id}/salary-template-excel")
def download_salary_template_excel(
    company_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    import openpyxl
    _get_company(db, company_id, current_user)
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == run_id, models.PayrollRun.company_id == company_id).first()
    if not run: raise HTTPException(404, "Run not found")
    
    employees = db.query(models.Employee).filter(models.Employee.company_id == company_id, models.Employee.is_active == True).all()
    payheads = db.query(models.PayrollPayhead).filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.is_active == True).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary Template"
    
    headers = ["Employee ID", "Employee Name", "Department", "Project", "Segment", "Payable Days", "TDS Amount"]
    for ph in payheads:
        headers.append(f"[{ph.id}] {ph.name}")
        
    ws.append(headers)
    
    payslips = db.query(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    payslip_by_emp = {p.employee_id: p for p in payslips}
    
    payslip_lines = db.query(models.PayrollPayslipLine).join(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    lines_by_slip = {}
    for l in payslip_lines:
        lines_by_slip.setdefault(l.payslip_id, {})[l.payhead_id] = float(l.amount or 0)
    payslips = db.query(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    payslip_by_emp = {p.employee_id: p for p in payslips}
    
    payslip_lines = db.query(models.PayrollPayslipLine).join(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    lines_by_slip = {}
    for l in payslip_lines:
        lines_by_slip.setdefault(l.payslip_id, {})[l.payhead_id] = float(l.amount or 0)

    for emp in employees:
        slip = payslip_by_emp.get(emp.id)
        if slip:
            dept = emp.department.name if emp.department else ""
            proj = emp.project.name if emp.project else ""
            seg = emp.segment.name if emp.segment else ""
            row = [emp.id, emp.full_name, dept, proj, seg, float(slip.payable_days or 0), float(getattr(slip, "tds_amount", 0) or 0)]
            slines = lines_by_slip.get(slip.id, {})
            for ph in payheads:
                row.append(slines.get(ph.id, 0.0))
        else:
            dept = emp.department.name if emp.department else ""
            proj = emp.project.name if emp.project else ""
            seg = emp.segment.name if emp.segment else ""
            row = [emp.id, emp.full_name, dept, proj, seg, 30, 0]
            for ph in payheads:
                row.append(0.0)
        ws.append(row)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_template_{run.period_year}_{run.period_month}.xlsx"}
    )

@router.post("/runs/{run_id}/upload-salary-excel")
def upload_salary_excel(
    company_id: int,
    run_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    import openpyxl
    _get_company(db, company_id, current_user)
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == run_id, models.PayrollRun.company_id == company_id).first()
    if not run: raise HTTPException(404, "Run not found")
    if run.locked: raise HTTPException(400, "Run is locked")
    
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    if not headers or headers[0] != "Employee ID":
        raise HTTPException(400, "Invalid template format")
        
    # Get payheads
    payheads = {ph.id: ph for ph in db.query(models.PayrollPayhead).filter(models.PayrollPayhead.company_id == company_id).all()}
    
    payhead_cols = {}
    for idx, h in enumerate(headers):
        h_str = str(h or "").strip()
        if h_str.startswith("[") and "]" in h_str:
            try:
                pid = int(h_str.split("]")[0].strip("["))
                if pid in payheads:
                    payhead_cols[idx] = pid
            except:
                pass
                
    payable_idx = headers.index("Payable Days") if "Payable Days" in headers else 2
    tds_idx = headers.index("TDS Amount") if "TDS Amount" in headers else 3
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        try:
            emp_id = int(row[0])
            payable_days = float(row[payable_idx] if len(row) > payable_idx and row[payable_idx] is not None else 0)
            tds_amount = float(row[tds_idx] if len(row) > tds_idx and row[tds_idx] is not None else 0)
        except (ValueError, TypeError):
            continue
            
        slip = db.query(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id, models.PayrollPayslip.employee_id == emp_id).with_for_update().first()
        
        if slip is None:
            slip = models.PayrollPayslip(
                company_id=company_id,
                payroll_run_id=run_id,
                employee_id=emp_id,
            )
            db.add(slip)
            db.flush()
        else:
            db.query(models.PayrollPayslipLine).filter(models.PayrollPayslipLine.payslip_id == int(slip.id)).delete()
            
        slip.payable_days = payable_days
        slip.absent_days = 0
        slip.late_minutes = 0
        slip.overtime_minutes = 0
        slip.is_manual_override = True
        slip.override_reason = "Uploaded via Excel"
        
        earnings_total = 0.0
        deductions_total = tds_amount
        
        for idx, pid in payhead_cols.items():
            if idx < len(row):
                val = row[idx]
                if val is not None:
                    try:
                        amt = float(val)
                        if amt <= 0: continue
                        ph = payheads[pid]
                        if ph.type == models.PayrollPayheadType.EARNING:
                            earnings_total += amt
                        else:
                            deductions_total += amt
                            
                        db.add(models.PayrollPayslipLine(
                            company_id=company_id,
                            payslip_id=int(slip.id),
                            payhead_id=pid,
                            type=ph.type,
                            amount=amt
                        ))
                    except (ValueError, TypeError):
                        pass
                        
        slip.earnings_total = float(round(earnings_total, 2))
        slip.deductions_total = float(round(deductions_total, 2))
        slip.tds_amount = float(round(tds_amount, 2))
        slip.net_pay = float(round(earnings_total - deductions_total, 2))
        db.add(slip)
        
    run.status = models.PayrollRunStatus.COMPUTED
    run.computed_at = datetime.utcnow()
    db.commit()
    return {"detail": "Salary uploaded successfully"}


@router.get("/runs/{run_id}/salary-sheet-data")
def get_salary_sheet_data(
    company_id: int,
    run_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == run_id, models.PayrollRun.company_id == company_id).first()
    if not run: raise HTTPException(404, "Run not found")
    
    employees = db.query(models.Employee).filter(models.Employee.company_id == company_id, models.Employee.is_active == True).all()
    payheads = db.query(models.PayrollPayhead).filter(models.PayrollPayhead.company_id == company_id, models.PayrollPayhead.is_active == True).all()
    
    headers = ["Employee ID", "Employee Name", "Grade No.", "Designation", "Department", "Project", "Segment", "Payable Days", "TDS Amount"]
    for ph in payheads:
        headers.append(f"[{ph.id}] {ph.name}")
    headers.append("Net Amount")

    payslips = db.query(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    payslip_by_emp = {p.employee_id: p for p in payslips}

    payslip_lines = db.query(models.PayrollPayslipLine).join(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id).all()
    lines_by_slip = {}
    for l in payslip_lines:
        lines_by_slip.setdefault(l.payslip_id, {})[l.payhead_id] = float(l.amount or 0)

    rows = []
    for emp in employees:
        dept = emp.department.name if emp.department else ""
        proj = emp.project.name if emp.project else ""
        seg = emp.segment.name if emp.segment else ""

        grade_number = getattr(emp, "grade_number", None)
        designation = emp.designation.name if emp.designation else ""
        slip = payslip_by_emp.get(emp.id)
        if slip:
            row = [emp.id, emp.full_name, grade_number, designation, dept, proj, seg, float(slip.payable_days or 0), float(getattr(slip, "tds_amount", 0) or 0)]
            slines = lines_by_slip.get(slip.id, {})
            for ph in payheads:
                row.append(slines.get(ph.id, 0.0))
            row.append(float(slip.net_pay or 0))
        else:
            row = [emp.id, emp.full_name, grade_number, designation, dept, proj, seg, 30, 0]
            for ph in payheads:
                row.append(0.0)
            row.append(0.0)
        rows.append(row)

    return {"headers": headers, "rows": rows}

@router.post("/runs/{run_id}/upload-salary-json")
def upload_salary_json(
    company_id: int,
    run_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    run = db.query(models.PayrollRun).filter(models.PayrollRun.id == run_id, models.PayrollRun.company_id == company_id).first()
    if not run: raise HTTPException(404, "Run not found")
    if run.locked: raise HTTPException(400, "Run is locked")
    
    headers = payload.get("headers", [])
    rows = payload.get("rows", [])
    if not headers or headers[0] != "Employee ID":
        raise HTTPException(400, "Invalid template format")
        
    payheads = {ph.id: ph for ph in db.query(models.PayrollPayhead).filter(models.PayrollPayhead.company_id == company_id).all()}
    
    payhead_cols = {}
    for idx, h in enumerate(headers):
        h_str = str(h or "").strip()
        if h_str.startswith("[") and "]" in h_str:
            try:
                pid = int(h_str.split("]")[0].strip("["))
                if pid in payheads:
                    payhead_cols[idx] = pid
            except:
                pass
                
    payable_idx = headers.index("Payable Days") if "Payable Days" in headers else 2
    tds_idx = headers.index("TDS Amount") if "TDS Amount" in headers else 3
    
    for row in rows:
        if not row or not row[0]: continue
        try:
            emp_id = int(row[0])
            payable_days = float(row[payable_idx] if len(row) > payable_idx and row[payable_idx] is not None else 0)
            tds_amount = float(row[tds_idx] if len(row) > tds_idx and row[tds_idx] is not None else 0)
        except (ValueError, TypeError):
            continue
            
        slip = db.query(models.PayrollPayslip).filter(models.PayrollPayslip.payroll_run_id == run_id, models.PayrollPayslip.employee_id == emp_id).with_for_update().first()
        
        if slip is None:
            slip = models.PayrollPayslip(
                company_id=company_id,
                payroll_run_id=run_id,
                employee_id=emp_id,
            )
            db.add(slip)
            db.flush()
        else:
            db.query(models.PayrollPayslipLine).filter(models.PayrollPayslipLine.payslip_id == int(slip.id)).delete()
            
        slip.payable_days = payable_days
        slip.absent_days = 0
        slip.late_minutes = 0
        slip.overtime_minutes = 0
        slip.is_manual_override = True
        slip.override_reason = "Updated via Preview Grid"
        
        earnings_total = 0.0
        deductions_total = tds_amount
        
        for idx, pid in payhead_cols.items():
            if idx < len(row):
                val = row[idx]
                if val is not None and val != "":
                    try:
                        amt = float(val)
                        if amt <= 0: continue
                        ph = payheads[pid]
                        if ph.type == models.PayrollPayheadType.EARNING:
                            earnings_total += amt
                        else:
                            deductions_total += amt
                            
                        db.add(models.PayrollPayslipLine(
                            company_id=company_id,
                            payslip_id=int(slip.id),
                            payhead_id=pid,
                            type=ph.type,
                            amount=amt
                        ))
                    except (ValueError, TypeError):
                        pass
                        
        slip.earnings_total = float(round(earnings_total, 2))
        slip.deductions_total = float(round(deductions_total, 2))
        slip.tds_amount = float(round(tds_amount, 2))
        slip.net_pay = float(round(earnings_total - deductions_total, 2))
        db.add(slip)
        
    run.status = models.PayrollRunStatus.COMPUTED
    run.computed_at = datetime.utcnow()
    db.commit()
    return {"detail": "Salary data updated successfully"}



@router.get("/reports/salary-sheet")
def get_salary_sheet_report(
    company_id: int,
    year: int = Query(...),
    month: int = Query(None),
    employee_id: int = Query(None),
    department_id: int = Query(None),
    project_id: int = Query(None),
    segment_id: int = Query(None),
    calendar_mode: str = Query("AD"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    company = _get_company(db, company_id, current_user)
    company_mode = getattr(company, 'calendar_mode', 'AD')
    
    # We need to decide what the stored year/month represent.
    # In this system, they usually follow company_mode.
    
    target_year = year
    target_month = month
    
    # If the user is querying in a different mode than the company default,
    # we may need to translate the filter.
    # For now, let's assume the user wants to see data that MATCHES the selected period.
    # If they are in BS mode and select 2081-01, they expect to see runs from 2081-01 (BS).
    # If the database stores runs in AD, we must translate.
    
    # Translation logic for monthly filters
    if month is not None and calendar_mode != company_mode:
        from ..nepali_date import bs_to_ad_date, ad_to_bs_str
        from datetime import date
        try:
            if calendar_mode == "BS" and company_mode == "AD":
                # Convert BS year/month to AD year/month
                ad_date = bs_to_ad_date(f"{year}-{month:02d}-15")
                target_year = ad_date.year
                target_month = ad_date.month
            elif calendar_mode == "AD" and company_mode == "BS":
                # Convert AD year/month to BS year/month
                bs_str = ad_to_bs_str(date(year, month, 15))
                y, m, d = map(int, bs_str.split("-"))
                target_year = y
                target_month = m
        except:
            pass

    query = db.query(models.PayrollPayslip).join(models.PayrollRun).filter(
        models.PayrollPayslip.company_id == company_id
    )
    
    
    if month is None:
        # For year compile, if calendars differ, we need a range
        if calendar_mode != company_mode:
            if calendar_mode == "BS" and company_mode == "AD":
                # BS year Y roughly spans AD year Y-57 (Apr-Dec) and Y-56 (Jan-Apr)
                ad_y1 = year - 57
                ad_y2 = year - 56
                # We can be more precise: 
                # BS 2081-01-01 is 2024-04-13
                # BS 2081-12-31 is 2025-04-13
                query = query.filter(
                    ((models.PayrollRun.period_year == ad_y1) & (models.PayrollRun.period_month >= 4)) |
                    ((models.PayrollRun.period_year == ad_y2) & (models.PayrollRun.period_month <= 4))
                )
            elif calendar_mode == "AD" and company_mode == "BS":
                # AD year Y spans BS year Y+56 (Poush-Chaitra) and Y+57 (Baisakh-Mangsir)
                bs_y1 = year + 56
                bs_y2 = year + 57
                query = query.filter(
                    ((models.PayrollRun.period_year == bs_y1) & (models.PayrollRun.period_month >= 9)) |
                    ((models.PayrollRun.period_year == bs_y2) & (models.PayrollRun.period_month <= 9))
                )
            else:
                query = query.filter(models.PayrollRun.period_year == year)
        else:
            query = query.filter(models.PayrollRun.period_year == year)
    else:

        query = query.filter(
            models.PayrollRun.period_year == target_year,
            models.PayrollRun.period_month == target_month
        )
    
    if employee_id:
        query = query.filter(models.PayrollPayslip.employee_id == employee_id)
        
    payslips = query.all()
    
    if department_id or project_id or segment_id:
        filtered = []
        for p in payslips:
            emp = p.employee
            if department_id and emp.department_id != department_id: continue
            if project_id and emp.project_id != project_id: continue
            if segment_id and emp.segment_id != segment_id: continue
            filtered.append(p)
        payslips = filtered

    payhead_ids = set()
    for p in payslips:
        for line in p.lines:
            payhead_ids.add(line.payhead_id)
            
    payheads = db.query(models.PayrollPayhead).filter(models.PayrollPayhead.id.in_(payhead_ids)).all() if payhead_ids else []
    payheads.sort(key=lambda x: x.sort_order)
    
    
    report_rows = []
    if month is None:
        # Aggregate by employee for yearly report
        agg = {}
        for p in payslips:
            eid = p.employee_id
            if eid not in agg:
                agg[eid] = {
                    "employee_id": eid,
                    "employee_name": p.employee.full_name,
                    "employee_code": p.employee.code,
                    "designation": p.employee.designation.name if p.employee.designation else "",
                    "grade": getattr(p.employee, "grade", "") or "",
                    "grade_number": getattr(p.employee, "grade_number", None),
                    "grade_rate": float(p.employee.designation.grade_rate) if p.employee.designation and getattr(p.employee.designation, "grade_rate", None) is not None else None,

                    "department": p.employee.department.name if p.employee.department else "",
                    "project": p.employee.project.name if p.employee.project else "",
                    "segment": p.employee.segment.name if p.employee.segment else "",
                    "payable_days": 0.0,
                    "earnings_total": 0.0,
                    "deductions_total": 0.0,
                    "tds_amount": 0.0,
                    "net_pay": 0.0,
                }
            
            agg[eid]["payable_days"] += float(p.payable_days)
            agg[eid]["earnings_total"] += float(p.earnings_total)
            agg[eid]["deductions_total"] += float(p.deductions_total)
            agg[eid]["tds_amount"] += float(p.tds_amount)
            agg[eid]["net_pay"] += float(p.net_pay)
            
            for line in p.lines:
                key = f"ph_{line.payhead_id}"
                agg[eid][key] = agg[eid].get(key, 0.0) + float(line.amount)
        report_rows = list(agg.values())
    else:
        # List each payslip for monthly report
        for p in payslips:
            row = {
                "employee_id": p.employee_id,
                "employee_name": p.employee.full_name,
                "employee_code": p.employee.code,
                "designation": p.employee.designation.name if p.employee.designation else "",
                "grade": getattr(p.employee, "grade", "") or "",
                "grade_number": getattr(p.employee, "grade_number", None),
                "grade_rate": float(p.employee.designation.grade_rate) if p.employee.designation and getattr(p.employee.designation, "grade_rate", None) is not None else None,

                "department": p.employee.department.name if p.employee.department else "",
                "project": p.employee.project.name if p.employee.project else "",
                "segment": p.employee.segment.name if p.employee.segment else "",
                "month": p.run.period_month,
                "year": p.run.period_year,
                "payable_days": float(p.payable_days),
                "earnings_total": float(p.earnings_total),
                "deductions_total": float(p.deductions_total),
                "tds_amount": float(p.tds_amount),
                "net_pay": float(p.net_pay),
            }
            for line in p.lines:
                row[f"ph_{line.payhead_id}"] = float(line.amount)
            report_rows.append(row)

            
    return {
        "payheads": [
            {
                "id": ph.id,
                "name": ph.name,
                "type": ph.type.value if hasattr(ph.type, "value") else str(ph.type),
                "cost_center_option": getattr(ph, "cost_center_option", None),
            }
            for ph in payheads
        ],
        "rows": report_rows,
    }


@router.get("/employees/export-template")
def export_employee_template(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = [
        "Full Name", "Code", "Gender", "Marital Status", "Email", "Phone", "DOB (YYYY-MM-DD)", "PAN",
        "Join Date (YYYY-MM-DD)", "Grade Label", "Grade Number",
        "Employee Type Name", "Designation Name", "Department Name", 
        "Project Name", "Segment Name",
        "Base Monthly Salary", "Apply TDS (TRUE/FALSE)", "TDS Percent"
    ]
    ws.append(headers)
    
    # Add sample row
    ws.append([
        "John Doe", "EMP001", "Male", "Married", "john@example.com", "9841234567", "1990-01-01", "123456789",
        "2023-01-01", "A1", "3", "Permanent", "Accountant", "Finance", 
        "", "", "25000", "TRUE", "1.0"
    ])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"}
    )


@router.post("/employees/import-excel")
async def import_employees_excel(
    company_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _get_company(db, company_id, current_user)
    from openpyxl import load_workbook
    
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")
    
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    data_rows = rows[1:]
    
    # Pre-fetch lookups
    employee_types = {t.name.lower(): t.id for t in db.query(models.EmployeeType).filter(models.EmployeeType.company_id == company_id).all()}
    designations = {d.name.lower(): d for d in db.query(models.PayrollDesignation).filter(models.PayrollDesignation.company_id == company_id).all()}
    departments = {d.name.lower(): d.id for d in db.query(models.Department).filter(models.Department.company_id == company_id).all()}
    projects = {p.name.lower(): p.id for p in db.query(models.Project).filter(models.Project.company_id == company_id).all()}
    segments = {s.name.lower(): s.id for s in db.query(models.Segment).filter(models.Segment.company_id == company_id).all()}

    # Resolve "Payable Ledger"
    payable_ledger = db.query(models.Ledger).filter(
        models.Ledger.company_id == company_id,
        models.Ledger.name.ilike("%salary%payable%")
    ).first()
    payable_ledger_id = payable_ledger.id if payable_ledger else None

    created_count = 0
    errors = []

    for idx, row_data in enumerate(data_rows, start=2):
        row = dict(zip(headers, row_data))
        if not row.get("full name"):
            continue
            
        try:
            # Basic info
            full_name = str(row.get("full name")).strip()
            code = str(row.get("code")).strip() if row.get("code") else None
            
            # Uniqueness check for code
            if code:
                existing_in_db = (
                    db.query(models.Employee)
                    .filter(models.Employee.company_id == company_id, models.Employee.code == code)
                    .first()
                )
                if existing_in_db:
                    errors.append(f"Row {idx}: Duplicate ID/Code '{code}' already exists.")
                    continue

            gender = str(row.get("gender")).strip() if row.get("gender") else None
            marital_status = str(row.get("marital status")).strip() if row.get("marital status") else None
            email = str(row.get("email")).strip() if row.get("email") else None
            phone = str(row.get("phone")).strip() if row.get("phone") else None
            pan = str(row.get("pan")).strip() if row.get("pan") else None
            
            # Dates
            dob = None
            dob_val = row.get("dob (yyyy-mm-dd)")
            if dob_val:
                if isinstance(dob_val, (date, datetime)):
                    dob = dob_val.date() if isinstance(dob_val, datetime) else dob_val
                else:
                    try:
                        dob = datetime.strptime(str(dob_val).strip(), "%Y-%m-%d").date()
                    except:
                        pass

            join_date = None
            jd_val = row.get("join date (yyyy-mm-dd)")
            if jd_val:
                if isinstance(jd_val, (date, datetime)):
                    join_date = jd_val.date() if isinstance(jd_val, datetime) else jd_val
                else:
                    try:
                        join_date = datetime.strptime(str(jd_val).strip(), "%Y-%m-%d").date()
                    except:
                        pass
            
            grade = str(row.get("grade label")).strip() if row.get("grade label") else None
            grade_number = None
            try:
                gn_val = row.get("grade number")
                grade_number = int(float(gn_val)) if gn_val is not None else None
            except:
                pass

            # Lookups
            etype_name = str(row.get("employee type name") or "").lower()
            etype_id = employee_types.get(etype_name)
            
            desg_name = str(row.get("designation name") or "").lower()
            desg_obj = designations.get(desg_name)
            
            dept_name = str(row.get("department name") or "").lower()
            dept_id = departments.get(dept_name)
            
            proj_name = str(row.get("project name") or "").lower()
            proj_id = projects.get(proj_name)
            
            seg_name = str(row.get("segment name") or "").lower()
            seg_id = segments.get(seg_name)

            salary = None
            try:
                sal_val = row.get("base monthly salary")
                salary = float(sal_val) if sal_val is not None else None
            except:
                pass
            
            apply_tds_val = str(row.get("apply tds (true/false)") or "").lower()
            apply_tds = apply_tds_val == "true"
            
            tds_percent = 1.0
            try:
                tp_val = row.get("tds percent")
                tds_percent = float(tp_val) if tp_val is not None else 1.0
            except:
                pass

            emp = models.Employee(
                company_id=company_id,
                full_name=full_name,
                code=code,
                gender=gender,
                marital_status=marital_status,
                email=email,
                phone=phone,
                dob=dob,
                pan=pan,
                join_date=join_date,
                grade=grade,
                grade_number=grade_number,
                employee_type_id=etype_id,
                designation_id=desg_obj.id if desg_obj else None,
                department_id=dept_id,
                project_id=proj_id,
                segment_id=seg_id,
                base_monthly_salary=salary if salary is not None else (desg_obj.base_monthly_salary if desg_obj else None),
                apply_tds=apply_tds,
                tds_percent=tds_percent,
                payable_ledger_id=payable_ledger_id,
                payroll_mode=models.PayrollMode.MONTHLY,
                salary_mode=models.SalaryMode.PRO_RATA,
                is_active=True
            )
            db.add(emp)
            db.flush()

            # Apply template if designation exists
            if desg_obj:
                apply_designation_template(
                    db,
                    company_id=company_id,
                    employee_id=int(emp.id),
                    designation=desg_obj,
                    effective_from=join_date or date.today(),
                )
            
            created_count += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    if errors and created_count == 0:
        db.rollback()
        raise HTTPException(status_code=400, detail="\n".join(errors))
    
    db.commit()
    return {"detail": f"Imported {created_count} employees", "errors": errors}


